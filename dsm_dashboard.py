# dsm_dashboard.py
# Modern DSM Dashboard with Solar Analytics style

from __future__ import annotations

import json
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
import plotly.express as px

import dash
from dash import Dash, html, dcc, dash_table, Input, Output, State, ctx, ALL, MATCH, callback_context
from dash.exceptions import PreventUpdate
from dash import no_update
import dash_bootstrap_components as dbc

# =========================
# Models & light utils (drop-in)
# =========================
from dataclasses import dataclass
from typing import List, Tuple, Dict, Any
from io import BytesIO
import math

from core.dsm_engine import (
    Band,
    RATE_FLAT,
    RATE_FRAC,
    RATE_MULT,
    RATE_SCALED,
    MODE_DEFAULT,
    MODE_DYNAMIC,
    safe_mode,
    denominator_and_basis,
    direction_from,
    slice_pct,
    kwh_from_slice,
    band_rate,
    compute_error_pct,
    compute_basis_mw,
    _normalize_bands_df,
    parse_bands_from_settings,
    compute_slot_row,
    apply_bands,
    summarize,
    generate_label,
)

# =========================
# ---------- THEME --------
# =========================
APP_TITLE = "O2: DSM Analytics"
THEME = dbc.themes.BOOTSTRAP

app: Dash = Dash(
    __name__,
    external_stylesheets=[THEME],
    title=APP_TITLE,
    suppress_callback_exceptions=True,
)
server = app.server

# ==============================
# ---- DEFAULT PRESET BANDS ----
# ==============================
DEFAULT_BANDS = [
    {"direction": "UI", "lower_pct": 0.0,  "upper_pct": 15.0, "tolerance_cut_pct": 15.0,
     "rate_type": "flat_per_kwh", "rate_value": 0.0, "excess_slope_per_pct": 0.0, "loss_zone": False, "label": "UI ≤15% (no penalty)"},
    {"direction": "UI", "lower_pct": 15.0, "upper_pct": 20.0, "tolerance_cut_pct": 15.0,
     "rate_type": "ppa_fraction", "rate_value": 0.10, "excess_slope_per_pct": 0.0, "loss_zone": False, "label": "UI 15–20% (10% of PPA)"},
    {"direction": "UI", "lower_pct": 20.0, "upper_pct": 1_000.0, "tolerance_cut_pct": 20.0,
     "rate_type": "scaled_excess", "rate_value": 3.36, "excess_slope_per_pct": 0.08, "loss_zone": False, "label": "UI >20% (scaled)"},
    {"direction": "OI", "lower_pct": 0.0,  "upper_pct": 15.0, "tolerance_cut_pct": 15.0,
     "rate_type": "flat_per_kwh", "rate_value": 0.0, "excess_slope_per_pct": 0.0, "loss_zone": False, "label": "OI ≤15% (no penalty)"},
    {"direction": "OI", "lower_pct": 15.0, "upper_pct": 20.0, "tolerance_cut_pct": 15.0,
     "rate_type": "ppa_fraction", "rate_value": 0.10, "excess_slope_per_pct": 0.0, "loss_zone": False, "label": "OI 15–20% (10% of PPA)"},
    {"direction": "OI", "lower_pct": 20.0, "upper_pct": 1_000.0, "tolerance_cut_pct": 20.0,
     "rate_type": "scaled_excess", "rate_value": 3.36, "excess_slope_per_pct": 0.08, "loss_zone": True, "label": "OI >20% (scaled)"},
]


BANDS_COLUMNS = [
    {"name": "Direction", "id": "direction", "presentation": "dropdown", "editable": True},
    {"name": "Lower %", "id": "lower_pct", "type": "numeric", "format": {"specifier": ".1f"}},
    {"name": "Upper %", "id": "upper_pct", "type": "numeric", "format": {"specifier": ".1f"}},
    # Tolerance Cut % and Deviated On are now hidden - handled by regulation mode
    {"name": "Rate Type", "id": "rate_type", "presentation": "dropdown", "editable": True},
    {"name": "Rate Value", "id": "rate_value", "type": "numeric", "format": {"specifier": ".2f"}},
    {"name": "Excess Slope %", "id": "excess_slope_per_pct", "type": "numeric", "format": {"specifier": ".2f"}},
    {"name": "Loss Zone", "id": "loss_zone", "presentation": "dropdown", "editable": True},
    {"name": "Label", "id": "label", "editable": False},
]

BANDS_DROPDOWNS = {
    "direction": {"options": [{"label": "UI (Under Injection)", "value": "UI"}, {"label": "OI (Over Injection)", "value": "OI"}]},
    "rate_type": {"options": [
                {"label": "Flat per kWh", "value": "flat_per_kwh"},
                {"label": "PPA Fraction", "value": "ppa_fraction"},
                {"label": "PPA Multiple", "value": "ppa_multiple"},
                {"label": "Flat per MWh", "value": "flat_per_mwh"},
                {"label": "Scaled Excess", "value": "scaled_excess"},
    ]},
    "loss_zone": {"options": [
                {"label": "No", "value": False},
                {"label": "Yes", "value": True},
    ]},
    # Apply To and Deviated On removed; denominator/basis is driven by Error% Mode
}

# =====================================
# ---------- DATA LOADING FUNCTIONS ---
# =====================================
import os
from functools import lru_cache
from pathlib import Path
import re
from datetime import datetime, timedelta

def _norm_plant_name(name: str) -> str:
    try:
        return str(name).strip().upper()
    except Exception:
        return str(name).upper()

# ------------------------------------------
# Plant classification helpers (Solar vs Wind)
# ------------------------------------------
def classify_plants_by_type(regions: list[str], start_date: str, end_date: str) -> dict:
    """Return mapping {plant_name: 'SOLAR'|'WIND'|'UNKNOWN'} using schedule window.
    Rule: 
    - WIND: Has scheduled power in early morning blocks (blocks 1-20, before 05:00)
    - SOLAR: Scheduled power only from block 21 onwards (05:00 or later)
    - UNKNOWN: No schedule data
    """
    mapping: dict[str, str] = {}
    regs = [str(r).strip().lower() for r in (regions or []) if str(r).strip()]
    if not regs:
        return mapping

    db_path = get_master_db_path()
    try:
        conn = duckdb.connect(db_path, read_only=True)
        try:
            placeholders = ", ".join(["?"] * len(regs))
            q = f"""
                SELECT plant_name,
                       MIN(time_block) AS min_block,
                       MAX(time_block) AS max_block
                FROM master
                WHERE lower(region) IN ({placeholders})
                  AND date >= ?
                  AND date <= ?
                  AND forecasted_power > 0
                GROUP BY plant_name
            """
            agg = conn.execute(q, [*regs, start_date, end_date]).fetchdf()
        finally:
            conn.close()

        for _, row in agg.iterrows():
            plant = row["plant_name"]
            try:
                min_b = int(row["min_block"]) if pd.notna(row["min_block"]) else None
                max_b = int(row["max_block"]) if pd.notna(row["max_block"]) else None
            except Exception:
                min_b, max_b = None, None
            if min_b is None or max_b is None:
                mapping[_norm_plant_name(plant)] = "UNKNOWN"
                continue
            # If minimum block is before block 21 (before 05:00), it's WIND
            # Block 21 = 05:00 AM (each block is 15 mins, so block 1 = 00:00-00:15)
            if min_b < 21:
                plant_type = "WIND"
                mapping[_norm_plant_name(plant)] = plant_type
                print(f"DEBUG: {plant} classified as {plant_type} (min_block={min_b}, max_block={max_b})")
            elif min_b >= 21:
                plant_type = "SOLAR"
                mapping[_norm_plant_name(plant)] = plant_type
                print(f"DEBUG: {plant} classified as {plant_type} (min_block={min_b}, max_block={max_b})")
            else:
                mapping[_norm_plant_name(plant)] = "UNKNOWN"

    except Exception as e:
        print(f"Error classifying plants from master ({db_path}): {e}")

    return mapping

@lru_cache(maxsize=1)
def load_plant_renewable_mapping() -> dict[str, str]:
    """Load plant-to-renewable mapping from consolidated_plant_list.xlsx.
    Returns dict mapping normalized plant name to renewable type (Solar/Wind/Thermal).
    """
    mapping: dict[str, str] = {}

    # Prefer resolving relative to THIS file to avoid issues when cwd changes (Dash reloaders/callbacks).
    script_path = Path(__file__).resolve()
    script_dir = script_path.parent

    candidates: list[Path] = [
        script_dir / "consolidated_plant_list.xlsx",          # same folder as this script (DSM MASTER)
        script_dir.parent / "consolidated_plant_list.xlsx",   # project root (if someone moved the file)
        script_dir.parent / "DSM MASTER" / "consolidated_plant_list.xlsx",  # explicit DSM MASTER from root
        Path.cwd() / "consolidated_plant_list.xlsx",          # cwd fallback
        Path.cwd() / "DSM MASTER" / "consolidated_plant_list.xlsx",  # cwd + DSM MASTER fallback
    ]

    excel_path: Path | None = next((p for p in candidates if p.exists()), None)

    # Last-resort: cheap bounded search (only within the script dir and its parent).
    if excel_path is None:
        for base in (script_dir, script_dir.parent):
            try:
                found = next(base.rglob("consolidated_plant_list.xlsx"), None)
            except Exception:
                found = None
            if found is not None and found.exists():
                excel_path = found
                break

    if excel_path is None:
        attempted = ", ".join(str(p) for p in candidates)
        print(
            "[WARN] consolidated_plant_list.xlsx not found. "
            "Plant filtering by resource will not work. "
            f"(cwd={os.getcwd()}, __file__={script_path}, attempted={attempted})"
        )
        return mapping
    
    try:
        df = pd.read_excel(str(excel_path))
        if "plantname" not in df.columns or "renewable" not in df.columns:
            print(f"[WARN] {excel_path} missing required columns (plantname, renewable).")
            return mapping
        
        for _, row in df.iterrows():
            plant = str(row["plantname"]).strip()
            renewable = str(row["renewable"]).strip()
            if plant and renewable:
                mapping[_norm_plant_name(plant)] = renewable
    except Exception as e:
        print(f"[WARN] Error loading {excel_path}: {e}")
    
    return mapping

def get_filtered_plants_by_type(regions: list[str], type_value: str, start_date: str, end_date: str) -> list[str]:
    """Return plants filtered by type selection (ALL/SOLAR/WIND/THERMAL).
    Uses consolidated_plant_list.xlsx if available, otherwise falls back to classification logic.
    """
    all_plants = get_plants_from_duckdb(regions)
    t = (type_value or "ALL").upper()
    if t == "ALL":
        return all_plants
    
    # Try to use Excel mapping first
    excel_mapping = load_plant_renewable_mapping()
    if excel_mapping:
        # Map Excel renewable values to filter values (Solar->SOLAR, Wind->WIND, Thermal->THERMAL)
        renewable_map = {"Solar": "SOLAR", "Wind": "WIND", "Thermal": "THERMAL"}
        filtered = []
        for p in all_plants:
            plant_norm = _norm_plant_name(p)
            excel_type = excel_mapping.get(plant_norm, "").strip()
            mapped_type = renewable_map.get(excel_type, "").upper()
            if mapped_type == t:
                filtered.append(p)
        
        if filtered:
            return filtered
        # If no matches found with Excel, fall back to classification logic
    
    # Fallback to old classification logic (only handles SOLAR/WIND, not THERMAL)
    mapping = classify_plants_by_type(regions, start_date, end_date)
    filtered = [p for p in all_plants if mapping.get(_norm_plant_name(p), "UNKNOWN") == t]
    return filtered if filtered else all_plants

# ==========================================
# -------- DUCKDB HELPER FUNCTIONS ---------
# ==========================================
import duckdb

from utils.config import get_master_db_path
from data.loader import MasterQuery, list_plants, list_qcas, list_regions, load_master_frame
from data.health import summarize_health_from_df

# Dedupe DB error logging (avoid spam when multiple callbacks fail)
_db_unavailable_logged = False

def get_regions_from_duckdb() -> list[str]:
    """Get unique regions from master.duckdb (single source of truth)."""
    global _db_unavailable_logged
    db_path = get_master_db_path()
    try:
        out = list_regions(db_path)
        _db_unavailable_logged = False  # Reset on success
        return out
    except Exception as e:
        if not _db_unavailable_logged:
            _db_unavailable_logged = True
            err = str(e).lower()
            if "cannot open" in err or "in use" in err or "another process" in err:
                print("[DSM] Database is in use. Close other apps (dashboard, DBeaver, ingestion) and refresh.")
            else:
                print(f"[DSM] Database error: {e}")
        return []

def get_plants_from_duckdb(regions: list[str]) -> list[str]:
    """Get unique plant names from master.duckdb for selected regions."""
    db_path = get_master_db_path()
    try:
        return list_plants(db_path, regions)
    except Exception:
        return []


def get_qcas_from_duckdb(regions: list[str], plants: list[str] | None = None) -> list[str]:
    """Get unique QCA labels from master.duckdb for selected regions/plants."""
    db_path = get_master_db_path()
    try:
        return list_qcas(db_path, regions, plants)
    except Exception:
        return []

def load_data_from_duckdb(
    regions: list[str],
    start_date: str,
    end_date: str,
    plants: list[str],
    qcas: list[str] | None = None,
) -> pd.DataFrame:
    """Load data from master.duckdb for specified regions, date range, and plants."""
    db_path = get_master_db_path()
    try:
        regs = [r for r in (regions or []) if r and str(r).strip()]
        if not regs:
            return pd.DataFrame()
        qca_list = [q for q in (qcas or []) if q and q != "SELECT_ALL"]
        q = MasterQuery(
            regions=regs,
            start_date=start_date,
            end_date=end_date,
            plants=plants or [],
            qcas=qca_list or None,
        )
        return load_master_frame(db_path, q)
    except Exception as e:
        print(f"Error loading data from master {db_path}: {e}")
        return pd.DataFrame()

# ==========================================
# -------- REAL DATA LOADING FUNCTIONS --------
# ==========================================

def load_nrpc_data(start_date, end_date, plants):
    """Load NRPC data - Legacy function, now redirects to DuckDB"""
    return load_data_from_duckdb(["NRPC"], start_date, end_date, plants)

def load_srpc_data(start_date, end_date, plants):
    """Load SRPC data - Legacy function, now redirects to DuckDB"""
    return load_data_from_duckdb(["SRPC"], start_date, end_date, plants)

def load_wrpc_data(start_date, end_date, plants):
    """Load WRPC data - Legacy function, now redirects to DuckDB"""
    return load_data_from_duckdb(["WRPC"], start_date, end_date, plants)

def load_region_data(regions, start_date, end_date, plants, qcas: list[str] | None = None):
    """Load data from master.duckdb for selected region(s). Supports multi-region."""
    regs = [regions] if isinstance(regions, str) else (regions or [])
    return load_data_from_duckdb(regs, start_date, end_date, plants, qcas=qcas)

def make_sample_blocks(start_date: str, end_date: str, plant: str) -> pd.DataFrame:
    start = datetime.fromisoformat(start_date)
    end = datetime.fromisoformat(end_date)
    if end < start:
        end = start

    times = pd.date_range(start, end + timedelta(days=1), freq="15min", inclusive="left")
    df = pd.DataFrame({"date_time": times})
    df["Plant"] = plant

    base_avc = 51.3 if "A" in plant or "B" in plant else 75.0
    rng = np.random.default_rng(42)
    df["AvC_MW"] = (base_avc + rng.normal(0, 0.3, len(df))).round(3)

    hours = df["date_time"].dt.hour + df["date_time"].dt.minute / 60.0
    df["Scheduled_MW"] = (
        (0.55 * df["AvC_MW"]) +
        8 * np.sin((hours / 24) * 2 * np.pi) +
        rng.normal(0, 1.2, len(df))
    ).clip(lower=0).round(4)

    bias = np.where((hours >= 13) & (hours <= 16), -6.0, 0.0)
    df["Actual_MW"] = (df["Scheduled_MW"] + bias + rng.normal(0, 2.0, len(df))).clip(lower=0).round(4)

    df["date"] = df["date_time"].dt.date
    daily_ppa = {d: (4.20 if i % 2 == 0 else 3.95) for i, d in enumerate(sorted(df["date"].unique()))}
    df["PPA"] = df["date"].map(daily_ppa)

    df["block"] = df.groupby("date").cumcount() + 1
    return df

# ==========================================
# --------- ENGINE: see core.dsm_engine -------
# ==========================================

def xlsx_col(idx: int) -> str:
    s = ""
    n = idx + 1
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

def export_with_formulas(detail_rows: pd.DataFrame, bands_df: pd.DataFrame, mode: str, dyn_x: float) -> bytes:
    """Build an Excel file (Bands, Config, Detail) with embedded formulas."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as xw:
        wb = xw.book
        # Bands
        bands_out = (bands_df or pd.DataFrame(columns=["direction","lower_pct","upper_pct","rate_type","rate_value","rate_slope","loss_zone"]))
        bands_out = bands_out[[
            "direction","lower_pct","upper_pct",
            "rate_type","rate_value","rate_slope","loss_zone"
        ]].reset_index(drop=True)
        bands_out.to_excel(xw, sheet_name="Bands", index=False, startrow=0)
        nrows = len(bands_out)
        def NR(col_letter: str) -> str:
            return f"Bands!${col_letter}$2:${col_letter}${nrows+1}" if nrows > 0 else f"Bands!${col_letter}$2:${col_letter}$2"
        wb.define_name("Bands_Dir",      f"={NR('A')}")
        wb.define_name("Bands_Lower",    f"={NR('B')}")
        wb.define_name("Bands_Upper",    f"={NR('C')}")
        wb.define_name("Bands_RateType", f"={NR('D')}")
        wb.define_name("Bands_RateVal",  f"={NR('E')}")
        wb.define_name("Bands_RateSlope",f"={NR('F')}")
        wb.define_name("Bands_LossZone", f"={NR('G')}")
        # Config
        cfg = pd.DataFrame({"Key": ["MODE","DYN_X"], "Value": [mode, dyn_x]})
        cfg.to_excel(xw, sheet_name="Config", index=False)
        wb.define_name("CFG_MODE", "=Config!$B$2")
        wb.define_name("CFG_DYNX", "=Config!$B$3")
        # Detail
        ws = wb.add_worksheet("Detail")
        headers = [
            "region","plant_name","date","time_block","from_time","to_time",
            "AvC_MW","Scheduled_MW","Actual_MW","PPA",
            "error_pct","direction","abs_err","band_level",
            "UI_Energy_deviation_bands","OI_Energy_deviation_bands",
            "Revenue_as_per_generation","Scheduled_Revenue_as_per_generation",
            "UI_DSM","OI_DSM","OI_Loss","Total_DSM","Revenue_Loss"
        ]
        for c, h in enumerate(headers):
            ws.write(0, c, h)
        COL = {h: i for i, h in enumerate(headers)}
        BASIS_COL = len(headers)
        ws.write(0, BASIS_COL, "_basis_helper")
        start_row = 1
        for r, row in enumerate(detail_rows.itertuples(index=False), start=start_row):
            ws.write(r, COL["region"],       getattr(row, "region"))
            ws.write(r, COL["plant_name"],   getattr(row, "plant_name"))
            ws.write(r, COL["date"],         getattr(row, "date"))
            ws.write(r, COL["time_block"],   getattr(row, "time_block"))
            ws.write(r, COL["from_time"],    getattr(row, "from_time"))
            ws.write(r, COL["to_time"],      getattr(row, "to_time"))
            ws.write_number(r, COL["AvC_MW"],       float(getattr(row, "AvC_MW")))
            ws.write_number(r, COL["Scheduled_MW"], float(getattr(row, "Scheduled_MW")))
            ws.write_number(r, COL["Actual_MW"],    float(getattr(row, "Actual_MW")))
            ws.write_number(r, COL["PPA"],          float(getattr(row, "PPA")))
            row1 = r + 1
            avc_ref = f"{xlsx_col(COL['AvC_MW'])}{row1}"
            sch_ref = f"{xlsx_col(COL['Scheduled_MW'])}{row1}"
            act_ref = f"{xlsx_col(COL['Actual_MW'])}{row1}"
            ppa_ref = f"{xlsx_col(COL['PPA'])}{row1}"
            basis_formula = f'=IF(CFG_MODE="DEFAULT",{avc_ref}, CFG_DYNX*{avc_ref} + (1-CFG_DYNX)*{sch_ref})'
            ws.write_formula(r, BASIS_COL, basis_formula)
            basis_ref = f"{xlsx_col(BASIS_COL)}{row1}"
            err_formula = f"=IF({basis_ref}=0,0, ({act_ref}-{sch_ref})/{basis_ref}*100)"
            ws.write_formula(r, COL["error_pct"], err_formula)
            dir_formula = f'=IF({act_ref}<{sch_ref},"UI",IF({act_ref}>{sch_ref},"OI","FLAT"))'
            ws.write_formula(r, COL["direction"], dir_formula)
            abs_formula = f"=ABS({xlsx_col(COL['error_pct'])}{row1})"
            ws.write_formula(r, COL["abs_err"], abs_formula)
            ws.write(r, COL["band_level"], "")
            ui_dev_formula = f'=IF({xlsx_col(COL["direction"])}{row1}="UI",{xlsx_col(COL["abs_err"])}{row1}/100*{basis_ref}*0.25*1000,0)'
            oi_dev_formula = f'=IF({xlsx_col(COL["direction"])}{row1}="OI",{xlsx_col(COL["abs_err"])}{row1}/100*{basis_ref}*0.25*1000,0)'
            ws.write_formula(r, COL["UI_Energy_deviation_bands"], ui_dev_formula)
            ws.write_formula(r, COL["OI_Energy_deviation_bands"], oi_dev_formula)
            rev_act_formula = f"={act_ref}*0.25*1000*{ppa_ref}"
            rev_sch_formula = f"={sch_ref}*0.25*1000*{ppa_ref}"
            ws.write_formula(r, COL["Revenue_as_per_generation"], rev_act_formula)
            ws.write_formula(r, COL["Scheduled_Revenue_as_per_generation"], rev_sch_formula)
            abs_ref = f"{xlsx_col(COL['abs_err'])}{row1}"
            slice_factor = f"MAX(0, MIN({abs_ref}, Bands_Upper) - Bands_Lower)/100 * {basis_ref} * 0.25*1000"
            rate_expr = (
                'IF(Bands_RateType="PPA_FRAC", ' + ppa_ref + '*Bands_RateVal, '
                'IF(Bands_RateType="PPA_MULT", ' + ppa_ref + '*Bands_RateVal, '
                'IF(Bands_RateType="FLAT", Bands_RateVal, '
                'IF(Bands_RateType="SCALED", Bands_RateVal + Bands_RateSlope*' + abs_ref + ', 0))))'
            )
            ui_sp = '=SUMPRODUCT(--(Bands_Dir="UI"),' + slice_factor + ',' + rate_expr + ')'
            oi_dsm_sp = '=SUMPRODUCT(--(Bands_Dir="OI"),--(Bands_LossZone=FALSE),' + slice_factor + ',' + rate_expr + ')'
            oi_loss_sp = '=SUMPRODUCT(--(Bands_Dir="OI"),--(Bands_LossZone=TRUE),' + slice_factor + ',' + rate_expr + ')'
            ws.write_formula(r, COL["UI_DSM"], ui_sp)
            ws.write_formula(r, COL["OI_DSM"], oi_dsm_sp)
            ws.write_formula(r, COL["OI_Loss"], oi_loss_sp)
            ws.write_formula(r, COL["Total_DSM"], f"={xlsx_col(COL['UI_DSM'])}{row1}+{xlsx_col(COL['OI_DSM'])}{row1}")
            ws.write_formula(r, COL["Revenue_Loss"], f"={xlsx_col(COL['Total_DSM'])}{row1}+{xlsx_col(COL['OI_Loss'])}{row1}")
        ws.set_column(BASIS_COL, BASIS_COL, None, None, {'hidden': True})
    output.seek(0)
    return output.getvalue()

def export_with_formulas_openpyxl(detail_rows: pd.DataFrame, bands_df: pd.DataFrame, mode: str, dyn_x: float) -> bytes:
    """OpenPyXL implementation of the same formula-driven workbook with named ranges."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    # Try to import DefinedName (optional). If unavailable, formulas will use absolute refs instead of names.
    try:
        from openpyxl.workbook.defined_name import DefinedName  # type: ignore
    except Exception:
        DefinedName = None  # type: ignore

    wb = Workbook()
    # Clear default sheet
    ws_default = wb.active
    wb.remove(ws_default)

    # Bands sheet
    ws_b = wb.create_sheet("Bands")
    bands_out = bands_df.copy().reset_index(drop=True)
    cols = ["direction","lower_pct","upper_pct","rate_type","rate_value","rate_slope","loss_zone"]
    for c_idx, h in enumerate(cols, start=1):
        ws_b.cell(row=1, column=c_idx, value=h)
    for r_idx, row in enumerate(bands_out.itertuples(index=False), start=2):
        ws_b.cell(row=r_idx, column=1, value=getattr(row, "direction"))
        ws_b.cell(row=r_idx, column=2, value=float(getattr(row, "lower_pct")))
        ws_b.cell(row=r_idx, column=3, value=float(getattr(row, "upper_pct")))
        ws_b.cell(row=r_idx, column=4, value=str(getattr(row, "rate_type")))
        ws_b.cell(row=r_idx, column=5, value=float(getattr(row, "rate_value")))
        ws_b.cell(row=r_idx, column=6, value=float(getattr(row, "rate_slope")))
        ws_b.cell(row=r_idx, column=7, value=bool(getattr(row, "loss_zone")))
    nrows = len(bands_out)
    last_row = 1 + (nrows if nrows > 0 else 1)
    def ref(col):
        return f"Bands!${col}$2:${col}${last_row}"
    # Named ranges (best-effort)
    if DefinedName is not None:
        wb.defined_names.append(DefinedName(name="Bands_Dir", attr_text=ref('A')))
        wb.defined_names.append(DefinedName(name="Bands_Lower", attr_text=ref('B')))
        wb.defined_names.append(DefinedName(name="Bands_Upper", attr_text=ref('C')))
        wb.defined_names.append(DefinedName(name="Bands_RateType", attr_text=ref('D')))
        wb.defined_names.append(DefinedName(name="Bands_RateVal", attr_text=ref('E')))
        wb.defined_names.append(DefinedName(name="Bands_RateSlope", attr_text=ref('F')))
        wb.defined_names.append(DefinedName(name="Bands_LossZone", attr_text=ref('G')))
    # Absolute ranges (always available)
    BANDS_DIR = ref('A')
    BANDS_LOWER = ref('B')
    BANDS_UPPER = ref('C')
    BANDS_RATETYPE = ref('D')
    BANDS_RATEVAL = ref('E')
    BANDS_RATESLOPE = ref('F')
    BANDS_LOSSZONE = ref('G')

    # Config
    ws_c = wb.create_sheet("Config")
    ws_c.cell(row=1, column=1, value="Key")
    ws_c.cell(row=1, column=2, value="Value")
    ws_c.cell(row=2, column=1, value="MODE")
    ws_c.cell(row=2, column=2, value=mode)
    ws_c.cell(row=3, column=1, value="DYN_X")
    ws_c.cell(row=3, column=2, value=float(dyn_x))
    # Named cells (best-effort)
    if DefinedName is not None:
        wb.defined_names.append(DefinedName(name="CFG_MODE", attr_text="Config!$B$2"))
        wb.defined_names.append(DefinedName(name="CFG_DYNX", attr_text="Config!$B$3"))
    CFG_MODE_REF = "Config!$B$2"
    CFG_DYNX_REF = "Config!$B$3"

    # Detail with formulas
    ws = wb.create_sheet("Detail")
    headers = [
        "region","plant_name","date","time_block","from_time","to_time",
        "AvC_MW","Scheduled_MW","Actual_MW","PPA",
        "error_pct","direction","abs_err","band_level",
        "UI_Energy_deviation_bands","OI_Energy_deviation_bands",
        "Revenue_as_per_generation","Scheduled_Revenue_as_per_generation",
        "UI_DSM","OI_DSM","OI_Loss","Total_DSM","Revenue_Loss"
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    col_index = {h: i+1 for i, h in enumerate(headers)}
    BASIS_COL = len(headers) + 1
    ws.cell(row=1, column=BASIS_COL, value="_basis_helper")

    for r_idx, row in enumerate(detail_rows.itertuples(index=False), start=2):
        ws.cell(row=r_idx, column=col_index["region"], value=getattr(row, "region"))
        ws.cell(row=r_idx, column=col_index["plant_name"], value=getattr(row, "plant_name"))
        ws.cell(row=r_idx, column=col_index["date"], value=getattr(row, "date"))
        ws.cell(row=r_idx, column=col_index["time_block"], value=getattr(row, "time_block"))
        ws.cell(row=r_idx, column=col_index["from_time"], value=getattr(row, "from_time"))
        ws.cell(row=r_idx, column=col_index["to_time"], value=getattr(row, "to_time"))
        ws.cell(row=r_idx, column=col_index["AvC_MW"], value=float(getattr(row, "AvC_MW")))
        ws.cell(row=r_idx, column=col_index["Scheduled_MW"], value=float(getattr(row, "Scheduled_MW")))
        ws.cell(row=r_idx, column=col_index["Actual_MW"], value=float(getattr(row, "Actual_MW")))
        ws.cell(row=r_idx, column=col_index["PPA"], value=float(getattr(row, "PPA")))

        row1 = r_idx
        def A1(col_name):
            return f"{get_column_letter(col_index[col_name])}{row1}"
        avc_ref = A1("AvC_MW")
        sch_ref = A1("Scheduled_MW")
        act_ref = A1("Actual_MW")
        ppa_ref = A1("PPA")
        basis_ref = f"{get_column_letter(BASIS_COL)}{row1}"

        ws.cell(row=row1, column=BASIS_COL, value=f'=IF({CFG_MODE_REF}="DEFAULT",{avc_ref}, {CFG_DYNX_REF}*{avc_ref} + (1-{CFG_DYNX_REF})*{sch_ref})')
        ws.cell(row=row1, column=col_index["error_pct"], value=f"=IF({basis_ref}=0,0, ({act_ref}-{sch_ref})/{basis_ref}*100)")
        ws.cell(row=row1, column=col_index["direction"], value=f'=IF({act_ref}<{sch_ref},"UI",IF({act_ref}>{sch_ref},"OI","FLAT"))')
        ws.cell(row=row1, column=col_index["abs_err"], value=f"=ABS({A1('error_pct')})")
        # band_level left blank
        ws.cell(row=row1, column=col_index["UI_Energy_deviation_bands"], value=f'=IF({A1("direction")}="UI",{A1("abs_err")}/100*{basis_ref}*0.25*1000,0)')
        ws.cell(row=row1, column=col_index["OI_Energy_deviation_bands"], value=f'=IF({A1("direction")}="OI",{A1("abs_err")}/100*{basis_ref}*0.25*1000,0)')
        ws.cell(row=row1, column=col_index["Revenue_as_per_generation"], value=f"={act_ref}*0.25*1000*{ppa_ref}")
        ws.cell(row=row1, column=col_index["Scheduled_Revenue_as_per_generation"], value=f"={sch_ref}*0.25*1000*{ppa_ref}")

        abs_ref = A1("abs_err")
        slice_factor = (
            f"MAX(0, MIN({abs_ref}, {BANDS_UPPER}) - {BANDS_LOWER})/100 * {basis_ref} * 0.25*1000"
        )
        rate_expr = (
            'IF(' + BANDS_RATETYPE + '="PPA_FRAC", ' + ppa_ref + '*' + BANDS_RATEVAL + ', '
            'IF(' + BANDS_RATETYPE + '="PPA_MULT", ' + ppa_ref + '*' + BANDS_RATEVAL + ', '
            'IF(' + BANDS_RATETYPE + '="FLAT", ' + BANDS_RATEVAL + ', '
            'IF(' + BANDS_RATETYPE + '="SCALED", ' + BANDS_RATEVAL + ' + ' + BANDS_RATESLOPE + '*' + abs_ref + ', 0))))'
        )
        ui_sp = '=SUMPRODUCT(--(' + BANDS_DIR + '="UI"),' + slice_factor + ',' + rate_expr + ')'
        oi_dsm_sp = '=SUMPRODUCT(--(' + BANDS_DIR + '="OI"),--(' + BANDS_LOSSZONE + '=FALSE),' + slice_factor + ',' + rate_expr + ')'
        oi_loss_sp = '=SUMPRODUCT(--(' + BANDS_DIR + '="OI"),--(' + BANDS_LOSSZONE + '=TRUE),' + slice_factor + ',' + rate_expr + ')'
        ws.cell(row=row1, column=col_index["UI_DSM"], value=ui_sp)
        ws.cell(row=row1, column=col_index["OI_DSM"], value=oi_dsm_sp)
        ws.cell(row=row1, column=col_index["OI_Loss"], value=oi_loss_sp)
        ws.cell(row=row1, column=col_index["Total_DSM"], value=f"={get_column_letter(col_index['UI_DSM'])}{row1}+{get_column_letter(col_index['OI_DSM'])}{row1}")
        ws.cell(row=row1, column=col_index["Revenue_Loss"], value=f"={get_column_letter(col_index['Total_DSM'])}{row1}+{get_column_letter(col_index['OI_Loss'])}{row1}")

    # Hide basis column
    ws.column_dimensions[get_column_letter(BASIS_COL)].hidden = True

    from io import BytesIO
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

def export_with_formulas_opc(detail_rows: pd.DataFrame, bands_df: pd.DataFrame, mode: str, dyn_x: float) -> bytes:
    """Create Excel workbook (Detail/Bands/Config) with formulas using only stdlib (zip + XML)."""
    from zipfile import ZipFile, ZIP_DEFLATED
    from datetime import datetime
    import xml.sax.saxutils as saxutils

    headers = [
        "region","plant_name","date","time_block","from_time","to_time",
        "AvC_MW","Scheduled_MW","Actual_MW","PPA",
        "error_pct","direction","abs_err","band_level",
        "UI_Energy_deviation_bands","OI_Energy_deviation_bands",
        "Revenue_as_per_generation","Scheduled_Revenue_as_per_generation",
        "UI_DSM","OI_DSM","OI_Loss","Total_DSM","Revenue_Loss"
    ]
    basis_col_num = len(headers) + 1
    basis_col_letter = xlsx_col(basis_col_num - 1)

    def coord(col_num: int, row_num: int) -> str:
        return f"{xlsx_col(col_num - 1)}{row_num}"

    def cell_inline(col_num: int, row_num: int, text: str) -> str:
        txt = saxutils.escape(text or "")
        return f'<c r="{coord(col_num, row_num)}" t="inlineStr"><is><t>{txt}</t></is></c>'

    def cell_number(col_num: int, row_num: int, value) -> str:
        val = "" if value is None else ("0" if value == 0 else f"{value}")
        return f'<c r="{coord(col_num, row_num)}"><v>{val}</v></c>'

    def cell_formula(col_num: int, row_num: int, formula: str) -> str:
        frm = saxutils.escape(formula)
        return f'<c r="{coord(col_num, row_num)}"><f>{frm}</f></c>'

    # Bands ranges for formulas
    bands_len = len(bands_df)
    bands_last = 1 + (bands_len if bands_len > 0 else 1)
    band_range = lambda col_letter: f'Bands!${col_letter}$2:${col_letter}${bands_last}'
    bands_dir_ref = band_range('A')
    bands_lower_ref = band_range('B')
    bands_upper_ref = band_range('C')
    bands_ratetype_ref = band_range('D')
    bands_rateval_ref = band_range('E')
    bands_rateslope_ref = band_range('F')
    bands_losszone_ref = band_range('G')

    cfg_mode_ref = 'Config!$B$2'
    cfg_dnx_ref = 'Config!$B$3'

    rows_xml = []
    header_cells = [cell_inline(idx + 1, 1, hdr) for idx, hdr in enumerate(headers)]
    header_cells.append(cell_inline(basis_col_num, 1, '_basis_helper'))
    rows_xml.append(f'<row r="1">{"".join(header_cells)}</row>')

    for idx, row in enumerate(detail_rows.itertuples(index=False), start=2):
        cells = []
        cells.append(cell_inline(1, idx, str(getattr(row, 'region', ''))))
        cells.append(cell_inline(2, idx, str(getattr(row, 'plant_name', ''))))
        cells.append(cell_inline(3, idx, str(getattr(row, 'date', ''))))
        cells.append(cell_number(4, idx, getattr(row, 'time_block', '')))
        cells.append(cell_inline(5, idx, str(getattr(row, 'from_time', ''))))
        cells.append(cell_inline(6, idx, str(getattr(row, 'to_time', ''))))
        cells.append(cell_number(7, idx, float(getattr(row, 'AvC_MW'))))
        cells.append(cell_number(8, idx, float(getattr(row, 'Scheduled_MW'))))
        cells.append(cell_number(9, idx, float(getattr(row, 'Actual_MW'))))
        cells.append(cell_number(10, idx, float(getattr(row, 'PPA'))))

        avc_ref = coord(7, idx)
        sch_ref = coord(8, idx)
        act_ref = coord(9, idx)
        ppa_ref = coord(10, idx)
        basis_ref = f'${basis_col_letter}${idx}'
        error_ref = coord(11, idx)
        direction_ref = coord(12, idx)
        abs_ref = coord(13, idx)

        basis_formula = f'IF({cfg_mode_ref}="DEFAULT",{avc_ref}, {cfg_dnx_ref}*{avc_ref} + (1-{cfg_dnx_ref})*{sch_ref})'
        cells.append(cell_formula(basis_col_num, idx, basis_formula))
        cells.append(cell_formula(11, idx, f'IF({basis_ref}=0,0, ({act_ref}-{sch_ref})/{basis_ref}*100)'))
        cells.append(cell_formula(12, idx, f'IF({act_ref}<{sch_ref},"UI",IF({act_ref}>{sch_ref},"OI","FLAT"))'))
        cells.append(cell_formula(13, idx, f'ABS({error_ref})'))
        cells.append(cell_inline(14, idx, ""))  # band_level blank
        cells.append(cell_formula(15, idx, f'IF({direction_ref}="UI",{abs_ref}/100*{basis_ref}*0.25*1000,0)'))
        cells.append(cell_formula(16, idx, f'IF({direction_ref}="OI",{abs_ref}/100*{basis_ref}*0.25*1000,0)'))
        cells.append(cell_formula(17, idx, f'{act_ref}*0.25*1000*{ppa_ref}'))
        cells.append(cell_formula(18, idx, f'{sch_ref}*0.25*1000*{ppa_ref}'))

        slice_factor = f'MAX(0, MIN({abs_ref}, {bands_upper_ref}) - {bands_lower_ref})/100 * {basis_ref} * 0.25*1000'
        rate_expr = (
            f'IF({bands_ratetype_ref}="PPA_FRAC", {ppa_ref}*{bands_rateval_ref}, '
            f'IF({bands_ratetype_ref}="PPA_MULT", {ppa_ref}*{bands_rateval_ref}, '
            f'IF({bands_ratetype_ref}="FLAT", {bands_rateval_ref}, '
            f'IF({bands_ratetype_ref}="SCALED", {bands_rateval_ref} + {bands_rateslope_ref}*{abs_ref}, 0))))'
        )
        cells.append(cell_formula(19, idx, f'SUMPRODUCT(--({bands_dir_ref}="UI"),{slice_factor},{rate_expr})'))
        cells.append(cell_formula(20, idx, f'SUMPRODUCT(--({bands_dir_ref}="OI"),--({bands_losszone_ref}=FALSE),{slice_factor},{rate_expr})'))
        cells.append(cell_formula(21, idx, f'SUMPRODUCT(--({bands_dir_ref}="OI"),--({bands_losszone_ref}=TRUE),{slice_factor},{rate_expr})'))
        ui_ref = coord(19, idx)
        oi_ref = coord(20, idx)
        oi_loss_ref = coord(21, idx)
        cells.append(cell_formula(22, idx, f'{ui_ref}+{oi_ref}'))
        cells.append(cell_formula(23, idx, f'{coord(22, idx)}+{oi_loss_ref}'))

        rows_xml.append(f'<row r="{idx}">{"".join(cells)}</row>')

    cols_xml = f'<cols><col min="{basis_col_num}" max="{basis_col_num}" hidden="1" width="0"/></cols>'
    sheet_data_xml = ''.join(rows_xml)
    detail_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'{cols_xml}<sheetData>{sheet_data_xml}</sheetData></worksheet>'
    )

    # Bands sheet
    bands_rows = []
    band_headers = ["direction","lower_pct","upper_pct","rate_type","rate_value","rate_slope","loss_zone"]
    bands_header_cells = [cell_inline(i+1, 1, h) for i, h in enumerate(band_headers)]
    bands_rows.append(f'<row r="1">{"".join(bands_header_cells)}</row>')
    for idx, row in enumerate(bands_df.reset_index(drop=True).itertuples(index=False), start=2):
        cells = [
            cell_inline(1, idx, str(getattr(row, 'direction'))),
            cell_number(2, idx, float(getattr(row, 'lower_pct'))),
            cell_number(3, idx, float(getattr(row, 'upper_pct'))),
            cell_inline(4, idx, str(getattr(row, 'rate_type'))),
            cell_number(5, idx, float(getattr(row, 'rate_value'))),
            cell_number(6, idx, float(getattr(row, 'rate_slope'))),
            cell_inline(7, idx, 'TRUE' if bool(getattr(row, 'loss_zone')) else 'FALSE'),
        ]
        bands_rows.append(f'<row r="{idx}">{"".join(cells)}</row>')
    bands_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<sheetData>{"".join(bands_rows)}</sheetData></worksheet>'
    )

    # Config sheet
    config_rows = [
        '<row r="1">' + cell_inline(1,1,"Key") + cell_inline(2,1,"Value") + '</row>',
        '<row r="2">' + cell_inline(1,2,"MODE") + cell_inline(2,2,str(mode)) + '</row>',
        '<row r="3">' + cell_inline(1,3,"DYN_X") + cell_number(2,3,float(dyn_x)) + '</row>',
    ]
    config_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<sheetData>{"".join(config_rows)}</sheetData></worksheet>'
    )

    # Styles (minimal)
    styles_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<fonts count="1"><font><sz val="11"/><color theme="1"/><name val="Calibri"/><family val="2"/></font></fonts>'
        '<fills count="1"><fill><patternFill patternType="none"/></fill></fills>'
        '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>'
        '</styleSheet>'
    )

    # Workbook + rels
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets>'
        '<sheet name="Detail" sheetId="1" r:id="rId1"/>'
        '<sheet name="Bands" sheetId="2" r:id="rId2"/>'
        '<sheet name="Config" sheetId="3" r:id="rId3"/>'
        '</sheets>'
        '</workbook>'
    )

    workbook_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet2.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet3.xml"/>'
        '<Relationship Id="rId4" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
        '</Relationships>'
    )

    content_types_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/worksheets/sheet2.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/worksheets/sheet3.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
        '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
        '</Types>'
    )

    root_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
        '</Relationships>'
    )

    # docProps
    created_ts = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')
    core_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        'xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" '
        'xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        f'<dcterms:created xsi:type="dcterms:W3CDTF">{created_ts}</dcterms:created>'
        '<dc:creator>DSM Dashboard</dc:creator>'
        '<cp:lastModifiedBy>DSM Dashboard</cp:lastModifiedBy>'
        f'<dcterms:modified xsi:type="dcterms:W3CDTF">{created_ts}</dcterms:modified>'
        '</cp:coreProperties>'
    )

    app_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
        'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
        '<Application>DSM Dashboard</Application>'
        '<DocSecurity>0</DocSecurity>'
        '<ScaleCrop>0</ScaleCrop>'
        '<HeadingPairs><vt:vector size="2" baseType="variant">'
        '<vt:variant><vt:lpstr>Worksheets</vt:lpstr></vt:variant>'
        '<vt:variant><vt:i4>3</vt:i4></vt:variant>'
        '</vt:vector></HeadingPairs>'
        '<TitlesOfParts><vt:vector size="3" baseType="lpstr">'
        '<vt:lpstr>Detail</vt:lpstr><vt:lpstr>Bands</vt:lpstr><vt:lpstr>Config</vt:lpstr>'
        '</vt:vector></TitlesOfParts>'
        '</Properties>'
    )

    out = BytesIO()
    with ZipFile(out, 'w', ZIP_DEFLATED) as zf:
        zf.writestr('[Content_Types].xml', content_types_xml)
        zf.writestr('_rels/.rels', root_rels_xml)
        zf.writestr('docProps/core.xml', core_xml)
        zf.writestr('docProps/app.xml', app_xml)
        zf.writestr('xl/workbook.xml', workbook_xml)
        zf.writestr('xl/_rels/workbook.xml.rels', workbook_rels_xml)
        zf.writestr('xl/styles.xml', styles_xml)
        zf.writestr('xl/worksheets/sheet1.xml', detail_xml)
        zf.writestr('xl/worksheets/sheet2.xml', bands_xml)
        zf.writestr('xl/worksheets/sheet3.xml', config_xml)
    out.seek(0)
    return out.getvalue()

def build_summary_for_screen(detail_numeric_df: pd.DataFrame) -> Dict[str, Any]:
    avc_mode = safe_mode(detail_numeric_df["AvC_MW"].tolist()) if "AvC_MW" in detail_numeric_df.columns else 0.0
    ppa_mode = safe_mode(detail_numeric_df["PPA"].tolist()) if "PPA" in detail_numeric_df.columns else 0.0
    rev_loss_sum = float(detail_numeric_df.get("Revenue_Loss", pd.Series(dtype=float)).sum())
    rev_act_sum = float(detail_numeric_df.get("Revenue_as_per_generation", pd.Series(dtype=float)).sum())
    dsm_loss_sum = float(detail_numeric_df.get("Total_DSM", pd.Series(dtype=float)).sum())
    rev_loss_pct = 0.0 if rev_act_sum == 0 else (rev_loss_sum / rev_act_sum) * 100.0
    return {
        "plant_capacity_mode_AvC": avc_mode,
        "ppa_mode": ppa_mode,
        "revenue_loss_pct": rev_loss_pct,
        "dsm_loss": dsm_loss_sum,
    }

# ==========================================
# ---------------- LAYOUT ------------------
# ==========================================
# Reusable styles for consistent card-based UI
SECTION_CARD_STYLE = {
    "padding": "2rem",
    "backgroundColor": "#fff",
    "borderRadius": "12px",
    "marginBottom": "2rem",
    "boxShadow": "0 2px 8px rgba(0,0,0,0.05)",
    "border": "1px solid rgba(0,0,0,0.06)",
}
RESULTS_CARD_CLASS = "shadow-sm rounded-3 border-0"

def kpi_card(title: str, value: str) -> dbc.Card:
    return dbc.Card(
            dbc.CardBody([
            html.Div(title, className="text-muted", style={"fontSize": "0.9rem"}),
            html.H4(value, className="mt-1")
        ]),
        className="shadow-sm rounded-4"
    )

def sidebar():
    return html.Div([
        # Branding
        html.Div([
            html.Span("⚡", style={"fontSize": "1.5rem", "marginRight": "8px"}),
            html.Span("O2: DSM Analytics", style={"fontSize": "1.2rem", "fontWeight": "600", "color": "white"})
        ], style={"padding": "1.5rem 1rem", "borderBottom": "1px solid #444"}),
        
        # Navigation Menu
        html.Div([
            html.Div([
                html.Span("🏠", style={"marginRight": "12px"}),
                html.Span("Welcome", style={"color": "#ff6b35", "fontWeight": "500"})
            ], id="nav-welcome", className="nav-item active", style={
                "padding": "12px 20px", "cursor": "pointer", "borderLeft": "3px solid #ff6b35",
                "backgroundColor": "rgba(255, 107, 53, 0.1)", "marginBottom": "4px"
            }),
            html.Div([
                html.Span("⚙️", style={"marginRight": "12px"}),
                html.Span("Custom Settings", style={"color": "white", "fontWeight": "400"})
            ], id="nav-settings", className="nav-item", style={
                "padding": "12px 20px", "cursor": "pointer", "marginBottom": "4px"
            }),
            html.Div([
                html.Span("📊", style={"marginRight": "12px"}),
                html.Span("Analysis", style={"color": "white", "fontWeight": "400"})
            ], id="nav-analysis", className="nav-item", style={
                "padding": "12px 20px", "cursor": "pointer", "marginBottom": "4px"
            }),
            html.Div([
                html.Span("📊", style={"marginRight": "12px"}),
                html.Span("Aggregation Analysis", style={"color": "white", "fontWeight": "400"})
            ], id="nav-aggregation-analysis", className="nav-item", style={
                "padding": "12px 20px", "cursor": "pointer", "marginBottom": "4px"
            }),
            html.Div([
                html.Span("📤", style={"marginRight": "12px"}),
                html.Span("Custom Upload", style={"color": "white", "fontWeight": "400"})
            ], id="nav-custom-upload", className="nav-item", style={
                "padding": "12px 20px", "cursor": "pointer", "marginBottom": "4px"
            }),
            html.Div([
                html.Span("📈", style={"marginRight": "12px"}),
                html.Span("Data Statistics", style={"color": "white", "fontWeight": "400"})
            ], id="nav-stats", className="nav-item", style={
                "padding": "12px 20px", "cursor": "pointer", "marginBottom": "4px"
            }),
        ], style={"padding": "1rem 0"}),
        
        # Logout
        html.Div([
            html.Div([
                html.Span("🌙", style={"marginRight": "12px"}),
                html.Span("Logout", style={"color": "#ccc"})
            ], style={
                "padding": "12px 20px", "cursor": "pointer", "borderTop": "1px solid #444",
                "marginTop": "auto"
            })
        ], style={"position": "absolute", "bottom": "0", "left": "0", "right": "0"})
    ], style={
        "backgroundColor": "#1a1a1a", "height": "100vh", "width": "280px", "position": "fixed", 
        "left": "0", "top": "0", "zIndex": "1000", "display": "flex", "flexDirection": "column"
    })

def main_content():
    return html.Div([
        # Header (dynamic per selected sidebar section)
        html.Div([
            html.H2(id="page-title", style={"margin": "0", "fontWeight": "600", "color": "#333"}),
            html.P(id="page-subtitle", style={"margin": "0", "color": "#666", "fontSize": "0.9rem", "marginTop": "0.25rem"})
        ], style={"padding": "2rem 2.5rem 1.25rem", "borderBottom": "1px solid #eee", "backgroundColor": "#fff"}),
        
        # Content
        html.Div([
            # Welcome Dashboard Content
            html.Div([
                html.Div([
                    html.Div([
                        html.H1("Welcome to DSM Analytics Dashboard", style={
                            "fontSize": "2.5rem", "fontWeight": "700", "color": "#1a1a1a", 
                            "marginBottom": "1rem", "textAlign": "center"
                        }),
                        html.P("Real-time Deviation Settlement Mechanism Analytics", style={
                            "fontSize": "1.2rem", "color": "#666", "textAlign": "center", "marginBottom": "3rem"
                        }),
                    ], style={"marginBottom": "3rem"}),
                    
                    dbc.Row([
                        dbc.Col([
                            dbc.Card([
                                dbc.CardBody([
                                    html.Div("⚙️", style={"fontSize": "3rem", "marginBottom": "1rem"}),
                                    html.H4("Step 1: Configure Settings", style={"color": "#ff6b35", "fontWeight": "600"}),
                                    html.P("Go to Custom Settings to configure your analysis parameters, bands, and rates.", 
                                          style={"color": "#666", "lineHeight": "1.8"}),
                                    html.Ul([
                                        html.Li("Set Error% calculation mode"),
                                        html.Li("Define deviation bands and penalties"),
                                        html.Li("Save your custom settings for reuse")
                                    ], style={"textAlign": "left", "color": "#666"}),
                                ])
                            ], className="shadow-sm rounded-4", style={"height": "100%"})
                        ], md=4),
                        
                            dbc.Col([
                                dbc.Card([
                                    dbc.CardBody([
                                    html.Div("📊", style={"fontSize": "3rem", "marginBottom": "1rem"}),
                                    html.H4("Step 2: Run Analysis", style={"color": "#28a745", "fontWeight": "600"}),
                                    html.P("Navigate to Analysis to select region, plants, and date range for your analysis.", 
                                          style={"color": "#666", "lineHeight": "1.8"}),
                                    html.Ul([
                                        html.Li("Select region(s) from DuckDB"),
                                        html.Li("Choose specific plants or select all"),
                                        html.Li("Pick date range and plot results")
                                    ], style={"textAlign": "left", "color": "#666"}),
                                ])
                            ], className="shadow-sm rounded-4", style={"height": "100%"})
                            ], md=4),
                        
                            dbc.Col([
                                dbc.Card([
                                    dbc.CardBody([
                                    html.Div("📈", style={"fontSize": "3rem", "marginBottom": "1rem"}),
                                    html.H4("Step 3: View Results", style={"color": "#17a2b8", "fontWeight": "600"}),
                                    html.P("Analyze comprehensive DSM calculations, KPIs, and blockwise breakdown.", 
                                          style={"color": "#666", "lineHeight": "1.8"}),
                                    html.Ul([
                                        html.Li("View detailed KPIs and penalties"),
                                        html.Li("Explore blockwise analysis"),
                                        html.Li("Download full calculation in Excel")
                                    ], style={"textAlign": "left", "color": "#666"}),
                                ])
                            ], className="shadow-sm rounded-4", style={"height": "100%"})
                            ], md=4),
                    ], className="mb-4 g-4"),
                    
                    html.Div([
                        dbc.Card([
                            dbc.CardBody([
                                html.H5("🚀 Quick Start", style={"color": "#333", "marginBottom": "1rem"}),
                                html.P([
                                    "1. Click on ",
                                    html.Strong("Custom Settings", style={"color": "#ff6b35"}),
                                    " to configure your parameters and save settings"
                                ], style={"marginBottom": "0.5rem"}),
                                html.P([
                                    "2. Navigate to ",
                                    html.Strong("Analysis", style={"color": "#28a745"}),
                                    " to select data and run calculations"
                                ], style={"marginBottom": "0.5rem"}),
                                html.P([
                                    "3. Review results and download Excel report when needed"
                                ], style={"marginBottom": "0"}),
                            ])
                        ], className="shadow-sm rounded-4", style={"backgroundColor": "#f8f9fa"})
                    ], style={"marginTop": "3rem"}),
                    
                    html.Div([
                        dbc.Button([
                            html.Span("Get Started ", style={"marginRight": "8px"}),
                            html.Span("→")
                        ], id="btn-get-started", color="primary", size="lg", 
                                 style={"backgroundColor": "#ff6b35", "border": "none", "fontWeight": "600", "padding": "12px 40px"})
                    ], style={"textAlign": "center", "marginTop": "3rem"}),
                    
                ], style={"maxWidth": "1200px", "margin": "0 auto"})
            ], id="welcome-content", style={"display": "block"}),
            
            # Custom Settings Content
            html.Div([
                # Global Regulation Controls Section (Zero Basis Guard only)
                html.Div([
                    html.H5("Global Regulation Controls", style={"marginBottom": "1rem", "color": "#555"}),
                    dbc.Row([
                        dbc.Col([
                            html.Label("Zero Basis Guard", style={"fontWeight": "500", "marginBottom": "8px", "fontSize": "0.95rem", "color": "#555"}),
                            dbc.Checklist(
                                id="zero-basis-guard",
                                options=[{"label": "Skip night blocks (AvC≈0 & Schedule≈0)", "value": "on"}],
                                value=["on"],
                                inline=False,
                                style={"fontSize": "0.9rem"}
                            ),
                        ], md=12),
                    ], className="mb-3"),
                ], style={"marginBottom": "2rem", "padding": "1.5rem", "backgroundColor": "#f8f9fa", "borderRadius": "8px"}),
                
                # Error% Mode Section
                html.Div([
                    html.H5("Error% Mode", style={"marginBottom": "1rem", "color": "#555"}),
            dcc.RadioItems(
                        id="err-mode",
                options=[
                            {"label": "Default: (Actual - Scheduled) / AvC × 100", "value": "default"},
                            {"label": "Dynamic: 100×(Actual - Scheduled) / (X%·AvC + (100−X)%·Scheduled)", "value": "dynamic"},
                ],
                value="default",
                        inputStyle={"marginRight": "8px", "accentColor": "#ff6b35"},
                        labelStyle={"display": "block", "marginBottom": "12px", "color": "#666", "fontSize": "1rem"},
                    ),
                    html.Div([
                        html.Label(id="x-pct-label", children="X% (for Dynamic Error%)", 
                                  style={"color": "#555", "fontSize": "1rem", "marginBottom": "8px", "fontWeight": "500"}),
                        dcc.Slider(
                            id="x-pct", 
                        min=0,
                        max=100,
                            value=50, 
                        step=1,
                            marks={i: str(i) for i in range(0, 101, 10)},
                            tooltip={"placement": "bottom", "always_visible": True}
                        ),
                        html.Div(id="x-pct-readout", style={
                            "textAlign": "center", "color": "#ff6b35", "fontWeight": "bold", 
                            "fontSize": "1.2rem", "marginTop": "8px"
                        }),
                    ], id="x-pct-container", style={"display": "none", "marginTop": "1rem"}),
                ], style={"marginBottom": "2rem", "padding": "1.5rem", "backgroundColor": "#f8f9fa", "borderRadius": "8px"}),
                
                # User-defined Bands & Rates Section
                html.Div([
                    html.H5("User-defined Bands & Rates", style={"marginBottom": "1rem", "color": "#555"}),
                    
                    
                    # Add New Row Form
                    dbc.Card([
                        dbc.CardHeader([
                            html.H6("Add New Band", style={"margin": "0", "color": "#555", "fontWeight": "600"})
                        ]),
                        dbc.CardBody([
                            dbc.Row([
                                # Direction
                                dbc.Col([
                                    html.Label("Direction", style={"fontWeight": "500", "marginBottom": "4px", "fontSize": "0.9rem"}),
            dcc.Dropdown(
                                        id="form-direction",
                                        options=BANDS_DROPDOWNS["direction"]["options"],
                                        value="UI",
                clearable=False,
                                        style={"fontSize": "0.9rem"}
                                    ),
                                ], md=2),
                                
                                # Lower %
                                dbc.Col([
                                    html.Label("Lower %", style={"fontWeight": "500", "marginBottom": "4px", "fontSize": "0.9rem"}),
                                    dbc.Input(
                                        id="form-lower-pct",
                                        type="number",
                                        value=0.0,
                                        step=0.1,
                                        min=0,
                                        style={"fontSize": "0.9rem"}
                                    ),
                                ], md=1),
                                
                                # Upper %
                                dbc.Col([
                                    html.Label("Upper %", style={"fontWeight": "500", "marginBottom": "4px", "fontSize": "0.9rem"}),
                                    dbc.Input(
                                        id="form-upper-pct",
                                        type="number",
                                        value=10.0,
                                        step=0.1,
                                        min=0,
                                        style={"fontSize": "0.9rem"}
                                    ),
                                ], md=1),
                # Tolerance Cut % (hidden)
                dbc.Col([
                    html.Label("Tolerance Cut %", style={"fontWeight": "500", "marginBottom": "4px", "fontSize": "0.9rem"}),
                    dbc.Input(
                        id="form-tolerance-cut-pct",
                        type="number",
                        value=10.0,
                        step=0.1,
                        min=0,
                        style={"fontSize": "0.9rem"}
                    ),
                ], md=1, style={"display": "none"}),
                
                # Loss Zone
                dbc.Col([
                    html.Label("Loss Zone", style={"fontWeight": "500", "marginBottom": "4px", "fontSize": "0.9rem"}),
                    dcc.Dropdown(
                        id="form-loss-zone",
                        options=BANDS_DROPDOWNS["loss_zone"]["options"],
                        value=False,
                        clearable=False,
                        style={"fontSize": "0.9rem"}
                    ),
                ], md=2),
                            ], className="mb-2"),
                            
                            dbc.Row([
                                # Rate Type
                                dbc.Col([
                                    html.Label("Rate Type", style={"fontWeight": "500", "marginBottom": "4px", "fontSize": "0.9rem"}),
                                    dcc.Dropdown(
                                        id="form-rate-type",
                                        options=BANDS_DROPDOWNS["rate_type"]["options"],
                                        value="flat_per_kwh",
                                        clearable=False,
                                        style={"fontSize": "0.9rem"}
                                    ),
                                ], md=2),
                                
                                # Rate Value
                                dbc.Col([
                                    html.Label("Rate Value", style={"fontWeight": "500", "marginBottom": "4px", "fontSize": "0.9rem"}),
                                    dbc.Input(
                                        id="form-rate-value",
                                        type="number",
                                        value=0.0,
                                        step=0.01,
                                        min=0,
                                        style={"fontSize": "0.9rem"}
                                    ),
                                ], md=1),
                                
                                # Excess Slope %
                                dbc.Col([
                                    html.Label("Excess Slope %", style={"fontWeight": "500", "marginBottom": "4px", "fontSize": "0.9rem"}),
                                    dbc.Input(
                                        id="form-excess-slope-per-pct",
                                        type="number",
                                        value=0.0,
                                        step=0.01,
                                        min=0,
                                        style={"fontSize": "0.9rem"}
                                    ),
                                ], md=1),
                                
                                # Apply To removed (routing is via Loss Zone + direction)
                                
                                # Add Button
                                dbc.Col([
                                    html.Label("&nbsp;", style={"marginBottom": "4px"}),  # Spacer
        dbc.Button(
                                        "Add to Table", 
                                        id="add-from-form", 
                                        color="success", 
                                        size="sm", 
                                        className="w-100",
                                        style={"fontSize": "0.9rem"}
                                    ),
                                ], md=2),
                            ], className="mb-2"),
                            
                            # Auto-generated Label Preview
                dbc.Row([
                                dbc.Col([
                                    html.Label("Preview Label:", style={"fontWeight": "500", "marginBottom": "4px", "fontSize": "0.9rem"}),
                                    html.Div(
                                        id="form-label-preview",
                                        style={
                                            "padding": "8px", 
                                            "backgroundColor": "#f8f9fa", 
                                            "border": "1px solid #dee2e6", 
                                            "borderRadius": "4px",
                                            "fontSize": "0.9rem",
                                            "color": "#495057"
                                        }
                                    ),
                                ], md=12),
                            ]),
                        ]),
                    ], className="mb-3"),
                    
        dash_table.DataTable(
                        id="bands-table",
                        columns=BANDS_COLUMNS,
                        data=DEFAULT_BANDS.copy(),
                        editable=True,
                        row_deletable=True,
                        sort_action="native",
                        page_size=8,
                        style_table={"overflow": "visible", "overflowX": "auto", "overflowY": "visible"},
                        style_cell={"padding": "12px", "fontSize": "0.9rem"},
                        style_header={"backgroundColor": "#f8f9fa", "fontWeight": "600"},
                        dropdown={
                            "direction": {**BANDS_DROPDOWNS["direction"], "clearable": False},
                            "rate_type": {**BANDS_DROPDOWNS["rate_type"], "clearable": False},
                            "loss_zone": {**BANDS_DROPDOWNS["loss_zone"], "clearable": False},
                        },
                    ),
                    html.Div([
                        dbc.Button("Reset Table", id="reset-bands", size="sm", className="mt-3 me-2",
                                 style={"backgroundColor": "#6c757d", "border": "none"}),
                        dcc.Upload(
                            id="upload-bands",
                            children=dbc.Button("Load from File", size="sm", className="mt-3 me-2",
                                             style={"backgroundColor": "#17a2b8", "border": "none"}),
                            multiple=False
                        ),
                        dbc.Button("Save to File", id="save-bands", size="sm", className="mt-3",
                                 style={"backgroundColor": "#28a745", "border": "none"}),
                        dcc.Download(id="download-bands-json"),
                    ]),
                ], style={"padding": "1.5rem", "backgroundColor": "#f8f9fa", "borderRadius": "8px", "marginBottom": "2rem"}),
                
                # Save Settings Section
                html.Div([
                    html.H5("Save Configuration", style={"marginBottom": "1rem", "color": "#555"}),
                    html.P("Save your current settings to apply them when running analysis.", style={"color": "#666", "marginBottom": "1rem"}),
                    dbc.Button([
                        html.Span("💾 ", style={"marginRight": "8px"}),
                        "Save Settings"
                    ], id="btn-save-settings", color="success", size="lg", className="w-100",
                             style={"fontWeight": "500", "padding": "12px"}),
                    html.Div(id="settings-save-message", style={"marginTop": "1rem"}),
                    # Preset controls
                    dbc.Row([
                        dbc.Col([
                            html.Label("Preset Name", style={"fontWeight": "600"}),
                            dbc.Input(id="preset-name", placeholder="e.g., A – Default, B – Dynamic X=50", type="text"),
                        ], md=6),
                        dbc.Col([
                            html.Label(" "),
                            dbc.Button("Save as Preset", id="btn-save-preset", color="secondary", className="w-100"),
                        ], md=3),
                        dbc.Col([
                            html.Label(" "),
                            dbc.Button("Delete Selected Preset(s)", id="btn-delete-preset", color="danger", outline=True, className="w-100"),
                        ], md=3),
                    ], className="mt-3"),
                    html.Div(id="preset-save-message", className="mt-2"),
                    ], style={"padding": "1.5rem", "backgroundColor": "#f8f9fa", "borderRadius": "8px"}),
            ], id="custom-settings-content", style={"display": "none"}),
            
            # Analysis Content
            html.Div([
                # Analysis intro (header handled globally)
                html.Div([]),
                
                # Global Filters Section
                html.Div([
                    html.Div([
                        html.Span("🔍", style={"fontSize": "1.5rem", "marginRight": "10px"}),
                        html.Span("Data Selection", style={"fontSize": "1.1rem", "fontWeight": "600", "color": "#333"})
                    ], style={"marginBottom": "1.5rem"}),
                    
                        dbc.Row([
                            dbc.Col([
                                html.Label([
                                html.Span("🌐 ", style={"marginRight": "6px"}),
                                "Region"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.Loading(
                                id="region-loading",
                                type="circle",
                                children=dcc.Dropdown(
                                    id="region-dd",
                                    options=[],
                                    value=[],
                                    multi=True,
                                    clearable=False,
                                    placeholder="Select Region(s) from DuckDB...",
                                    style={"fontSize": "0.95rem"}
                                ),
                                ),
                            ], md=4),
                            dbc.Col([
                                html.Label([
                                html.Span("🏭 ", style={"marginRight": "6px"}),
                                "Resource"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                                dcc.Dropdown(
                                    id="resource-type-dd",
                                    options=[
                                        {"label": "All", "value": "ALL"},
                                        {"label": "Solar", "value": "SOLAR"},
                                        {"label": "Wind", "value": "WIND"},
                                        {"label": "Thermal", "value": "THERMAL"},
                                    ],
                                    value="ALL",
                                    clearable=False,
                                    style={"fontSize": "0.95rem"}
                                ),
                            ], md=2),
                            dbc.Col([
                                html.Label([
                                html.Span("🏭 ", style={"marginRight": "6px"}),
                                "Plant"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                                dcc.Loading(
                                    id="plant-loading",
                                    type="circle",
                                    children=dcc.Dropdown(
                                        id="plant-dd",
                                        options=[],
                                        value=[],
                                        multi=True,
                                        clearable=False,
                                    placeholder="Select Plant(s) or Select All...",
                                    style={"fontSize": "0.95rem"},
                                        optionHeight=35,
                                        maxHeight=400,
                                        searchable=True,
                                    ),
                                ),
                            ], md=2),
                            dbc.Col([
                                html.Label([
                                html.Span("📅 ", style={"marginRight": "6px"}),
                                "Date Range"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                                dcc.DatePickerRange(
                                    id="date-range",
                                    start_date=(datetime.now() - timedelta(days=6)).date(),
                                    end_date=datetime.now().date(),
                                    display_format="DD-MMM-YYYY",
                                    minimum_nights=0,
                                style={"fontSize": "0.95rem"}
                                ),
                            ], md=4),
                        ], className="mb-4"),

                        # Optional QCA filter (STU + RPC)
                        dbc.Row([
                            dbc.Col([
                                html.Label([
                                    html.Span("🧑‍💼 ", style={"marginRight": "6px"}),
                                    "QCA (Optional)"
                                ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                                dcc.Loading(
                                    id="qca-loading",
                                    type="circle",
                                    children=dcc.Dropdown(
                                        id="qca-dd",
                                        options=[],
                                        value=[],
                                        multi=True,
                                        clearable=True,
                                        placeholder="(Optional) Filter by QCA...",
                                        style={"fontSize": "0.95rem"},
                                        optionHeight=35,
                                        maxHeight=300,
                                        searchable=True,
                                    ),
                                ),
                            ], md=4),
                        ], className="mb-4"),
                        # Preset selection (optional) — legacy compare-presets mode
                        dbc.Row([
                            dbc.Col([
                                html.Label("Custom Setting(s)", style={"fontWeight": 600, "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                                dcc.Dropdown(
                                    id="analysis-preset-select",
                                    multi=True,
                                    placeholder="(Optional) Compare presets across all plants",
                                ),
                                html.Small("Compare same plants with different presets", className="text-muted d-block mt-1", style={"fontSize": "0.8rem"}),
                            ], md=12),
                        ], className="mb-2"),
                        # Add-more config rows (Region, Resource, Plant, QCA, Custom Setting per row)
                        dbc.Row([
                            dbc.Col([
                                html.Div([
                                    html.Label("Analysis configs (add multiple to compare)", style={"fontWeight": 600, "marginBottom": "8px", "fontSize": "0.95rem"}),
                                    dbc.Button("➕ Add more", id="analysis-add-config", color="secondary", size="sm", className="mb-2"),
                                    html.Div(id="analysis-config-container", children=[]),
                                ]),
                            ], md=12),
                        ], className="mb-3"),
                        
                        dbc.Row([
                            dbc.Col([
                                dbc.Button([
                                    html.Span("📊 ", style={"marginRight": "8px"}),
                                "Run Analysis"
                                ], id="plot-now", size="lg", className="w-100", disabled=True,
                                     style={"backgroundColor": "#ff6b35", "border": "none", "fontWeight": "600", "padding": "14px",
                                       "boxShadow": "0 4px 6px rgba(255, 107, 53, 0.3)"}),
                            ], md=6, className="offset-md-3"),
                        ]),
                    # Horizontal progress bar while computing
                        dbc.Row([
                            dbc.Col([
                            html.Div(id="progress-container", style={"display": "none", "marginTop": "1.5rem"}, children=[
                                html.Div([
                                    html.Span("⏳ ", style={"marginRight": "8px"}),
                                    html.Span("Processing data...", style={"fontWeight": "500", "color": "#333"})
                                ], style={"marginBottom": "10px", "textAlign": "center"}),
                                dbc.Progress(
                                    id="compute-progress",
                                    value=100,
                                    striped=True,
                                    animated=True,
                                    style={"height": "25px"},
                                    className="mb-2"
                                ),
                            ]),
                        ], md=8, className="offset-md-2")
                    ]),
                ], style=SECTION_CARD_STYLE),
                
                # Results Section (card-wrapped)
                html.Div(id="results-section", style={"display": "none", "marginBottom": "2rem"}, children=[
                    dbc.Card([
                        dbc.CardBody([
                            html.Div(id="tab-content", className="mt-0"),
                        ], className="p-4"),
                    ], className=RESULTS_CARD_CLASS),
                ]),
            ], id="analysis-content", style={"display": "none"}),
            # Aggregation Analysis Content
            html.Div([
                # Aggregation Analysis intro (header handled globally)
                html.Div([]),
                # Global Filters Section (mirrors Analysis)
                html.Div([
                    html.Div([
                        html.Span("🔍", style={"fontSize": "1.5rem", "marginRight": "10px"}),
                        html.Span("Data Selection", style={"fontSize": "1.1rem", "fontWeight": "600", "color": "#333"})
                    ], style={"marginBottom": "1.5rem"}),
                    dbc.Row([
                        dbc.Col([
                            html.Label([
                                html.Span("🌐 ", style={"marginRight": "6px"}),
                                "Region"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.Loading(
                                id="agg-region-loading",
                                type="circle",
                                children=dcc.Dropdown(
                                    id="agg-region-dd",
                                    options=[],
                                    value=[],
                                    multi=True,
                                    clearable=False,
                                    placeholder="Select Region(s) from DuckDB...",
                                    style={"fontSize": "0.95rem"}
                                ),
                            ),
                        ], md=4),
                        dbc.Col([
                            html.Label([
                                html.Span("🏭 ", style={"marginRight": "6px"}),
                                "Resource"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.Dropdown(
                                id="agg-resource-type-dd",
                                options=[
                                    {"label": "All", "value": "ALL"},
                                    {"label": "Solar", "value": "SOLAR"},
                                    {"label": "Wind", "value": "WIND"},
                                    {"label": "Thermal", "value": "THERMAL"},
                                ],
                                value="ALL",
                                clearable=False,
                                style={"fontSize": "0.95rem"}
                            ),
                        ], md=2),
                        dbc.Col([
                            html.Label([
                                html.Span("🏭 ", style={"marginRight": "6px"}),
                                "Plant"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.Loading(
                                id="agg-plant-loading",
                                type="circle",
                                children=dcc.Dropdown(
                                    id="agg-plant-dd",
                                    options=[],
                                    value=[],
                                    multi=True,
                                    clearable=False,
                                    placeholder="Select Plant(s) or Select All...",
                                    style={"fontSize": "0.95rem"},
                                    optionHeight=35,
                                    maxHeight=400,
                                    searchable=True,
                                ),
                            ),
                        ], md=2),
                        dbc.Col([
                            html.Label([
                                html.Span("📅 ", style={"marginRight": "6px"}),
                                "Date Range"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.DatePickerRange(
                                id="agg-date-range",
                                start_date=(datetime.now() - timedelta(days=6)).date(),
                                end_date=datetime.now().date(),
                                display_format="DD-MMM-YYYY",
                                minimum_nights=0,
                                style={"fontSize": "0.95rem"}
                            ),
                        ], md=4),
                    ], className="mb-4"),

                    # Optional QCA filter (STU + RPC)
                    dbc.Row([
                        dbc.Col([
                            html.Label([
                                html.Span("🧑‍💼 ", style={"marginRight": "6px"}),
                                "QCA (Optional)"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.Loading(
                                id="agg-qca-loading",
                                type="circle",
                                children=dcc.Dropdown(
                                    id="agg-qca-dd",
                                    options=[],
                                    value=[],
                                    multi=True,
                                    clearable=True,
                                    placeholder="(Optional) Filter by QCA...",
                                    style={"fontSize": "0.95rem"},
                                    optionHeight=35,
                                    maxHeight=300,
                                    searchable=True,
                                ),
                            ),
                        ], md=4),
                    ], className="mb-4"),
                    # Exclude Plants Section (appears only when SELECT_ALL is active)
                    html.Div([
                        dbc.Row([
                            dbc.Col([
                                html.Div([
                                    html.Label([
                                        html.Span("❌ ", style={"marginRight": "6px", "color": "#dc3545"}),
                                        "Exclude Plants from Aggregation (Optional)",
                                        html.Small(" - Only available when 'Select All' is active", 
                                                  style={"marginLeft": "8px", "color": "#666", "fontWeight": "normal", "fontSize": "0.85rem"})
                                    ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                                    dcc.Loading(
                                        id="agg-exclude-plant-loading",
                                        type="circle",
                                        children=dcc.Dropdown(
                                            id="agg-exclude-plant-dd",
                                            options=[],
                                            value=[],
                                            multi=True,
                                            clearable=True,
                                            placeholder="Select plants to exclude from aggregation...",
                                            style={"fontSize": "0.95rem"},
                                            optionHeight=35,
                                            maxHeight=300,
                                            searchable=True,
                                        ),
                                    ),
                                ], style={
                                    "padding": "1rem",
                                    "backgroundColor": "#fff3cd",
                                    "border": "1px solid #ffc107",
                                    "borderRadius": "8px",
                                    "marginBottom": "1rem"
                                })
                            ], md=12),
                        ], className="mb-3"),
                    ], id="agg-exclude-section", style={"display": "none"}),
                    # Custom Settings (presets) selection
                    dbc.Row([
                        dbc.Col([
                            html.Label("Custom Setting(s)", style={"fontWeight": 600, "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.Dropdown(
                                id="agg-analysis-preset-select",
                                multi=True,
                                placeholder="(Optional) Compare presets across aggregated profile",
                            ),
                            html.Small("Compare aggregated profile with different presets", className="text-muted d-block mt-1", style={"fontSize": "0.8rem"}),
                        ], md=6),
                        dbc.Col([
                            dbc.Checklist(
                                id="agg-per-plant-toggle",
                                options=[{"label": " Per-plant settings (each plant with its preset)", "value": "on"}],
                                value=[],
                                inline=True,
                                style={"fontSize": "0.9rem"},
                            ),
                        ], md=6, className="d-flex align-items-end pb-2"),
                    ], className="mb-3"),
                    html.Div([
                        dbc.Row([
                            dbc.Col([
                                html.Label("Assign preset per plant (runs each plant separately)", style={"fontWeight": 600, "marginBottom": "8px", "fontSize": "0.9rem"}),
                                dash_table.DataTable(
                                    id="agg-plant-preset-table",
                                    columns=[
                                        {"name": "Plant", "id": "Plant", "editable": False, "type": "text"},
                                        {"name": "Custom Setting", "id": "Preset", "editable": True, "presentation": "dropdown"},
                                    ],
                                    data=[],
                                    dropdown={"Preset": {"options": [{"label": "Current (form bands)", "value": "__CURRENT__"}]}},
                                    style_table={"overflowX": "auto"},
                                    style_cell={"padding": "8px", "fontSize": "0.9rem"},
                                    style_header={"fontWeight": "600"},
                                    row_deletable=False,
                                ),
                            ], md=12),
                        ], className="mb-3"),
                    ], id="agg-per-plant-section", style={"display": "none"}),
                    # PPA Configuration
                    html.Div([
                        html.Div([
                            html.Span("⚡", style={"fontSize": "1.3rem", "marginRight": "8px"}),
                            html.Span("PPA Configuration", style={"fontSize": "1.05rem", "fontWeight": "600", "color": "#333"})
                        ], style={"marginBottom": "0.75rem"}),
                        dbc.Row([
                            dbc.Col([
                                dcc.Dropdown(
                                    id="agg-ppa-mode",
                                    options=[
                                        {"label": "Mean", "value": "mean"},
                                        {"label": "Median", "value": "median"},
                                        {"label": "Mode", "value": "mode"},
                                        {"label": "Weighted Average", "value": "weighted"},
                                        {"label": "Numeric", "value": "numeric"},
                                    ],
                                    value=["mean"],
                                    multi=True,
                                    clearable=False,
                                    placeholder="Select one or more PPA methods...",
                                    style={"fontSize": "0.95rem"},
                                ),
                            ], md=6),
                            dbc.Col([
                                html.Div([
                                    html.Label("Numeric PPA (₹/kWh)", style={"fontWeight": "500", "marginBottom": "6px", "fontSize": "0.9rem", "color": "#555"}),
                                    dbc.Input(
                                        id="agg-ppa-value",
                                        type="number",
                                        value=None,
                                        min=0,
                                        step=0.01,
                                        placeholder="Enter PPA value",
                                        style={"fontSize": "0.9rem"}
                                    ),
                                ], id="agg-ppa-numeric-container", style={"display": "none"}),
                            ], md=6),
                        ]),
                    ], style={"marginBottom": "1.5rem", "padding": "1.25rem", "backgroundColor": "#f8f9fa", "borderRadius": "8px"}),
                    # Action buttons
                    dbc.Row([
                        dbc.Col([
                            dbc.Button([
                                html.Span("📊 ", style={"marginRight": "8px"}),
                                "Plot Now"
                            ], id="agg-plot-now", size="md", className="w-100", disabled=True,
                               style={"backgroundColor": "#ff6b35", "border": "none", "fontWeight": "600", "padding": "10px",
                                      "boxShadow": "0 4px 6px rgba(255, 107, 53, 0.3)"}),
                        ], md=4, className="offset-md-4"),
                    ]),
                    # Progress bar
                    dbc.Row([
                        dbc.Col([
                            html.Div(id="agg-progress-container", style={"display": "none", "marginTop": "1.5rem"}, children=[
                                html.Div([
                                    html.Span("⏳ ", style={"marginRight": "8px"}),
                                    html.Span("Processing aggregated data...", style={"fontWeight": "500", "color": "#333"})
                                ], style={"marginBottom": "10px", "textAlign": "center"}),
                                dbc.Progress(
                                    id="agg-compute-progress",
                                    value=100,
                                    striped=True,
                                    animated=True,
                                    style={"height": "25px"},
                                    className="mb-2"
                                ),
                            ]),
                        ], md=8, className="offset-md-2")
                    ]),
                ], style=SECTION_CARD_STYLE),
                # Results Section (card-wrapped)
                html.Div(id="agg-results-section", style={"display": "none", "marginBottom": "2rem"}, children=[
                    dbc.Card([
                        dbc.CardBody([
                            html.Div(id="agg-tab-content", className="mt-0"),
                        ], className="p-4"),
                    ], className=RESULTS_CARD_CLASS),
                ]),
            ], id="aggregation-analysis-content", style={"display": "none"}),
            # Custom Upload Content
            html.Div([
                # Custom Upload intro (header handled globally)
                html.Div([]),
                # Download sample buttons
                html.Div([
                    dbc.Row([
                        dbc.Col([
                            dbc.Button("Download XLSX Sample", id="btn-download-sample-xlsx", color="secondary", className="me-2"),
                            dbc.Button("Download CSV Sample", id="btn-download-sample-csv", color="secondary"),
                        ], md=12),
                    ], className="mb-3"),
                ], style={**SECTION_CARD_STYLE, "marginBottom": "1rem"}),
                dcc.Download(id="dl-sample-xlsx"),
                dcc.Download(id="dl-sample-csv"),

                # Upload control and preset select
                html.Div([
                    dcc.Upload(
                        id="upload-custom-file",
                        children=html.Div(["Drag & drop or ", html.A("choose a CSV/XLSX")]),
                        multiple=False,
                        style={"border":"1px dashed #bbb", "padding":"16px", "borderRadius":"8px"}
                    ),
                    html.Div(id="custom-upload-validate", className="mt-2"),

                    html.Label("Custom Setting(s)", className="mt-3", style={"fontWeight": 600, "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                    dcc.Dropdown(id="custom-upload-preset-select", multi=True, placeholder="Choose saved preset(s)"),
                    dbc.Button("Run Analysis", id="btn-run-custom", color="primary", className="mt-3", disabled=True),
                ], style={**SECTION_CARD_STYLE, "marginBottom": "1rem"}),

                # Results and download (card-wrapped)
                html.Div([
                    dbc.Card([
                        dbc.CardBody([
                            html.Div(id="custom-results"),
                            dbc.Button("Download Output (Excel)", id="download-custom-output-btn", className="mt-3", color="success"),
                            dcc.Download(id="download-custom-output"),
                        ], className="p-4"),
                    ], className=RESULTS_CARD_CLASS),
                ], style=SECTION_CARD_STYLE),
            ], id="custom-upload-content", style={"display": "none"}),
            # Data Statistics Content
            html.Div([
                # Data Statistics intro (header handled globally)
                html.Div([]),
                html.Div([
                    dbc.Row([
                        dbc.Col([
                            html.Label([html.Span("🌐 ", style={"marginRight": "6px"}), "Region"], 
                                       style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.Dropdown(id="stats-region-dd", options=[], value=[], multi=True, clearable=False,
                                         placeholder="Select Region(s) from DuckDB...", style={"fontSize": "0.95rem"}),
                        ], md=4),
                        dbc.Col([
                            html.Label([html.Span("🏭 ", style={"marginRight": "6px"}), "Resource"], 
                                       style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.Dropdown(id="stats-resource-dd", options=[
                                {"label": "All", "value": "ALL"},
                                {"label": "Solar", "value": "SOLAR"},
                                {"label": "Wind", "value": "WIND"},
                            ], value="ALL", clearable=False, style={"fontSize": "0.95rem"}),
                        ], md=2),
                        dbc.Col([
                            html.Label([html.Span("🏭 ", style={"marginRight": "6px"}), "Plant"], 
                                       style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.Dropdown(id="stats-plant-dd", options=[], value=[], multi=True, clearable=False,
                                         placeholder="Select Plant(s) or Select All...", style={"fontSize": "0.95rem"}),
                        ], md=4),
                        dbc.Col([
                            html.Label([html.Span("📅 ", style={"marginRight": "6px"}), "Date Range"], 
                                       style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.DatePickerRange(id="stats-date-range", start_date=(datetime.now()-timedelta(days=6)).date(),
                                                end_date=datetime.now().date(), display_format="DD-MMM-YYYY", minimum_nights=0,
                                                style={"fontSize": "0.95rem"}),
                        ], md=2),
                    ], className="mb-3"),

                    # Optional QCA filter (STU + RPC)
                    dbc.Row([
                        dbc.Col([
                            html.Label([html.Span("🧑‍💼 ", style={"marginRight": "6px"}), "QCA (Optional)"],
                                       style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.Dropdown(
                                id="stats-qca-dd",
                                options=[],
                                value=[],
                                multi=True,
                                clearable=True,
                                placeholder="(Optional) Filter by QCA...",
                                style={"fontSize": "0.95rem"},
                            ),
                        ], md=4),
                    ], className="mb-3"),
                    dbc.Row([
                        dbc.Col([
                            dbc.Button([html.Span("📈 ", style={"marginRight": "8px"}), "Run"], id="stats-run", size="lg", className="w-100",
                                       style={"backgroundColor": "#ff6b35", "border": "none", "fontWeight": "600", "padding": "14px"}),
                        ], md=6, className="offset-md-3")
                    ]),
                ], style=SECTION_CARD_STYLE),
                html.Div([
                    dbc.Card([
                        dbc.CardBody([
                            html.Div(id="stats-results"),
                        ], className="p-4"),
                    ], className=RESULTS_CARD_CLASS),
                ], style={"marginBottom": "2rem"}),
            ], id="stats-content", style={"display": "none"}),
            
        ], style={"padding": "2rem 2.5rem 3rem"})
    ], style={"marginLeft": "280px", "minHeight": "100vh", "backgroundColor": "#f8f9fa"})

app.layout = html.Div([
    dcc.Store(id="results-store"),
    dcc.Store(id="agg-results-store"),
    dcc.Store(id="nav-store", data="welcome"),
    dcc.Store(id="saved-settings-store", storage_type="local"),
    dcc.Store(id="presets-store", storage_type="local"),
    dcc.Store(id="analysis-config-store", data=[{"region": None, "resource": "ALL", "plant": None, "qca": None, "preset": "__CURRENT__"}]),
    dcc.Store(id="custom-upload-store"),
    dcc.Store(id="custom-results-store"),
    dcc.Interval(id="progress-interval", interval=400, n_intervals=0, disabled=True),
    sidebar(),
    main_content(),
    dcc.Download(id="download-excel"),
])

# ==========================================
# ---------------- CALLBACKS ---------------
# ==========================================
@app.callback(
    Output("nav-welcome", "style"),
    Output("nav-settings", "style"),
    Output("nav-analysis", "style"),
    Output("nav-aggregation-analysis", "style"),
    Output("nav-custom-upload", "style"),
    Output("nav-stats", "style"),
    Output("welcome-content", "style"),
    Output("custom-settings-content", "style"),
    Output("analysis-content", "style"),
    Output("aggregation-analysis-content", "style"),
    Output("custom-upload-content", "style"),
    Output("stats-content", "style"),
    Output("nav-store", "data"),
    Input("nav-welcome", "n_clicks"),
    Input("nav-settings", "n_clicks"),
    Input("nav-analysis", "n_clicks"),
    Input("nav-aggregation-analysis", "n_clicks"),
    Input("nav-custom-upload", "n_clicks"),
    Input("nav-stats", "n_clicks"),
    Input("btn-get-started", "n_clicks"),
    State("nav-store", "data"),
)
def switch_nav_tabs(welcome_clicks, settings_clicks, analysis_clicks, agg_analysis_clicks, custom_upload_clicks, nav_stats, get_started_clicks, current):
    ctx_triggered = ctx.triggered_id
    target = current or "welcome"

    if ctx_triggered == "nav-welcome":
        target = "welcome"
    elif ctx_triggered == "nav-settings":
        target = "settings"
    elif ctx_triggered == "nav-analysis":
        target = "analysis"
    elif ctx_triggered == "nav-aggregation-analysis":
        target = "aggregation_analysis"
    elif ctx_triggered == "nav-custom-upload":
        target = "custom_upload"
    elif ctx_triggered == "nav-stats":
        target = "stats"
    elif ctx_triggered == "btn-get-started":
        target = "settings"

    # Define nav item styles
    active_style = {
        "padding": "12px 20px", "cursor": "pointer", "borderLeft": "3px solid #ff6b35",
        "backgroundColor": "rgba(255, 107, 53, 0.1)", "marginBottom": "4px",
        "color": "#ff6b35", "fontWeight": "500"
    }
    inactive_style = {
        "padding": "12px 20px", "cursor": "pointer", "marginBottom": "4px", "color": "white"
    }

    if target == "welcome":
        return (
            active_style,
            inactive_style,
            inactive_style,
            inactive_style,
            inactive_style,
            inactive_style,
            {"display": "block"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            target,
        )
    elif target == "settings":
        return (
            inactive_style,
            active_style,
            inactive_style,
            inactive_style,
            inactive_style,
            inactive_style,
            {"display": "none"},
            {"display": "block"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            target,
        )
    elif target == "analysis":
        return (
            inactive_style,
            inactive_style,
            active_style,
            inactive_style,
            inactive_style,
            inactive_style,
            {"display": "none"},
            {"display": "none"},
            {"display": "block"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            target,
        )
    elif target == "aggregation_analysis":
        return (
            inactive_style,
            inactive_style,
            inactive_style,
            active_style,
            inactive_style,
            inactive_style,
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "block"},
            {"display": "none"},
            {"display": "none"},
            target,
        )
    elif target == "custom_upload":
        return (
            inactive_style,
            inactive_style,
            inactive_style,
            inactive_style,
            active_style,
            inactive_style,
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "block"},
            {"display": "none"},
            target,
        )
    else:  # stats
        return (
            inactive_style,
            inactive_style,
            inactive_style,
            inactive_style,
            inactive_style,
            active_style,
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "block"},
            target,
        )


@app.callback(
    Output("page-title", "children"),
    Output("page-subtitle", "children"),
    Input("nav-store", "data"),
)
def update_page_header(nav_key):
    """Set main page header based on current sidebar selection."""
    nav_key = nav_key or "welcome"
    title_map = {
        "welcome": "Welcome",
        "settings": "Custom Settings",
        "analysis": "Analysis",
        "aggregation_analysis": "Aggregation Analysis",
        "custom_upload": "Custom Upload",
        "stats": "Data Statistics",
    }
    subtitle_map = {
        "welcome": "Start here to understand how to use the DSM analytics tools.",
        "settings": "Configure error modes and DSM deviation bands.",
        "analysis": "Run DSM deviation analysis for selected plants.",
        "aggregation_analysis": "Aggregate multiple plants, then run DSM analysis on the combined profile.",
        "custom_upload": "Upload your own DSM dataset and apply saved presets.",
        "stats": "Compute availability statistics for selected plants.",
    }
    return title_map.get(nav_key, "Welcome"), subtitle_map.get(nav_key, "")

# Callback to load regions from DuckDB on page load
@app.callback(
    Output("region-dd", "options"),
    Input("nav-store", "data"),
    prevent_initial_call=False
)
def load_regions_from_duckdb(nav_data):
    """Load regions from DuckDB databases"""
    try:
        regions = get_regions_from_duckdb()
        return [{"label": r, "value": r} for r in regions]
    except Exception as e:
        print(f"Error loading regions: {e}")
        return []


@app.callback(
    Output("agg-region-dd", "options"),
    Input("nav-store", "data"),
    prevent_initial_call=False
)
def load_regions_for_agg(nav_data):
    """Load regions for Aggregation Analysis from DuckDB databases."""
    try:
        regions = get_regions_from_duckdb()
        return [{"label": r, "value": r} for r in regions]
    except Exception as e:
        print(f"Error loading regions (agg): {e}")
        return []

# Regions for stats page
@app.callback(
    Output("stats-region-dd", "options"),
    Input("nav-store", "data"),
    prevent_initial_call=False
)
def load_regions_stats(nav_data):
    try:
        regions = get_regions_from_duckdb()
        return [{"label": r, "value": r} for r in regions]
    except Exception as e:
        print(f"Error loading regions: {e}")
        return []

@app.callback(
    Output("plant-dd", "options"),
    Output("plant-dd", "value"),
    Input("region-dd", "value"),
    Input("resource-type-dd", "value"),
    Input("date-range", "start_date"),
    Input("date-range", "end_date"),
    State("plant-dd", "value"),
    prevent_initial_call=False
)
def update_plants_from_duckdb(regions, resource_type, start_date, end_date, current_plant_value):
    """Update plant options based on selected regions and resource type.
    Only resets plant selection when region changes, not when date range changes.
    """
    if not regions:
        return [], []

    # Filter by type
    try:
        sd = str(start_date) if start_date else str(datetime.now().date())
        ed = str(end_date) if end_date else sd
        filtered_plants = get_filtered_plants_by_type(regions, resource_type or "ALL", sd, ed)
    except Exception:
        filtered_plants = get_plants_from_duckdb(regions)

    if not filtered_plants:
        return [], []

    def truncate_label(name, max_length=50):
        return name if len(name) <= max_length else name[:max_length-3] + "..."

    options = [{"label": "✓ Select All", "value": "SELECT_ALL"}] + [
        {"label": truncate_label(p), "value": p} for p in filtered_plants
    ]
    
    # Only reset plant selection if region changed (not date range)
    trigger = ctx.triggered_id
    if trigger is None or trigger == "region-dd":
        # Initial load or region changed - reset selection only if region changed
        if trigger == "region-dd":
            return options, []
        else:
            # Initial load - no selection
            return options, []
    else:
        # Date range or resource type changed - preserve current selection if still valid
        if current_plant_value:
            # Check if current selections are still in the filtered list
            if isinstance(current_plant_value, list):
                valid_selections = [p for p in current_plant_value if p in filtered_plants or p == "SELECT_ALL"]
                if valid_selections:
                    return options, valid_selections
            elif current_plant_value in filtered_plants or current_plant_value == "SELECT_ALL":
                return options, current_plant_value
        # If no valid current selection, return options with current selection preserved
        return options, current_plant_value if current_plant_value else []


@app.callback(
    Output("qca-dd", "options"),
    Output("qca-dd", "value"),
    Input("region-dd", "value"),
    Input("plant-dd", "value"),
    State("qca-dd", "value"),
    prevent_initial_call=False,
)
def update_qcas_from_duckdb(regions, plants, current_qcas):
    """Populate optional QCA filter based on selected regions/plants."""
    if not regions:
        return [], []

    # If SELECT_ALL is chosen (or nothing chosen), don't restrict by plant.
    plant_filter = None
    if plants:
        p_list = plants if isinstance(plants, list) else [plants]
        if "SELECT_ALL" not in p_list:
            plant_filter = p_list

    qcas = get_qcas_from_duckdb(regions, plant_filter)
    options = [{"label": "✓ Select All", "value": "SELECT_ALL"}] + [{"label": q, "value": q} for q in qcas]

    if not current_qcas:
        return options, []
    cur_list = current_qcas if isinstance(current_qcas, list) else [current_qcas]
    valid = [q for q in cur_list if q == "SELECT_ALL" or q in qcas]
    # If Select All is chosen, keep it as the only selection.
    if "SELECT_ALL" in valid:
        return options, ["SELECT_ALL"]
    return options, valid


@app.callback(
    Output("agg-plant-dd", "options"),
    Output("agg-plant-dd", "value"),
    Input("agg-region-dd", "value"),
    Input("agg-resource-type-dd", "value"),
    Input("agg-date-range", "start_date"),
    Input("agg-date-range", "end_date"),
    State("agg-plant-dd", "value"),
    prevent_initial_call=False
)
def update_agg_plants_from_duckdb(regions, resource_type, start_date, end_date, current_plant_value):
    """Update plant options for Aggregation Analysis based on selected regions and resource type.
    Only resets plant selection when region changes, not when date range changes.
    """
    if not regions:
        return [], []

    try:
        sd = str(start_date) if start_date else str(datetime.now().date())
        ed = str(end_date) if end_date else sd
        filtered_plants = get_filtered_plants_by_type(regions, resource_type or "ALL", sd, ed)
    except Exception:
        filtered_plants = get_plants_from_duckdb(regions)

    if not filtered_plants:
        return [], []

    def truncate_label(name, max_length=50):
        return name if len(name) <= max_length else name[:max_length-3] + "..."

    options = [{"label": "✓ Select All", "value": "SELECT_ALL"}] + [
        {"label": truncate_label(p), "value": p} for p in filtered_plants
    ]

    trigger = ctx.triggered_id
    if trigger is None or trigger == "agg-region-dd":
        return options, []
    else:
        if current_plant_value:
            if isinstance(current_plant_value, list):
                valid_selections = [p for p in current_plant_value if p in filtered_plants or p == "SELECT_ALL"]
                if valid_selections:
                    return options, valid_selections
            elif current_plant_value in filtered_plants or current_plant_value == "SELECT_ALL":
                return options, current_plant_value
        return options, current_plant_value if current_plant_value else []


@app.callback(
    Output("agg-qca-dd", "options"),
    Output("agg-qca-dd", "value"),
    Input("agg-region-dd", "value"),
    Input("agg-plant-dd", "value"),
    State("agg-qca-dd", "value"),
    prevent_initial_call=False,
)
def update_agg_qcas_from_duckdb(regions, plants, current_qcas):
    if not regions:
        return [], []

    plant_filter = None
    if plants:
        p_list = plants if isinstance(plants, list) else [plants]
        if "SELECT_ALL" not in p_list:
            plant_filter = p_list

    qcas = get_qcas_from_duckdb(regions, plant_filter)
    options = [{"label": "✓ Select All", "value": "SELECT_ALL"}] + [{"label": q, "value": q} for q in qcas]

    if not current_qcas:
        return options, []
    cur_list = current_qcas if isinstance(current_qcas, list) else [current_qcas]
    valid = [q for q in cur_list if q == "SELECT_ALL" or q in qcas]
    if "SELECT_ALL" in valid:
        return options, ["SELECT_ALL"]
    return options, valid


@app.callback(
    Output("agg-exclude-section", "style"),
    Output("agg-exclude-plant-dd", "options"),
    Output("agg-exclude-plant-dd", "value"),
    Input("agg-plant-dd", "value"),
    Input("agg-region-dd", "value"),
    Input("agg-resource-type-dd", "value"),
    Input("agg-date-range", "start_date"),
    Input("agg-date-range", "end_date"),
    State("agg-exclude-plant-dd", "value"),
    prevent_initial_call=False
)
def toggle_exclude_section(plant_value, regions, resource_type, start_date, end_date, current_exclude_value):
    """Show/hide exclude section and populate exclude dropdown when SELECT_ALL is active."""
    # Check if SELECT_ALL is selected
    is_select_all = False
    if isinstance(plant_value, list):
        is_select_all = "SELECT_ALL" in plant_value
    elif plant_value == "SELECT_ALL":
        is_select_all = True
    
    # Hide section if SELECT_ALL is not active
    if not is_select_all:
        return {"display": "none"}, [], []
    
    # Show section and populate options
    if not regions:
        return {"display": "block"}, [], current_exclude_value if current_exclude_value else []
    
    try:
        sd = str(start_date) if start_date else str(datetime.now().date())
        ed = str(end_date) if end_date else sd
        filtered_plants = get_filtered_plants_by_type(regions, resource_type or "ALL", sd, ed)
    except Exception:
        filtered_plants = get_plants_from_duckdb(regions)
    
    if not filtered_plants:
        return {"display": "block"}, [], current_exclude_value if current_exclude_value else []
    
    def truncate_label(name, max_length=50):
        return name if len(name) <= max_length else name[:max_length-3] + "..."
    
    exclude_options = [
        {"label": truncate_label(p), "value": p} for p in filtered_plants
    ]
    
    # Filter out invalid excluded plants if filter changed
    if current_exclude_value:
        if isinstance(current_exclude_value, list):
            valid_excluded = [p for p in current_exclude_value if p in filtered_plants]
            return {"display": "block"}, exclude_options, valid_excluded
        elif current_exclude_value in filtered_plants:
            return {"display": "block"}, exclude_options, current_exclude_value
    
    return {"display": "block"}, exclude_options, current_exclude_value if current_exclude_value else []

# Stats plants based on filters
@app.callback(
    Output("stats-plant-dd", "options"),
    Output("stats-plant-dd", "value"),
    Input("stats-region-dd", "value"),
    Input("stats-resource-dd", "value"),
    Input("stats-date-range", "start_date"),
    Input("stats-date-range", "end_date"),
    prevent_initial_call=False
)
def update_stats_plants(regions, resource_type, start_date, end_date):
    if not regions:
        return [], []
    sd = str(start_date) if start_date else str(datetime.now().date())
    ed = str(end_date) if end_date else sd
    try:
        filtered_plants = get_filtered_plants_by_type(regions, resource_type or "ALL", sd, ed)
    except Exception:
        filtered_plants = get_plants_from_duckdb(regions)
    if not filtered_plants:
        return [], []
    def trunc(n, L=50):
        return n if len(n) <= L else n[:L-3] + "..."
    opts = [{"label": "✓ Select All", "value": "SELECT_ALL"}] + [{"label": trunc(p), "value": p} for p in filtered_plants]
    return opts, []


@app.callback(
    Output("stats-qca-dd", "options"),
    Output("stats-qca-dd", "value"),
    Input("stats-region-dd", "value"),
    Input("stats-plant-dd", "value"),
    State("stats-qca-dd", "value"),
    prevent_initial_call=False,
)
def update_stats_qcas(regions, plants, current_qcas):
    if not regions:
        return [], []

    plant_filter = None
    if plants:
        p_list = plants if isinstance(plants, list) else [plants]
        if "SELECT_ALL" not in p_list:
            plant_filter = p_list

    qcas = get_qcas_from_duckdb(regions, plant_filter)
    options = [{"label": "✓ Select All", "value": "SELECT_ALL"}] + [{"label": q, "value": q} for q in qcas]

    if not current_qcas:
        return options, []
    cur_list = current_qcas if isinstance(current_qcas, list) else [current_qcas]
    valid = [q for q in cur_list if q == "SELECT_ALL" or q in qcas]
    if "SELECT_ALL" in valid:
        return options, ["SELECT_ALL"]
    return options, valid

# Disable Run Analysis until all inputs present
@app.callback(
    Output("plot-now", "disabled"),
    Input("region-dd", "value"),
    Input("resource-type-dd", "value"),
    Input("plant-dd", "value"),
    Input("date-range", "start_date"),
    Input("date-range", "end_date"),
    Input("analysis-config-store", "data"),
    prevent_initial_call=False
)
def toggle_plot_now(regions, resource_type, plants, start_date, end_date, config_rows):
    has_dates = bool(start_date) and bool(end_date)
    valid_configs = [r for r in (config_rows or []) if r.get("region") and r.get("plant")]
    use_config_rows = bool(valid_configs)
    use_main_filters = bool(regions) and bool(plants)
    enabled = has_dates and (use_config_rows or use_main_filters)
    return not enabled


@app.callback(
    Output("agg-plot-now", "disabled"),
    Input("agg-region-dd", "value"),
    Input("agg-resource-type-dd", "value"),
    Input("agg-plant-dd", "value"),
    Input("agg-exclude-plant-dd", "value"),
    Input("agg-date-range", "start_date"),
    Input("agg-date-range", "end_date"),
    prevent_initial_call=False
)
def toggle_agg_buttons(regions, resource_type, plants, excluded_plants, start_date, end_date):
    has_regions = bool(regions)
    has_plants = bool(plants)
    has_dates = bool(start_date) and bool(end_date)
    
    # Validate that not all plants are excluded when SELECT_ALL is active
    if has_plants:
        is_select_all = False
        if isinstance(plants, list):
            is_select_all = "SELECT_ALL" in plants
        elif plants == "SELECT_ALL":
            is_select_all = True
        
        if is_select_all and excluded_plants:
            # Check if all plants would be excluded
            try:
                filtered_plants = get_filtered_plants_by_type(regions, resource_type or "ALL", 
                                                             str(start_date) if start_date else str(datetime.now().date()),
                                                             str(end_date) if end_date else str(datetime.now().date()))
            except Exception:
                filtered_plants = get_plants_from_duckdb(regions) if regions else []
            
            if filtered_plants:
                excluded_list = excluded_plants if isinstance(excluded_plants, list) else [excluded_plants]
                remaining = [p for p in filtered_plants if p not in excluded_list]
                if not remaining:
                    # All plants excluded - disable button
                    return True
    
    disabled = not (has_regions and has_plants and has_dates)
    return disabled

# Enable stats Run button
@app.callback(
    Output("stats-run", "disabled"),
    Input("stats-region-dd", "value"),
    Input("stats-plant-dd", "value"),
    Input("stats-date-range", "start_date"),
    Input("stats-date-range", "end_date"),
    prevent_initial_call=False
)
def toggle_stats_run(regions, plants, start_date, end_date):
    has_regions = bool(regions)
    has_plants = bool(plants)
    has_dates = bool(start_date) and bool(end_date)
    return not (has_regions and has_plants and has_dates)

# Callback to load saved settings into form (without bands-table to avoid duplicate)
@app.callback(
    Output("err-mode", "value"),
    Output("x-pct", "value"),
    Input("nav-store", "data"),
    State("saved-settings-store", "data"),
    prevent_initial_call=False
)
def load_saved_settings(nav_data, saved_settings):
    """Load saved settings into form when navigating to settings page"""
    if nav_data == "settings" and saved_settings and isinstance(saved_settings, dict):
        # Load saved settings into form
        return (
            saved_settings.get("err_mode", "default"),
            saved_settings.get("x_pct", 50)
        )
    else:
        # Return defaults
        return "default", 50

# Callback to save settings
@app.callback(
    Output("settings-save-message", "children"),
    Output("saved-settings-store", "data"),
    Input("btn-save-settings", "n_clicks"),
    State("err-mode", "value"),
    State("x-pct", "value"),
    State("bands-table", "data"),
    State("zero-basis-guard", "value"),
    prevent_initial_call=True
)
def save_settings(n_clicks, err_mode, x_pct, bands_data, zero_basis_guard):
    """Save current settings to local storage"""
    if not n_clicks:
        raise PreventUpdate
    
    # Package settings
    settings = {
        "err_mode": err_mode,
        "x_pct": x_pct,
        "bands": bands_data,
        "zero_basis_guard": "on" in (zero_basis_guard or [])
    }
    
    # Success message
    message = dbc.Alert([
        html.Span("✓ ", style={"fontSize": "1.2rem", "marginRight": "8px"}),
        html.Strong("Settings saved successfully!"),
        html.Br(),
        html.Small("Your configuration will be used when running analysis.")
    ], color="success", dismissable=True, duration=4000)
    
    return message, settings

# Single callback to handle bands table data from saved settings
@app.callback(
    Output("bands-table", "data", allow_duplicate=True),
    Input("saved-settings-store", "data"),
    Input("nav-store", "data"),
    prevent_initial_call=True
)
def load_saved_bands(saved_settings, nav_data):
    """Load saved bands data when settings change or navigating to settings page"""
    if saved_settings and isinstance(saved_settings, dict):
        bands_data = saved_settings.get("bands", DEFAULT_BANDS.copy())
        print(f"DEBUG - Loading saved bands: {len(bands_data)} bands")
        return bands_data
    else:
        print(f"DEBUG - Using default bands: {len(DEFAULT_BANDS)} bands")
        return DEFAULT_BANDS.copy()

@app.callback(
    Output("x-pct-container", "style"),
    Output("x-pct-readout", "children"),
    Output("x-pct-label", "children"),
    Input("err-mode", "value"),
    Input("x-pct", "value"),
)
def toggle_xpct(err_mode, x):
    show = {"display": "block"} if err_mode == "dynamic" else {"display": "none"}
    readout = f"X = {x:.0f}%" if err_mode == "dynamic" else ""
    label = f"X% = {x:.0f}% (for Dynamic Error%)" if err_mode == "dynamic" else "X% (for Dynamic Error%)"
    return show, readout, label


@app.callback(
    Output("agg-ppa-numeric-container", "style"),
    Input("agg-ppa-mode", "value"),
)
def toggle_agg_numeric_ppa(mode):
    # agg-ppa-mode is multi-select; show numeric input if "numeric" is among selected modes
    if isinstance(mode, list):
        show_numeric = "numeric" in mode
    else:
        show_numeric = mode == "numeric"
    show = {"display": "block"} if show_numeric else {"display": "none"}
    return show

@app.callback(
    Output("bands-table", "data", allow_duplicate=True),
    Output("download-bands-json", "data"),
    Output("form-direction", "value"),
    Output("form-lower-pct", "value"),
    Output("form-upper-pct", "value"),
    Output("form-tolerance-cut-pct", "value"),
    Output("form-loss-zone", "value"),
    Output("form-rate-type", "value"),
    Output("form-rate-value", "value"),
    Output("form-excess-slope-per-pct", "value"),
    Input("reset-bands", "n_clicks"),
    Input("bands-table", "data_timestamp"),
    Input("upload-bands", "contents"),
    Input("save-bands", "n_clicks"),
    Input("add-from-form", "n_clicks"),
    State("bands-table", "data"),
    State("form-direction", "value"),
    State("form-lower-pct", "value"),
    State("form-upper-pct", "value"),
    State("form-tolerance-cut-pct", "value"),
    State("form-loss-zone", "value"),
    State("form-rate-type", "value"),
    State("form-rate-value", "value"),
    State("form-excess-slope-per-pct", "value"),
    prevent_initial_call=True
)
def manage_bands(reset_clicks, timestamp, uploaded, save_clicks, add_clicks, rows, 
                direction, lower_pct, upper_pct, tolerance_cut_pct, loss_zone, 
                rate_type, rate_value, excess_slope_per_pct):
    trig = ctx.triggered_id
    rows = rows or []
    
    # Default form reset values
    form_reset = ("UI", 0.0, 10.0, 10.0, False, "flat_per_kwh", 0.0, 0.0)
    
    # Handle reset
    if trig == "reset-bands":
        return DEFAULT_BANDS.copy(), dash.no_update, *form_reset
    
    # Handle upload
    if trig == "upload-bands" and uploaded:
        header, b64 = uploaded.split(",", 1)
        import base64, io
        payload = base64.b64decode(b64)
        try:
            js = json.loads(payload.decode("utf-8"))
            assert isinstance(js, list)
            return js, dash.no_update, *form_reset
        except Exception:
            return dash.no_update, dash.no_update, *form_reset
    
    # Handle save
    if trig == "save-bands" and save_clicks:
        content = json.dumps(rows or [], indent=2)
        return dash.no_update, dict(content=content, filename="bands_config_preset.json"), *form_reset
    
    # Handle add from form
    if trig == "add-from-form" and add_clicks:
        # Validate required fields
        if all([direction, lower_pct is not None, upper_pct is not None, tolerance_cut_pct is not None, 
                rate_type, rate_value is not None, excess_slope_per_pct is not None]):
            
            # Create new row
            new_row = {
                "direction": direction,
                "lower_pct": float(lower_pct),
                "upper_pct": float(upper_pct),
                "tolerance_cut_pct": float(tolerance_cut_pct),
                "rate_type": rate_type,
                "rate_value": float(rate_value),
                "excess_slope_per_pct": float(excess_slope_per_pct),
                "loss_zone": bool(loss_zone),
                "label": ""
            }
            
            # Generate label for new row
            new_row["label"] = generate_label(new_row)
            
            # Add to existing rows
            updated_rows = (rows or []) + [new_row]
            
            return updated_rows, dash.no_update, *form_reset
    
    # Auto-update labels when data changes
    if trig == "bands-table" and rows:
        for row in rows:
            if row.get("label") == "" or not row.get("label"):
                row["label"] = generate_label(row)
        return rows, dash.no_update, *form_reset
    
    return dash.no_update, dash.no_update, *form_reset


# Form label preview callback
@app.callback(
    Output("form-label-preview", "children"),
    Input("form-direction", "value"),
    Input("form-lower-pct", "value"),
    Input("form-upper-pct", "value"),
    Input("form-rate-type", "value"),
    Input("form-rate-value", "value"),
    prevent_initial_call=True
)
def preview_form_label(direction, lower_pct, upper_pct, rate_type, rate_value):
    if not all([direction, lower_pct is not None, upper_pct is not None, rate_type, rate_value is not None]):
        return "Fill in all fields to see preview..."
    
    # Generate label using the same logic as the table
    row = {
        "direction": direction,
        "lower_pct": float(lower_pct) if lower_pct else 0,
        "upper_pct": float(upper_pct) if upper_pct else 0,
        "rate_type": rate_type,
        "rate_value": float(rate_value) if rate_value else 0
    }
    return generate_label(row)


def _compute_pipeline(regions, plants, start_date, end_date, err_mode, x_pct, bands_rows, unpaid_oi_threshold=15.0, qcas=None):
    """Run analysis pipeline: load → error% (default/dynamic) → custom bands → summarize.
    Region-agnostic: STU behaves exactly like RPC (custom bands, dynamic error%, same exports)."""
    if not regions or not plants:
        return {"error": "Please select at least one region and plant"}
    
    # Handle "Select All" case - resolve from DuckDB only for consistency
    if "SELECT_ALL" in plants:
        plants = get_plants_from_duckdb(regions)
    
    # Load data for all selected regions (multi-region support)
    qca_filter = [q for q in (qcas or []) if q and q != "SELECT_ALL"]
    df = load_region_data(regions, str(start_date), str(end_date), plants, qcas=qca_filter or None)
    if df is None or df.empty:
        return {"error": "No data found for the selected filters (or database is busy). Try a different date range—data availability varies by plant."}
    # Ensure required columns exist before continuing
    required_cols = {"Actual_MW", "Scheduled_MW", "AvC_MW", "PPA", "date", "time_block"}
    if not required_cols.issubset(set(df.columns)):
        return {"error": f"Data is missing required columns: {', '.join(sorted(required_cols - set(df.columns)))}"}

    df["error_pct"] = compute_error_pct(df, err_mode, float(x_pct))
    df["basis_MW"] = compute_basis_mw(df, err_mode, float(x_pct))

    bands_df = _normalize_bands_df(pd.DataFrame(bands_rows or []))
    # Only core fields are truly required after normalization
    core_required = {"direction","lower_pct","upper_pct","rate_type","rate_value","excess_slope_per_pct"}
    if not core_required.issubset(set(bands_df.columns)):
        missing = core_required - set(bands_df.columns)
        raise ValueError(f"Bands config missing required columns: {', '.join(sorted(missing))}")
    bands_rows = bands_df.to_dict("records")

    df = apply_bands(df, bands_rows, unpaid_oi_threshold)
    summary = summarize(df, selected_plants=plants, bands_rows=bands_rows, err_mode=err_mode, x_pct=float(x_pct))
    return summary


def _compute_pipeline_per_plant(
    regions,
    plant_preset_map,
    start_date,
    end_date,
    bands_rows_form,
    saved_settings,
    presets_store,
    unpaid_oi_threshold=15.0,
    qcas=None,
):
    """Run analysis per plant with its assigned preset. plant_preset_map: list of {Plant, Preset}."""
    if not plant_preset_map:
        return {"error": "No plant-preset assignments. Add rows in the per-plant table."}

    name_to_settings = {}
    for p in (presets_store or []):
        if isinstance(p, dict) and p.get("name"):
            name_to_settings[p["name"]] = p.get("settings", {})

    def get_settings(preset_name):
        if preset_name == "__CURRENT__" or not preset_name:
            base = saved_settings if (isinstance(saved_settings, dict) and saved_settings) else {}
            return {
                "err_mode": base.get("err_mode", "default"),
                "x_pct": base.get("x_pct", 50),
                "bands": base.get("bands") or bands_rows_form or DEFAULT_BANDS.copy(),
                "unpaid_oi_threshold": base.get("unpaid_oi_threshold", 15.0),
            }
        st = name_to_settings.get(preset_name)
        if not st:
            return {"err_mode": "default", "x_pct": 50, "bands": bands_rows_form or DEFAULT_BANDS.copy(), "unpaid_oi_threshold": 15.0}
        return st

    results = []
    for row in plant_preset_map:
        plant = row.get("Plant") or row.get("plant")
        preset_name = row.get("Preset") or row.get("preset") or "__CURRENT__"
        if not plant:
            continue
        settings = get_settings(preset_name)
        err_mode = settings.get("err_mode", "default")
        x_pct = settings.get("x_pct", 50)
        bands = settings.get("bands") or bands_rows_form or DEFAULT_BANDS.copy()
        unpaid = float(settings.get("unpaid_oi_threshold", 15.0))
        res = _compute_pipeline(regions, [plant], start_date, end_date, err_mode, x_pct, bands, unpaid, qcas=qcas)
        if "error" in res:
            return res
        ps = res.get("plant_summary", pd.DataFrame())
        if isinstance(ps, pd.DataFrame) and not ps.empty:
            ps = ps.copy()
            ps["Custom Setting"] = preset_name if preset_name != "__CURRENT__" else "Current"
        res["plant_summary"] = ps
        res["_preset_name"] = preset_name if preset_name != "__CURRENT__" else "Current"
        res["_final_err_mode"] = err_mode
        res["_final_x_pct"] = x_pct
        res["_final_bands"] = bands
        results.append(res)

    if not results:
        return {"error": "No valid plant-preset rows to compute."}

    combined_ps = pd.concat([r["plant_summary"] for r in results if isinstance(r.get("plant_summary"), pd.DataFrame)], ignore_index=True)
    blockwise_parts = []
    for r in results:
        bw = r.get("blockwise")
        ps = r.get("plant_summary")
        if isinstance(bw, pd.DataFrame) and not bw.empty and isinstance(ps, pd.DataFrame) and not ps.empty:
            plant_name = ps["Plant name"].iloc[0] if "Plant name" in ps.columns else ""
            blockwise_parts.append(bw.assign(Plant=plant_name))
    combined_blockwise = pd.concat(blockwise_parts, ignore_index=True) if blockwise_parts else pd.DataFrame()

    detail_dfs = [r["df"] for r in results if "df" in r and not r["df"].empty]
    combined_df = pd.concat(detail_dfs, ignore_index=True) if detail_dfs else results[0]["df"]

    first = results[0]
    return {
        "df": combined_df,
        "kpis": first.get("kpis", {}),
        "blockwise": combined_blockwise if not combined_blockwise.empty else first.get("blockwise", pd.DataFrame()),
        "plant_summary": combined_ps,
        "_all_runs": [{"name": r["_preset_name"], "df": r["df"], "final_bands": r.get("_final_bands", []), "err_mode": r["_final_err_mode"], "x_pct": r["_final_x_pct"]} for r in results],
        "_final_err_mode": first["_final_err_mode"],
        "_final_x_pct": first["_final_x_pct"],
        "final_bands": first.get("_final_bands", []),
    }


def _compute_pipeline_aggregated(
    regions,
    plants,
    start_date,
    end_date,
    err_mode,
    x_pct,
    bands_rows,
    ppa_mode: str,
    ppa_value: float | None,
    unpaid_oi_threshold: float = 15.0,
    qcas=None,
):
    """Run the same analysis engine on an aggregated multi-plant time series.
    STU and RPC are treated identically (custom bands, dynamic error%, same summarization and exports).

    This function only changes the input dataset by first aggregating AvC/Scheduled/Actual
    across the selected plants per time block, and then computing PPA as per the
    requested aggregation mode. All downstream calculations reuse the existing
    analysis logic unchanged.
    """
    if not regions or not plants:
        return {"error": "Please select at least one region and plant"}

    # Resolve "Select All" - but plants should already be resolved by caller (with exclusions applied)
    # Handle both cases: list of plants or SELECT_ALL string (for backward compatibility)
    if isinstance(plants, list):
        if "SELECT_ALL" in plants:
            plants = get_plants_from_duckdb(regions)
    elif plants == "SELECT_ALL":
        plants = get_plants_from_duckdb(regions)
    elif not isinstance(plants, list):
        plants = [plants] if plants else []

    if not plants:
        return {"error": "No plants resolved for aggregation"}

    # Load data from ALL selected regions (single query for multi-region)
    qca_filter = [q for q in (qcas or []) if q and q != "SELECT_ALL"]
    df_raw = load_region_data(regions, str(start_date), str(end_date), plants, qcas=qca_filter or None)
    
    if df_raw is None or df_raw.empty:
        return {"error": "No data found for the selected filters"}
    
    # Create comma-separated region string for display
    regions_str = ",".join(sorted(regions))
    
    # Ensure core numeric fields are usable before aggregation
    df_raw["AvC_MW"] = pd.to_numeric(df_raw.get("AvC_MW"), errors="coerce").fillna(0.0)
    df_raw["Scheduled_MW"] = pd.to_numeric(df_raw.get("Scheduled_MW"), errors="coerce").fillna(0.0)
    df_raw["Actual_MW"] = pd.to_numeric(df_raw.get("Actual_MW"), errors="coerce").fillna(0.0)
    df_raw["PPA"] = pd.to_numeric(df_raw.get("PPA"), errors="coerce").fillna(0.0)

    # Group per time block/slot (aggregate across all regions)
    # Note: region column will be updated to show comma-separated regions after grouping
    group_cols = ["date", "time_block", "from_time", "to_time"]
    for col in group_cols:
        if col not in df_raw.columns:
            df_raw[col] = None

    agg_df = (
        df_raw.groupby(group_cols, as_index=False)
        .agg({
            "AvC_MW": "sum",
            "Scheduled_MW": "sum",
            "Actual_MW": "sum",
        })
    )

    # PPA aggregation per slot (across all regions)
    mode_norm = (ppa_mode or "mean").strip().lower()
    if mode_norm == "numeric" and ppa_value is not None:
        agg_df["PPA"] = float(ppa_value)
    elif mode_norm == "weighted":
        # Weighted PPA per time-block:
        #   Weighted_PPA[t] = sum(PPA * AvC_MW) / sum(AvC_MW)
        # Uses the exact tags and does not change any downstream formulas.
        df_raw["_ppa_x_avc"] = df_raw["PPA"] * df_raw["AvC_MW"]
        num = df_raw.groupby(group_cols, as_index=False)["_ppa_x_avc"].sum().rename(columns={"_ppa_x_avc": "_ppa_num"})
        agg_df = agg_df.merge(num, on=group_cols, how="left")
        agg_df["_ppa_num"] = pd.to_numeric(agg_df.get("_ppa_num"), errors="coerce").fillna(0.0)
        denom = pd.to_numeric(agg_df.get("AvC_MW"), errors="coerce").fillna(0.0)
        agg_df["PPA"] = np.where(denom > 0, agg_df["_ppa_num"] / denom, 0.0)
        agg_df = agg_df.drop(columns=["_ppa_num"], errors="ignore")
        df_raw = df_raw.drop(columns=["_ppa_x_avc"], errors="ignore")
    else:
        def _agg_ppa(series: pd.Series) -> float:
            vals = pd.to_numeric(series, errors="coerce").dropna()
            if vals.empty:
                return 0.0
            if mode_norm == "median":
                return float(vals.median())
            if mode_norm == "mode":
                # Use existing safe_mode helper to respect mode semantics
                return float(safe_mode(vals.tolist()))
            # Default to mean
            return float(vals.mean())

        ppa_by_slot = (
            df_raw.groupby(group_cols)["PPA"]
            .apply(_agg_ppa)
            .reset_index(name="PPA")
        )
        agg_df = agg_df.merge(ppa_by_slot, on=group_cols, how="left")
        agg_df["PPA"] = pd.to_numeric(agg_df["PPA"], errors="coerce").fillna(0.0)

    # Add region column with comma-separated region names
    agg_df["region"] = regions_str

    # Synthesize a single aggregated plant identity for downstream logic
    agg_df["Plant"] = "AGGREGATED"
    agg_df["plant_name"] = "AGGREGATED"

    # Recreate date_time for completeness (matches load_region_data helper)
    try:
        agg_df["date_time"] = pd.to_datetime(agg_df["date"]) + pd.to_timedelta(
            (agg_df["time_block"] - 1) * 15, unit="m"
        )
    except Exception:
        pass

    # Error% and basis computed on aggregated MW values using existing helpers
    agg_df["error_pct"] = compute_error_pct(agg_df, err_mode, float(x_pct))
    agg_df["basis_MW"] = compute_basis_mw(agg_df, err_mode, float(x_pct))

    # Bands normalization exactly as in the main pipeline
    bands_df = _normalize_bands_df(pd.DataFrame(bands_rows or []))
    core_required = {"direction", "lower_pct", "upper_pct", "rate_type", "rate_value", "excess_slope_per_pct"}
    if not core_required.issubset(set(bands_df.columns)):
        missing = core_required - set(bands_df.columns)
        raise ValueError(f"Bands config missing required columns: {', '.join(sorted(missing))}")
    bands_rows_norm = bands_df.to_dict("records")

    # Apply bands and summarize using the existing engine; do NOT pass per-plant selection
    agg_df = apply_bands(agg_df, bands_rows_norm, unpaid_oi_threshold)
    summary = summarize(agg_df, selected_plants=None, bands_rows=bands_rows_norm, err_mode=err_mode, x_pct=float(x_pct))
    summary["aggregated_plants"] = list(plants)
    summary["aggregated_region"] = regions_str  # Use comma-separated regions string
    return summary


# --- Preset management callbacks ---
@app.callback(
    Output("preset-save-message", "children"),
    Output("presets-store", "data"),
    Input("btn-save-preset", "n_clicks"),
    State("preset-name", "value"),
    State("err-mode", "value"),
    State("x-pct", "value"),
    State("bands-table", "data"),
    State("zero-basis-guard", "value"),
    State("presets-store", "data"),
    prevent_initial_call=True
)
def save_preset(n_clicks, name, err_mode, x_pct, bands_data, zero_basis_guard, presets):
    if not n_clicks:
        raise PreventUpdate
    name = (name or "").strip()
    if not name:
        return dbc.Alert("Please enter a preset name.", color="warning"), dash.no_update
    preset = {
        "name": name,
        "settings": {
            "err_mode": err_mode,
            "x_pct": x_pct,
            "bands": bands_data,
            "zero_basis_guard": "on" in (zero_basis_guard or [])
        }
    }
    presets = list(presets or [])
    names = [p.get("name") for p in presets]
    if name in names:
        presets[names.index(name)] = preset
        msg = f"Preset '{name}' updated."
    else:
        presets.append(preset)
        msg = f"Preset '{name}' saved."
    print(f"DEBUG - {msg} (total presets: {len(presets)})")
    return dbc.Alert(msg, color="success"), presets


@app.callback(
    Output("analysis-preset-select", "options"),
    Input("presets-store", "data"),
)
def load_preset_options(presets):
    presets = presets or []
    return [{"label": p.get("name", "Unnamed"), "value": p.get("name", "Unnamed")} for p in presets]


# ---- Add-more config rows ----
def _config_row(i: int, cfg: dict, region_opts: list, resource_opts: list, plant_opts: list, qca_opts: list, preset_opts: list) -> html.Div:
    """Build a single config row with Region, Resource, Plant, QCA, Custom Setting dropdowns."""
    return html.Div([
        dbc.Row([
            dbc.Col([
                dcc.Dropdown(
                    id={"type": "cfg-region", "index": i},
                    options=region_opts,
                    value=cfg.get("region"),
                    placeholder="Region",
                    clearable=False,
                    style={"fontSize": "0.9rem"}
                ),
            ], md=2),
            dbc.Col([
                dcc.Dropdown(
                    id={"type": "cfg-resource", "index": i},
                    options=resource_opts,
                    value=cfg.get("resource") or "ALL",
                    clearable=False,
                    style={"fontSize": "0.9rem"}
                ),
            ], md=2),
            dbc.Col([
                dcc.Dropdown(
                    id={"type": "cfg-plant", "index": i},
                    options=plant_opts,
                    value=cfg.get("plant"),
                    placeholder="Plant",
                    searchable=True,
                    style={"fontSize": "0.9rem"}
                ),
            ], md=2),
            dbc.Col([
                dcc.Dropdown(
                    id={"type": "cfg-qca", "index": i},
                    options=qca_opts,
                    value=cfg.get("qca"),
                    placeholder="QCA (optional)",
                    clearable=True,
                    style={"fontSize": "0.9rem"}
                ),
            ], md=2),
            dbc.Col([
                dcc.Dropdown(
                    id={"type": "cfg-preset", "index": i},
                    options=preset_opts,
                    value=cfg.get("preset") or "__CURRENT__",
                    clearable=False,
                    style={"fontSize": "0.9rem"}
                ),
            ], md=2),
            dbc.Col([
                dbc.Button("✕", id={"type": "cfg-remove", "index": i}, color="danger", size="sm", outline=True, n_clicks=0),
            ], md=1, className="d-flex align-items-center"),
        ], className="mb-2 align-items-center"),
    ], id={"type": "cfg-row-wrapper", "index": i})


@app.callback(
    Output("analysis-config-store", "data", allow_duplicate=True),
    Input("analysis-add-config", "n_clicks"),
    State("analysis-config-store", "data"),
    prevent_initial_call=True,
)
def add_config_row(n_clicks, store):
    if not n_clicks:
        raise PreventUpdate
    store = store or []
    new_row = {"region": None, "resource": "ALL", "plant": None, "qca": None, "preset": "__CURRENT__"}
    return store + [new_row]


@app.callback(
    Output("analysis-config-store", "data", allow_duplicate=True),
    Input({"type": "cfg-remove", "index": ALL}, "n_clicks"),
    State("analysis-config-store", "data"),
    prevent_initial_call=True,
)
def remove_config_row(remove_clicks, store):
    if not remove_clicks or not any(remove_clicks):
        raise PreventUpdate
    trigger = ctx.triggered_id
    if not trigger or trigger.get("type") != "cfg-remove":
        raise PreventUpdate
    idx = trigger.get("index")
    if idx is None:
        raise PreventUpdate
    store = list(store or [])
    if len(store) <= 1:
        return store
    store.pop(idx)
    return store


@app.callback(
    Output("analysis-config-store", "data", allow_duplicate=True),
    Input({"type": "cfg-region", "index": ALL}, "value"),
    Input({"type": "cfg-resource", "index": ALL}, "value"),
    Input({"type": "cfg-plant", "index": ALL}, "value"),
    Input({"type": "cfg-qca", "index": ALL}, "value"),
    Input({"type": "cfg-preset", "index": ALL}, "value"),
    State("analysis-config-store", "data"),
    prevent_initial_call=True,
)
def sync_config_from_dropdowns(regions, resources, plants, qcas, presets, store):
    store = store or []
    n = len(store)
    if not n or (len(regions or []) != n):
        return no_update
    updated = []
    for i in range(n):
        updated.append({
            "region": (regions or [None] * n)[i] if i < len(regions or []) else store[i].get("region"),
            "resource": (resources or ["ALL"] * n)[i] if i < len(resources or []) else store[i].get("resource", "ALL"),
            "plant": (plants or [None] * n)[i] if i < len(plants or []) else store[i].get("plant"),
            "qca": (qcas or [None] * n)[i] if i < len(qcas or []) else store[i].get("qca"),
            "preset": (presets or ["__CURRENT__"] * n)[i] if i < len(presets or []) else store[i].get("preset", "__CURRENT__"),
        })
    return updated


@app.callback(
    Output("analysis-config-container", "children"),
    Input("analysis-config-store", "data"),
    Input("presets-store", "data"),
    Input("date-range", "start_date"),
    Input("date-range", "end_date"),
)
def build_config_container(store, presets, start_date, end_date):
    store = store or [{"region": None, "resource": "ALL", "plant": None, "qca": None, "preset": "__CURRENT__"}]
    region_opts = [{"label": r, "value": r} for r in get_regions_from_duckdb()]
    resource_opts = [{"label": "All", "value": "ALL"}, {"label": "Solar", "value": "SOLAR"}, {"label": "Wind", "value": "WIND"}, {"label": "Thermal", "value": "THERMAL"}]
    preset_opts = [{"label": "Current (form bands)", "value": "__CURRENT__"}]
    for p in (presets or []):
        if isinstance(p, dict) and p.get("name"):
            preset_opts.append({"label": p["name"], "value": p["name"]})

    children = []
    sd = str(start_date) if start_date else str(datetime.now().date())
    ed = str(end_date) if end_date else sd
    for i, cfg in enumerate(store):
        regions_for_row = [cfg.get("region")] if cfg.get("region") else []
        plant_opts = []
        if regions_for_row:
            try:
                plants = get_filtered_plants_by_type(regions_for_row, cfg.get("resource") or "ALL", sd, ed)
                plant_opts = [{"label": p, "value": p} for p in plants]
            except Exception:
                plant_opts = [{"label": p, "value": p} for p in get_plants_from_duckdb(regions_for_row)]
        qca_opts = []
        if regions_for_row:
            plant_for_qca = [cfg.get("plant")] if cfg.get("plant") else None
            qca_list = get_qcas_from_duckdb(regions_for_row, plant_for_qca)
            qca_opts = [{"label": q, "value": q} for q in qca_list]
        children.append(_config_row(i, cfg, region_opts, resource_opts, plant_opts, qca_opts, preset_opts))
    return children


@app.callback(
    Output("agg-analysis-preset-select", "options"),
    Input("presets-store", "data"),
)
def load_agg_preset_options(presets):
    presets = presets or []
    return [{"label": p.get("name", "Unnamed"), "value": p.get("name", "Unnamed")} for p in presets]


@app.callback(
    Output("agg-per-plant-section", "style"),
    Input("agg-per-plant-toggle", "value"),
)
def toggle_agg_per_plant_section(toggle_val):
    return {"display": "block"} if toggle_val and "on" in toggle_val else {"display": "none"}


@app.callback(
    Output("agg-plant-preset-table", "data"),
    Output("agg-plant-preset-table", "dropdown"),
    Input("agg-plant-dd", "value"),
    Input("agg-per-plant-toggle", "value"),
    Input("presets-store", "data"),
    State("agg-region-dd", "value"),
    State("agg-resource-type-dd", "value"),
    State("agg-date-range", "start_date"),
    State("agg-date-range", "end_date"),
    State("agg-exclude-plant-dd", "value"),
    prevent_initial_call=False,
)
def update_agg_plant_preset_table(plants, toggle_val, presets, regions, resource_type, start_date, end_date, excluded_plants):
    show = toggle_val and "on" in toggle_val
    if not show or not plants or not regions:
        return [], {"Preset": {"options": [{"label": "Current (form bands)", "value": "__CURRENT__"}]}}

    plant_list = plants if isinstance(plants, list) else [plants]
    if "SELECT_ALL" in plant_list:
        try:
            sd = str(start_date) if start_date else str(datetime.now().date())
            ed = str(end_date) if end_date else sd
            all_p = get_filtered_plants_by_type(regions, resource_type or "ALL", sd, ed)
            excluded_list = excluded_plants if isinstance(excluded_plants, list) else (excluded_plants or [])
            plant_list = [p for p in all_p if p not in excluded_list]
        except Exception:
            plant_list = get_plants_from_duckdb(regions)
    if not plant_list:
        return [], {"Preset": {"options": [{"label": "Current (form bands)", "value": "__CURRENT__"}]}}

    preset_opts = [{"label": "Current (form bands)", "value": "__CURRENT__"}]
    for p in (presets or []):
        if isinstance(p, dict) and p.get("name"):
            preset_opts.append({"label": p["name"], "value": p["name"]})
    return [{"Plant": p, "Preset": "__CURRENT__"} for p in plant_list], {"Preset": {"options": preset_opts}}


@app.callback(
    Output("presets-store", "data", allow_duplicate=True),
    Input("btn-delete-preset", "n_clicks"),
    State("analysis-preset-select", "value"),
    State("presets-store", "data"),
    prevent_initial_call=True
)
def delete_presets(n_clicks, selected_names, presets):
    if not n_clicks or not selected_names:
        raise PreventUpdate
    presets = list(presets or [])
    keep = [p for p in presets if p.get("name") not in (selected_names or [])]
    print(f"DEBUG - Deleted presets: {set(selected_names) - {p.get('name') for p in keep}}; remaining: {len(keep)}")
    return keep

@app.callback(
    Output("results-store", "data"),
    Output("results-section", "style"),
    Output("progress-container", "style"),
    Input("plot-now", "n_clicks"),
    State("region-dd", "value"),
    State("plant-dd", "value"),
    State("qca-dd", "value"),
    State("date-range", "start_date"),
    State("date-range", "end_date"),
    State("err-mode", "value"),
    State("x-pct", "value"),
    State("bands-table", "data"),
    State("saved-settings-store", "data"),
    State("analysis-preset-select", "value"),
    State("presets-store", "data"),
    State("analysis-config-store", "data"),
    State({"type": "cfg-region", "index": ALL}, "value"),
    State({"type": "cfg-resource", "index": ALL}, "value"),
    State({"type": "cfg-plant", "index": ALL}, "value"),
    State({"type": "cfg-qca", "index": ALL}, "value"),
    State({"type": "cfg-preset", "index": ALL}, "value"),
    prevent_initial_call=True
)
def compute_on_click(n, regions, plants, qcas, start_date, end_date, err_mode, x_pct, bands_rows, saved_settings, selected_preset_names, presets_store, config_rows, cfg_regions, cfg_resources, cfg_plants, cfg_qcas, cfg_presets):
    if not n:
        raise PreventUpdate

    # Build valid_configs from live dropdown values (preferred) or store
    config_rows = config_rows or []
    cfg_regions = cfg_regions or []
    cfg_resources = cfg_resources or []
    cfg_plants = cfg_plants or []
    cfg_qcas = cfg_qcas or []
    cfg_presets = cfg_presets or []

    if cfg_regions and cfg_plants and len(cfg_regions) == len(cfg_plants):
        valid_configs = []
        for i in range(len(cfg_regions)):
            if cfg_regions[i] and cfg_plants[i]:
                valid_configs.append({
                    "region": cfg_regions[i],
                    "resource": cfg_resources[i] if i < len(cfg_resources) else "ALL",
                    "plant": cfg_plants[i],
                    "qca": cfg_qcas[i] if i < len(cfg_qcas) and cfg_qcas[i] else None,
                    "preset": cfg_presets[i] if i < len(cfg_presets) and cfg_presets[i] else "__CURRENT__",
                })
    else:
        valid_configs = [r for r in config_rows if r.get("region") and r.get("plant")]

    try:
        if valid_configs:
            base_settings = saved_settings if (isinstance(saved_settings, dict) and saved_settings) else {
                "err_mode": err_mode,
                "x_pct": x_pct,
                "bands": bands_rows or DEFAULT_BANDS.copy(),
            }
            name_to_settings = {p["name"]: p.get("settings", {}) for p in (presets_store or []) if isinstance(p, dict) and p.get("name")}

            def get_settings(preset_name):
                if preset_name == "__CURRENT__" or not preset_name:
                    return {
                        "err_mode": base_settings.get("err_mode", err_mode),
                        "x_pct": base_settings.get("x_pct", x_pct),
                        "bands": base_settings.get("bands") or bands_rows or DEFAULT_BANDS.copy(),
                        "unpaid_oi_threshold": base_settings.get("unpaid_oi_threshold", 15.0),
                    }
                st = name_to_settings.get(preset_name)
                return st if st else {"err_mode": err_mode, "x_pct": x_pct, "bands": bands_rows or DEFAULT_BANDS.copy(), "unpaid_oi_threshold": 15.0}

            results = []
            for row in valid_configs:
                r_regions = [row["region"]]
                r_plant = row["plant"]
                r_qcas = [row["qca"]] if row.get("qca") else None
                preset_name = row.get("preset") or "__CURRENT__"
                settings = get_settings(preset_name)
                res = _compute_pipeline(
                    r_regions, [r_plant], start_date, end_date,
                    settings.get("err_mode", err_mode),
                    settings.get("x_pct", x_pct),
                    settings.get("bands") or bands_rows or DEFAULT_BANDS.copy(),
                    float(settings.get("unpaid_oi_threshold", 15.0)),
                    qcas=r_qcas,
                )
                if "error" in res:
                    if r_qcas:
                        res = _compute_pipeline(
                            r_regions, [r_plant], start_date, end_date,
                            settings.get("err_mode", err_mode),
                            settings.get("x_pct", x_pct),
                            settings.get("bands") or bands_rows or DEFAULT_BANDS.copy(),
                            float(settings.get("unpaid_oi_threshold", 15.0)),
                            qcas=None,
                        )
                    if "error" in res:
                        return (res, {"display": "block"}, {"display": "none"})
                ps = res.get("plant_summary", pd.DataFrame())
                if isinstance(ps, pd.DataFrame) and not ps.empty:
                    ps = ps.copy()
                    ps["Custom Setting"] = preset_name if preset_name != "__CURRENT__" else "Current"
                res["plant_summary"] = ps
                res["_preset_name"] = preset_name if preset_name != "__CURRENT__" else "Current"
                res["_final_err_mode"] = settings.get("err_mode", err_mode)
                res["_final_x_pct"] = settings.get("x_pct", x_pct)
                res["_final_bands"] = settings.get("bands", bands_rows or DEFAULT_BANDS.copy())
                results.append(res)

            combined_ps = pd.concat([r["plant_summary"] for r in results if isinstance(r.get("plant_summary"), pd.DataFrame)], ignore_index=True)
            blockwise_parts = []
            for r in results:
                bw = r.get("blockwise")
                ps = r.get("plant_summary")
                if isinstance(bw, pd.DataFrame) and not bw.empty and isinstance(ps, pd.DataFrame) and not ps.empty:
                    plant_name = ps["Plant name"].iloc[0] if "Plant name" in ps.columns else ""
                    blockwise_parts.append(bw.assign(Plant=plant_name))
            combined_blockwise = pd.concat(blockwise_parts, ignore_index=True) if blockwise_parts else pd.DataFrame()
            detail_dfs = [r["df"] for r in results if "df" in r and not r["df"].empty]
            combined_df = pd.concat(detail_dfs, ignore_index=True) if detail_dfs else results[0]["df"]

            first = results[0]
            out_payload = {
                "df": combined_df.to_json(date_format="iso", orient="records") if hasattr(combined_df, "to_json") else combined_df,
                "kpis": first.get("kpis", {}),
                "blockwise": (combined_blockwise.to_dict("records") if hasattr(combined_blockwise, "to_dict") else combined_blockwise) if not combined_blockwise.empty else [],
                "plant_summary": combined_ps.to_dict("records") if hasattr(combined_ps, "to_dict") else combined_ps,
                "filters": {"regions": [r["region"] for r in valid_configs], "plants": [r["plant"] for r in valid_configs], "qcas": qcas, "start_date": str(start_date) if start_date else None, "end_date": str(end_date) if end_date else None},
                "used_settings": {"multi": [r["_preset_name"] for r in results]},
                "band_labels": [b.get("label", "no-label") for b in (first.get("final_bands") or [])[:5]],
                "final_bands": first.get("final_bands", []),
                "_all_runs": [
                    {"name": r["_preset_name"], "df": r["df"].to_json(date_format="iso", orient="records") if hasattr(r.get("df"), "to_json") else str(r.get("df", "")), "final_bands": r.get("_final_bands", []), "err_mode": r["_final_err_mode"], "x_pct": r["_final_x_pct"]}
                    for r in results
                ],
            }
            return (out_payload, {"display": "block"}, {"display": "none"})

        # Helper to run once with a given settings dict
        def run_once(with_settings: dict):
            final_err_mode = with_settings.get("err_mode", err_mode)
            final_x_pct = with_settings.get("x_pct", x_pct)
            saved_bands = with_settings.get("bands", None)
            final_bands = saved_bands if (saved_bands and len(saved_bands) > 0) else (bands_rows or DEFAULT_BANDS.copy())
            unpaid_threshold = float(with_settings.get("unpaid_oi_threshold", 15.0))
            print(f"DEBUG - Analysis using: err_mode={final_err_mode}, x_pct={final_x_pct}, bands_count={len(final_bands)}")
            print(f"DEBUG - Bands: {[b.get('label','no-label') for b in final_bands[:3]]}...")
            res = _compute_pipeline(regions, plants, start_date, end_date, final_err_mode, final_x_pct, final_bands, unpaid_threshold, qcas=qcas)
            return res, final_err_mode, final_x_pct, final_bands

        multi = []
        if selected_preset_names:
            name_to_settings = {p["name"]: p.get("settings", {}) for p in (presets_store or []) if isinstance(p, dict) and "name" in p}
            print(f"DEBUG - Running multi-preset analysis for: {selected_preset_names}")
            for nm in selected_preset_names:
                st = name_to_settings.get(nm)
                if not st:
                    print(f"DEBUG - Preset '{nm}' not found in store; skipping")
                    continue
                res, fm, fx, fb = run_once(st)
                if "error" in res:
                    return (res, {"display": "block"}, {"display": "none"})
                ps = res.get("plant_summary", pd.DataFrame())
                if isinstance(ps, pd.DataFrame) and not ps.empty:
                    ps = ps.copy()
                    ps["Custom Setting"] = nm
                    res["plant_summary"] = ps
                res["_preset_name"] = nm
                res["_final_err_mode"] = fm
                res["_final_x_pct"] = fx
                res["_final_bands"] = fb
                multi.append(res)
        else:
            # Backward-compatible single setting path (prefer saved settings if present)
            base_settings = saved_settings if (isinstance(saved_settings, dict) and saved_settings) else {
                "err_mode": err_mode,
                "x_pct": x_pct,
                "bands": bands_rows
            }
            res, fm, fx, fb = run_once(base_settings)
            if "error" in res:
                return (res, {"display": "block"}, {"display": "none"})
            ps = res.get("plant_summary", pd.DataFrame())
            if isinstance(ps, pd.DataFrame) and not ps.empty:
                ps = ps.copy()
                ps["Custom Setting"] = (base_settings.get("name") or base_settings.get("label") or "Current")
                res["plant_summary"] = ps
            res["_preset_name"] = ps["Custom Setting"].iloc[0] if isinstance(ps, pd.DataFrame) and not ps.empty else "Current"
            res["_final_err_mode"] = fm
            res["_final_x_pct"] = fx
            res["_final_bands"] = fb
            multi.append(res)

        combined_ps = pd.concat([
            m["plant_summary"] for m in multi if "plant_summary" in m and isinstance(m["plant_summary"], pd.DataFrame)
        ], ignore_index=True) if multi else pd.DataFrame()

        first = multi[0]
        out_payload = {
            "df": first["df"].to_json(date_format="iso", orient="records"),
            "kpis": first.get("kpis", {}),
            "blockwise": first.get("blockwise", pd.DataFrame()).to_dict("records"),
            "plant_summary": combined_ps.to_dict("records"),
            "filters": {
                "regions": regions,
                "plants": plants,
                "qcas": qcas,
                "start_date": str(start_date) if start_date else None,
                "end_date": str(end_date) if end_date else None,
            },
            "used_settings": {
                "multi": [m["_preset_name"] for m in multi],
            },
            "band_labels": [band.get("label", "no-label") for band in (first.get("final_bands") or first.get("_final_bands") or [])[:5]],
            "final_bands": first.get("final_bands", first.get("_final_bands", [])),
            "_all_runs": [
                {
                    "name": m["_preset_name"],
                    "df": m["df"].to_json(date_format="iso", orient="records"),
                    "final_bands": m.get("_final_bands", []),
                    "err_mode": m.get("_final_err_mode", "default"),
                    "x_pct": m.get("_final_x_pct", 0)
                } for m in multi
            ]
        }

        return (out_payload, {"display": "block"}, {"display": "none"})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return ({"error": str(e)}, {"display": "block"}, {"display": "none"})


@app.callback(
    Output("agg-results-store", "data"),
    Output("agg-results-section", "style"),
    Output("agg-progress-container", "style"),
    Input("agg-plot-now", "n_clicks"),
    State("agg-region-dd", "value"),
    State("agg-resource-type-dd", "value"),
    State("agg-plant-dd", "value"),
    State("agg-exclude-plant-dd", "value"),
    State("agg-qca-dd", "value"),
    State("agg-date-range", "start_date"),
    State("agg-date-range", "end_date"),
    State("err-mode", "value"),
    State("x-pct", "value"),
    State("bands-table", "data"),
    State("saved-settings-store", "data"),
    State("agg-analysis-preset-select", "value"),
    State("presets-store", "data"),
    State("agg-ppa-mode", "value"),
    State("agg-ppa-value", "value"),
    State("agg-per-plant-toggle", "value"),
    State("agg-plant-preset-table", "data"),
    prevent_initial_call=True
)
def compute_agg_on_click(
    n_plot,
    regions,
    resource_type,
    plants,
    excluded_plants,
    qcas,
    start_date,
    end_date,
    err_mode,
    x_pct,
    bands_rows,
    saved_settings,
    selected_preset_names,
    presets_store,
    agg_ppa_mode,
    agg_ppa_value,
    agg_per_plant_toggle,
    agg_plant_preset_data,
):
    if not n_plot:
        raise PreventUpdate

    use_agg_per_plant = agg_per_plant_toggle and "on" in agg_per_plant_toggle and agg_plant_preset_data

    try:
        if use_agg_per_plant:
            name_to_settings = {p["name"]: p.get("settings", {}) for p in (presets_store or []) if isinstance(p, dict) and "name" in p}
            base_settings = saved_settings if (isinstance(saved_settings, dict) and saved_settings) else {"err_mode": err_mode, "x_pct": x_pct, "bands": bands_rows or DEFAULT_BANDS.copy()}
            ppa_modes = [str(agg_ppa_mode)] if not isinstance(agg_ppa_mode, list) else (agg_ppa_mode or ["mean"])
            if isinstance(ppa_modes, list) and len(ppa_modes) == 0:
                ppa_modes = ["mean"]

            def _get_settings(pn):
                if pn == "__CURRENT__" or not pn:
                    return base_settings
                return name_to_settings.get(pn, base_settings)

            multi = []
            for row in agg_plant_preset_data:
                plant = row.get("Plant") or row.get("plant")
                preset_name = row.get("Preset") or "__CURRENT__"
                if not plant:
                    continue
                st = _get_settings(preset_name)
                fb = st.get("bands") or bands_rows or DEFAULT_BANDS.copy()
                fm = st.get("err_mode", err_mode)
                fx = st.get("x_pct", x_pct)
                unpaid = float(st.get("unpaid_oi_threshold", 15.0))
                pm = ppa_modes[0]
                res = _compute_pipeline_aggregated(regions, [plant], start_date, end_date, fm, fx, fb, pm, agg_ppa_value, unpaid, qcas=qcas)
                if "error" in res:
                    return (res, {"display": "block"}, {"display": "none"})
                ps = res.get("plant_summary", pd.DataFrame())
                if isinstance(ps, pd.DataFrame) and not ps.empty:
                    ps = ps.copy()
                    ps["Custom Setting"] = preset_name if preset_name != "__CURRENT__" else "Current"
                res["plant_summary"] = ps
                res["_preset_name"] = preset_name if preset_name != "__CURRENT__" else "Current"
                res["_final_err_mode"] = fm
                res["_final_x_pct"] = fx
                res["_final_bands"] = fb
                multi.append(res)

            if not multi:
                return ({"error": "No valid plant-preset rows."}, {"display": "block"}, {"display": "none"})
            combined_ps = pd.concat([r["plant_summary"] for r in multi if isinstance(r.get("plant_summary"), pd.DataFrame)], ignore_index=True)
            first = multi[0]
            out_payload = {
                "df": first["df"].to_json(date_format="iso", orient="records"),
                "kpis": first.get("kpis", {}),
                "blockwise": (first.get("blockwise", pd.DataFrame()) or pd.DataFrame()).to_dict("records"),
                "plant_summary": combined_ps.to_dict("records"),
                "filters": {"regions": regions, "plants": [row.get("Plant") for row in agg_plant_preset_data if row.get("Plant")], "qcas": qcas, "start_date": str(start_date), "end_date": str(end_date)},
                "used_settings": {"multi": [m["_preset_name"] for m in multi], "aggregated": True, "aggregated_plants": [row.get("Plant") for row in agg_plant_preset_data if row.get("Plant")], "ppa_modes": ppa_modes},
                "final_bands": first.get("_final_bands", []),
                "_all_runs": [{"name": m["_preset_name"], "df": m["df"].to_json(date_format="iso", orient="records"), "final_bands": m.get("_final_bands", []), "err_mode": m["_final_err_mode"], "x_pct": m["_final_x_pct"]} for m in multi],
            }
            return (out_payload, {"display": "block"}, {"display": "none"})

        # Apply exclusion logic: Final Plants = Selected Plants - Excluded Plants
        final_plants = plants
        if isinstance(plants, list):
            is_select_all = "SELECT_ALL" in plants
        else:
            is_select_all = (plants == "SELECT_ALL")
        
        if is_select_all:
            # Resolve SELECT_ALL to filtered plants (respecting resource type filter)
            try:
                sd = str(start_date) if start_date else str(datetime.now().date())
                ed = str(end_date) if end_date else sd
                all_plants = get_filtered_plants_by_type(regions, resource_type or "ALL", sd, ed)
            except Exception:
                # Fallback to all plants if filtering fails
                all_plants = get_plants_from_duckdb(regions)
            # Apply exclusion: remove excluded plants
            excluded_list = excluded_plants if (excluded_plants and isinstance(excluded_plants, list)) else (excluded_plants if excluded_plants else [])
            final_plants = [p for p in all_plants if p not in excluded_list]
            
            # Validation: ensure at least one plant remains
            if not final_plants:
                return (
                    {"error": "All plants have been excluded. Please revise your exclusion selection."},
                    {"display": "block"},
                    {"display": "none"}
                )
        else:
            # Manual selection - no exclusion applied
            final_plants = plants if isinstance(plants, list) else [plants]
        
        # Normalize PPA mode selection (multi-select dropdown)
        if agg_ppa_mode is None or agg_ppa_mode == "":
            ppa_modes = ["mean"]
        elif isinstance(agg_ppa_mode, list):
            ppa_modes = agg_ppa_mode if len(agg_ppa_mode) > 0 else ["mean"]
        else:
            ppa_modes = [str(agg_ppa_mode)]

        def _ppa_label(pm: str) -> str:
            pm = (pm or "").strip().lower()
            if pm == "median":
                return "PPA Median"
            if pm == "mode":
                return "PPA Mode"
            if pm == "weighted":
                return "PPA Weighted"
            if pm == "numeric":
                v = agg_ppa_value
                return f"PPA Numeric ({v})" if v is not None else "PPA Numeric"
            return "PPA Mean"

        def run_once(with_settings: dict, ppa_mode_one: str):
            final_err_mode = with_settings.get("err_mode", err_mode)
            final_x_pct = with_settings.get("x_pct", x_pct)
            saved_bands = with_settings.get("bands", None)
            final_bands = saved_bands if (saved_bands and len(saved_bands) > 0) else (bands_rows or DEFAULT_BANDS.copy())
            unpaid_threshold = float(with_settings.get("unpaid_oi_threshold", 15.0))
            ppa_mode_one = (ppa_mode_one or "mean").strip().lower()
            if ppa_mode_one == "numeric" and agg_ppa_value is None:
                return {"error": "Please enter a numeric PPA value for 'Numeric' mode."}, final_err_mode, final_x_pct, final_bands
            print(f"DEBUG - Aggregation Analysis using: err_mode={final_err_mode}, x_pct={final_x_pct}, bands_count={len(final_bands)}, ppa_mode={ppa_mode_one}")
            res = _compute_pipeline_aggregated(
                regions,
                final_plants,  # Use final_plants (with exclusions applied)
                start_date,
                end_date,
                final_err_mode,
                final_x_pct,
                final_bands,
                ppa_mode_one,
                agg_ppa_value,
                unpaid_threshold,
                qcas=qcas,
            )
            return res, final_err_mode, final_x_pct, final_bands

        multi = []
        if selected_preset_names:
            name_to_settings = {p["name"]: p.get("settings", {}) for p in (presets_store or []) if isinstance(p, dict) and "name" in p}
            print(f"DEBUG - Running multi-preset aggregation analysis for: {selected_preset_names}")
            for nm in selected_preset_names:
                st = name_to_settings.get(nm)
                if not st:
                    print(f"DEBUG - Preset '{nm}' not found in store; skipping (agg)")
                    continue
                for pm in ppa_modes:
                    res, fm, fx, fb = run_once(st, pm)
                    if "error" in res:
                        return (res, {"display": "block"}, {"display": "none"})
                    ps = res.get("plant_summary", pd.DataFrame())
                    scenario = f"{nm} - {_ppa_label(pm)}"
                    if isinstance(ps, pd.DataFrame) and not ps.empty:
                        ps = ps.copy()
                        ps["Custom Setting"] = scenario
                        res["plant_summary"] = ps
                    res["_preset_name"] = scenario
                    res["_final_err_mode"] = fm
                    res["_final_x_pct"] = fx
                    res["_final_bands"] = fb
                    res["_ppa_mode"] = pm
                    multi.append(res)
        else:
            base_settings = saved_settings if (isinstance(saved_settings, dict) and saved_settings) else {
                "err_mode": err_mode,
                "x_pct": x_pct,
                "bands": bands_rows
            }
            base_label = (base_settings.get("name") or base_settings.get("label") or "Current")
            for pm in ppa_modes:
                res, fm, fx, fb = run_once(base_settings, pm)
                if "error" in res:
                    return (res, {"display": "block"}, {"display": "none"})
                ps = res.get("plant_summary", pd.DataFrame())
                scenario = f"{base_label} - {_ppa_label(pm)}" if len(ppa_modes) > 1 else base_label
                if isinstance(ps, pd.DataFrame) and not ps.empty:
                    ps = ps.copy()
                    ps["Custom Setting"] = scenario
                    res["plant_summary"] = ps
                res["_preset_name"] = scenario
                res["_final_err_mode"] = fm
                res["_final_x_pct"] = fx
                res["_final_bands"] = fb
                res["_ppa_mode"] = pm
                multi.append(res)

        combined_ps = pd.concat([
            m["plant_summary"] for m in multi if "plant_summary" in m and isinstance(m["plant_summary"], pd.DataFrame)
        ], ignore_index=True) if multi else pd.DataFrame()

        first = multi[0]
        used_settings = {
            "multi": [m["_preset_name"] for m in multi],
            "aggregated": True,
            "aggregated_plants": first.get("aggregated_plants", []),
            "aggregated_region": first.get("aggregated_region", ""),
            "ppa_modes": ppa_modes,
        }
        out_payload = {
            "df": first["df"].to_json(date_format="iso", orient="records"),
            "kpis": first.get("kpis", {}),
            "blockwise": first.get("blockwise", pd.DataFrame()).to_dict("records"),
            "plant_summary": combined_ps.to_dict("records"),
            "filters": {
                "regions": regions,
                "plants": final_plants,
                "qcas": qcas,
                "start_date": str(start_date) if start_date else None,
                "end_date": str(end_date) if end_date else None,
            },
            "used_settings": used_settings,
            "band_labels": [band.get("label", "no-label") for band in (first.get("final_bands") or first.get("_final_bands") or [])[:5]],
            "final_bands": first.get("final_bands", first.get("_final_bands", [])),
            "_all_runs": [
                {
                    "name": m["_preset_name"],
                    "df": m["df"].to_json(date_format="iso", orient="records"),
                    "final_bands": m.get("_final_bands", []),
                    "err_mode": m.get("_final_err_mode", "default"),
                    "x_pct": m.get("_final_x_pct", 0),
                    "ppa_mode": m.get("_ppa_mode", None),
                    "ppa_value": agg_ppa_value,
                } for m in multi
            ]
        }

        return (out_payload, {"display": "block"}, {"display": "none"})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return ({"error": str(e)}, {"display": "block"}, {"display": "none"})

# --------- Stats availability compute ---------
@app.callback(
    Output("stats-results", "children"),
    Input("stats-run", "n_clicks"),
    State("stats-region-dd", "value"),
    State("stats-plant-dd", "value"),
    State("stats-qca-dd", "value"),
    State("stats-date-range", "start_date"),
    State("stats-date-range", "end_date"),
    prevent_initial_call=True
)
def run_stats(n, regions, plants, qcas, start_date, end_date):
    if not n:
        raise PreventUpdate
    if not regions or not plants:
        return dbc.Alert("Select regions and plants", color="warning")

    if "SELECT_ALL" in (plants or []):
        plants = get_plants_from_duckdb(regions)

    qca_filter = [q for q in (qcas or []) if q and q != "SELECT_ALL"]
    df = load_region_data(regions, str(start_date), str(end_date), plants, qcas=qca_filter or None)
    if df.empty:
        return dbc.Alert("No data for selection", color="warning")

    # availability logic: consider rows where Scheduled_MW>0 then check PPA and Actual_MW present
    df["Scheduled_MW"] = pd.to_numeric(df["Scheduled_MW"], errors="coerce").fillna(0)
    df["Actual_MW"] = pd.to_numeric(df["Actual_MW"], errors="coerce")
    df["PPA"] = pd.to_numeric(df["PPA"], errors="coerce")

    df["is_candidate"] = df["Scheduled_MW"] > 0
    df["is_ok"] = df["is_candidate"] & df["Actual_MW"].notna() & df["PPA"].notna()

    def summarize_availability(pdf):
        cand = int(pdf["is_candidate"].sum())
        ok = int(pdf["is_ok"].sum())
        pct = (ok / cand * 100.0) if cand > 0 else 0.0
        missing = cand - ok
        return pd.Series({"availability": round(pct, 2), "missing": int(missing)})

    stats = df.groupby("Plant").apply(summarize_availability).reset_index().rename(columns={"Plant": "Plant name"})
    stats["Date Range"] = f"{start_date} → {end_date}"
    # Data coverage per plant (after all filters)
    try:
        cov = (
            df.groupby("Plant")["date"]
            .agg(["min", "max"])
            .reset_index()
            .rename(columns={"Plant": "Plant name", "min": "Data Available From", "max": "Data Available To"})
        )
        stats = stats.merge(cov, on="Plant name", how="left")
    except Exception:
        stats["Data Available From"] = None
        stats["Data Available To"] = None

    stats = stats[["Plant name", "Date Range", "Data Available From", "Data Available To", "availability", "missing"]]

    table = dash_table.DataTable(
        columns=[{"name": c, "id": c} for c in stats.columns],
        data=stats.to_dict("records"),
        sort_action="native",
        style_table={"overflowX": "auto"},
        style_cell={"padding": "6px"},
        style_header={"fontWeight": "600"},
        page_size=15,
    )
    return table

# Callback to show progress bar when plot button is clicked
@app.callback(
    Output("progress-container", "style", allow_duplicate=True),
    Input("plot-now", "n_clicks"),
    prevent_initial_call=True
)
def show_progress_bar(n):
    """Show progress bar when analysis starts"""
    if not n:
        raise PreventUpdate
    return {"display": "block", "marginTop": "1.5rem"}


@app.callback(
    Output("agg-progress-container", "style", allow_duplicate=True),
    Input("agg-plot-now", "n_clicks"),
    prevent_initial_call=True
)
def show_agg_progress_bar(n_plot):
    """Show progress bar when aggregation analysis starts."""
    if not n_plot:
        raise PreventUpdate
    return {"display": "block", "marginTop": "1.5rem"}

## Removed separate nav-store updater to avoid duplicate outputs; handled in switch_nav_tabs

@app.callback(
    Output("tab-content", "children"),
    Input("results-store", "data"),
)
def render_tabs(stored):
    try:
        if not stored or (isinstance(stored, dict) and stored.get("error")):
            err = stored.get("error") if isinstance(stored, dict) else ""
            msg = err or "Click Plot Now to compute."
            return dbc.Alert(msg, color="warning")

        df = pd.DataFrame(json.loads(stored["df"])) if isinstance(stored.get("df"), str) else pd.DataFrame(stored.get("df", []))
        plant_summary_df = pd.DataFrame(stored.get("plant_summary", []))
        used = stored.get("used_settings", {})
        filters = stored.get("filters", {}) if isinstance(stored, dict) else {}

        if df.empty:
            return dbc.Alert("No data found for the selected filters.", color="warning")

        # Analysis view simplified to only Plant Summary

        # Plant summary table
        plant_summary_table = dash_table.DataTable(
            columns=[
                {"name": c, "id": c, "type": "numeric", "format": {"specifier": ".2f"}} 
                if c in ["PPA", "Revenue Loss (%)", "Revenue Loss (p/k)", "Plant Capacity", "DSM Loss"]
                else {"name": c, "id": c}
                for c in plant_summary_df.columns
            ] if not plant_summary_df.empty else [],
            data=plant_summary_df.to_dict("records"),
            sort_action="native",
            style_table={"overflowX": "auto"},
            style_cell={"padding": "6px", "textAlign": "left"},
            style_header={"fontWeight": "600", "backgroundColor": "#f8f9fa"},
        )

        # Optional comparison chart
        graph_div = None
        chart_heading = None
        try:
            if not plant_summary_df.empty and ("Plant name" in plant_summary_df.columns) and ("Revenue Loss (%)" in plant_summary_df.columns):
                ps = plant_summary_df.copy()
                ps = ps.sort_values("Revenue Loss (%)")
                print(f"DEBUG - Creating chart with {len(ps)} rows, columns: {list(ps.columns)}")
                
                # Get X threshold values for each preset from stored data
                preset_x_map = {}
                all_runs = stored.get("_all_runs", []) if isinstance(stored, dict) else []
                for run in all_runs:
                    preset_name = run.get("name", "")
                    err_mode = str(run.get("err_mode", "default")).lower()
                    x_pct = float(run.get("x_pct", 100))
                    # For default mode, X is effectively 100%
                    x_threshold = 100.0 if err_mode == "default" else x_pct
                    preset_x_map[preset_name] = {
                        "x_value": x_threshold,
                        "err_mode": err_mode,
                        "x_pct": x_pct
                    }
                print(f"DEBUG - Preset X thresholds: {preset_x_map}")
                
                # Build per-row X label for bar text
                def label_for_setting(setting_name: str) -> str:
                    info = preset_x_map.get(setting_name, None)
                    if not info:
                        return ""
                    x_val = int(round(float(info.get("x_value", 0))))
                    return f"X = {x_val} %"

                if "Custom Setting" in ps.columns:
                    ps["X Label"] = ps["Custom Setting"].apply(label_for_setting)
                    fig = px.bar(
                        ps,
                        x="Revenue Loss (%)",
                        y="Plant name",
                        color="Custom Setting",
                        text="X Label",
                        orientation="h",
                        barmode="group",
                    )
                    fig.update_layout(legend_title_text="Setting")
                else:
                    # Single setting path - use first run's X from preset_x_map if present
                    if len(preset_x_map) > 0:
                        single_name = list(preset_x_map.keys())[0]
                        x_val = int(round(float(preset_x_map[single_name].get("x_value", 0))))
                        ps["X Label"] = [f"X = {x_val} %"] * len(ps)
                    else:
                        ps["X Label"] = [""] * len(ps)
                    fig = px.bar(
                        ps,
                        x="Revenue Loss (%)",
                        y="Plant name",
                        text="X Label",
                        orientation="h",
                    )
                
                # (Removed) Vertical threshold lines and top annotations per request
                
                # Ensure text labels are outside and colored like the bar
                for tr in getattr(fig, "data", []):
                    try:
                        tr.textposition = "outside"
                        if hasattr(tr, "marker") and hasattr(tr.marker, "color"):
                            tr.textfont.color = tr.marker.color
                            tr.textfont.size = 12
                    except Exception:
                        pass

                n_rows = len(ps.index)
                fig.update_layout(
                    template="plotly_white",
                    title="Revenue Loss % Comparison",
                    xaxis_title="Revenue Loss (%)",
                    yaxis_title="Plant Name",
                    margin=dict(l=150, r=60, t=60, b=60),  # Reduced top margin; no annotations/threshold lines
                    height=max(450, 220 + 40 * n_rows),
                    bargap=0.25,
                )
                chart_height = max(450, 220 + 40 * n_rows)
                graph_div = dcc.Graph(
                    id="revenue-loss-chart",
                    figure=fig,
                    config={"displayModeBar": True},
                    style={"width": "100%", "height": f"{chart_height}px"}
                )
                chart_heading = html.H6("Revenue Loss % — Comparison", 
                                       style={"marginTop": "1rem", "marginBottom": "0.5rem", "color": "#333", "fontWeight": "600"})
                print(f"DEBUG - Chart created successfully, height={chart_height}")
        except Exception as e:
            import traceback
            print(f"DEBUG - Chart build failed: {e}")
            traceback.print_exc()

        download_btn = dbc.Button("Download Full Calculation (Excel)", id="btn-download", color="primary", className="mt-2")

        # Debug: check if graph was created
        print(f"DEBUG - graph_div is None: {graph_div is None}")
        print(f"DEBUG - chart_heading is None: {chart_heading is None}")
        if graph_div:
            print(f"DEBUG - graph_div type: {type(graph_div)}")
            print(f"DEBUG - graph_div.id: {getattr(graph_div, 'id', 'no id')}")

        # Build children list - put chart BEFORE table
        children_list = [
            html.H5("Plant Summary", style={"marginTop": "0.5rem", "marginBottom": "1rem", "color": "#333"}),
        ]

        # -------------------------
        # Portfolio KPI cards (selected filters only)
        # -------------------------
        try:
            plant_col_kpi = "Plant" if "Plant" in df.columns else ("plant_name" if "plant_name" in df.columns else None)
            total_plants_kpi = int(pd.Series(df[plant_col_kpi]).nunique()) if plant_col_kpi else 0
            regions_kpi = int(pd.Series(df.get("region")).nunique()) if "region" in df.columns else 0
            qca_kpi = int(pd.Series(df.get("qca")).dropna().astype(str).str.strip().replace("", pd.NA).dropna().nunique()) if "qca" in df.columns else 0

            dts = pd.to_datetime(df.get("date"), errors="coerce")
            cov_from = dts.dt.date.min() if hasattr(dts, "dt") else None
            cov_to = dts.dt.date.max() if hasattr(dts, "dt") else None
            coverage = f"{cov_from} → {cov_to}" if cov_from and cov_to else "—"

            penalty_sum = None
            if "penalty" in df.columns:
                penalty_sum = float(pd.to_numeric(df["penalty"], errors="coerce").fillna(0.0).sum())

            children_list.append(
                dbc.Row(
                    [
                        dbc.Col(kpi_card("Total plants", f"{total_plants_kpi}"), md=2),
                        dbc.Col(kpi_card("Regions", f"{regions_kpi}"), md=2),
                        dbc.Col(kpi_card("QCAs", f"{qca_kpi}"), md=2),
                        dbc.Col(kpi_card("Coverage", coverage), md=3),
                        dbc.Col(
                            kpi_card("DSM penalty (sum)", f"{penalty_sum:,.2f}" if penalty_sum is not None else "—"),
                            md=3,
                        ),
                    ],
                    className="mb-3 g-3",
                )
            )
        except Exception as e:
            print(f"DEBUG - KPI build failed: {e}")
        
        # Add chart heading and graph if available (BEFORE table)
        if graph_div is not None:
            print(f"DEBUG - Adding chart to display...")
            if chart_heading:
                children_list.append(chart_heading)
            # Add graph with minimal wrapper
            children_list.append(graph_div)
            print(f"DEBUG - Chart added! Total children before table: {len(children_list)}")

        # -------------------------
        # Heatmap + Availability Pie (selected filters only)
        # -------------------------
        try:
            # Selected date range (prefer stored filters; fallback to df coverage)
            sd = filters.get("start_date") or str(pd.to_datetime(df["date"], errors="coerce").min().date())
            ed = filters.get("end_date") or str(pd.to_datetime(df["date"], errors="coerce").max().date())
            sd_d = pd.to_datetime(sd, errors="coerce").date()
            ed_d = pd.to_datetime(ed, errors="coerce").date()
            if ed_d < sd_d:
                ed_d = sd_d
            days = int((ed_d - sd_d).days) + 1

            # Expected blocks in range (per-plant, 96 blocks/day)
            plant_col = "Plant" if "Plant" in df.columns else "plant_name"
            n_plants = int(pd.Series(df[plant_col]).nunique()) if plant_col in df.columns else 1
            expected = max(0, days * 96 * max(1, n_plants))
            present = int(df[[plant_col, "date", "time_block"]].drop_duplicates().shape[0])
            missing = max(0, expected - present)

            pie_df = pd.DataFrame({"Status": ["Present blocks", "Missing blocks"], "Blocks": [present, missing]})
            pie_fig = px.pie(
                pie_df,
                names="Status",
                values="Blocks",
                title=f"Data Availability ({sd_d} → {ed_d})",
                hole=0.45,
            )
            pie_fig.update_layout(template="plotly_white", height=340, margin=dict(l=10, r=10, t=55, b=10))

            # Heatmap intensity
            hm = df.copy()
            hm["date"] = pd.to_datetime(hm["date"], errors="coerce").dt.date
            hm["time_block"] = pd.to_numeric(hm.get("time_block"), errors="coerce")
            if "penalty" in hm.columns:
                hm["intensity"] = pd.to_numeric(hm["penalty"], errors="coerce").fillna(0.0).abs()
                metric_name = "ABS(Penalty)"
            else:
                hm["intensity"] = pd.to_numeric(hm.get("error_pct"), errors="coerce").fillna(0.0).abs()
                metric_name = "ABS(Error%)"

            heat_long = (
                hm.dropna(subset=["date", "time_block"])
                .groupby(["date", "time_block"], as_index=False)["intensity"]
                .sum()
            )
            heat_fig = px.density_heatmap(
                heat_long,
                x="date",
                y="time_block",
                z="intensity",
                histfunc="sum",
                color_continuous_scale="YlOrRd",
                title=f"DSM Heatmap (selected filters) — {metric_name} (sum)",
            )
            heat_fig.update_layout(template="plotly_white", height=460, margin=dict(l=30, r=10, t=55, b=30))
            heat_fig.update_yaxes(title="Time Block", autorange="reversed")
            heat_fig.update_xaxes(title="Date")

            children_list.append(
                html.H6(
                    "DSM Heatmap & Data Availability",
                    style={"marginTop": "1.25rem", "marginBottom": "0.5rem", "color": "#333", "fontWeight": "600"},
                )
            )
            children_list.append(
                dbc.Row(
                    [
                        dbc.Col(dcc.Graph(id="dsm-heatmap", figure=heat_fig, config={"displayModeBar": True}), md=8),
                        dbc.Col(dcc.Graph(id="data-availability-pie", figure=pie_fig, config={"displayModeBar": False}), md=4),
                    ],
                    className="mb-3",
                )
            )
        except Exception as e:
            print(f"DEBUG - Heatmap/Pie build failed: {e}")

        # -------------------------
        # QCA Analytics (penalty & plants by QCA, selected filters only)
        # -------------------------
        try:
            qca_col = "QCA" if "QCA" in df.columns else ("qca" if "qca" in df.columns else None)
            if qca_col and qca_col in df.columns:
                qca_vals = df[qca_col].dropna().astype(str).str.strip()
                qca_vals = qca_vals[qca_vals != ""]
                if not qca_vals.empty:
                    qca_agg = df.loc[qca_vals.index].copy()
                    qca_agg["_qca_label"] = qca_agg[qca_col].astype(str).str.strip()
                    plant_col_qca = "Plant" if "Plant" in qca_agg.columns else "plant_name"
                    qca_agg["_penalty_num"] = pd.to_numeric(qca_agg.get("penalty"), errors="coerce").fillna(0.0) if "penalty" in qca_agg.columns else 0.0
                    agg_dict = {"Blocks": ("_qca_label", "size"), "Penalty_Sum": ("_penalty_num", "sum")}
                    if plant_col_qca in qca_agg.columns:
                        agg_dict["Plants"] = (plant_col_qca, "nunique")
                    else:
                        agg_dict["Plants"] = ("_qca_label", "size")
                    agg_by_qca = qca_agg.groupby("_qca_label", as_index=False).agg(**{k: v for k, v in agg_dict.items()})
                    agg_by_qca = agg_by_qca.rename(columns={"_qca_label": "QCA", "Penalty_Sum": "Penalty (₹)"})
                    agg_by_qca = agg_by_qca.sort_values("Penalty (₹)", ascending=False)

                    qca_fig = px.bar(
                        agg_by_qca,
                        x="QCA",
                        y="Penalty (₹)",
                        title="DSM Penalty by QCA (selected filters)",
                        text="Penalty (₹)",
                        text_auto=",.0f",
                    )
                    qca_fig.update_layout(template="plotly_white", height=320, margin=dict(l=10, r=10, t=50, b=80))
                    qca_fig.update_xaxes(tickangle=-45)

                    qca_table = dash_table.DataTable(
                        columns=[
                            {"name": c, "id": c, "type": "numeric", "format": {"specifier": ",.2f"}}
                            if c == "Penalty (₹)" else {"name": c, "id": c}
                            for c in agg_by_qca.columns
                        ],
                        data=agg_by_qca.to_dict("records"),
                        sort_action="native",
                        style_table={"overflowX": "auto"},
                        style_cell={"padding": "6px", "textAlign": "left"},
                        style_header={"fontWeight": "600", "backgroundColor": "#f8f9fa"},
                    )
                    children_list.append(
                        html.H6(
                            "QCA Analytics",
                            style={"marginTop": "1.25rem", "marginBottom": "0.5rem", "color": "#333", "fontWeight": "600"},
                        )
                    )
                    children_list.append(
                        dbc.Row(
                            [
                                dbc.Col(dcc.Graph(id="qca-penalty-chart", figure=qca_fig, config={"displayModeBar": True}), md=7),
                                dbc.Col(html.Div([html.H6("Summary by QCA", className="mb-2"), qca_table], style={"fontSize": "0.9rem"}), md=5),
                            ],
                            className="mb-3",
                        )
                    )
        except Exception as e:
            print(f"DEBUG - QCA analytics failed: {e}")

        # -------------------------
        # Data Health (selected filters only)
        # -------------------------
        try:
            hs, extra = summarize_health_from_df(
                df,
                start_date=filters.get("start_date"),
                end_date=filters.get("end_date"),
                plant_col="Plant",
            )
            health_lines = []
            if hs.missing_blocks > 0:
                health_lines.append(
                    f"Missing blocks: {hs.missing_blocks:,} (expected {hs.expected_blocks:,}, present {hs.present_blocks:,})"
                )
            if hs.ppa_missing_rows > 0:
                health_lines.append(f"Missing/zero PPA rows: {hs.ppa_missing_rows:,}")
            if hs.avc_zero_rows > 0:
                health_lines.append(f"AVC = 0 rows: {hs.avc_zero_rows:,}")
            if hs.all_zero_rows > 0:
                health_lines.append(
                    f"All-zero rows (AVC=0 & Schedule=0 & Actual=0): {hs.all_zero_rows:,}"
                )

            if not health_lines:
                health_alert = dbc.Alert(
                    "No data quality issues detected for the selected filters.", color="success"
                )
            else:
                health_alert = dbc.Alert([html.Div(x) for x in health_lines], color="warning")

            children_list.append(
                html.H6(
                    "Data Health (selected filters)",
                    style={
                        "marginTop": "0.5rem",
                        "marginBottom": "0.5rem",
                        "color": "#333",
                        "fontWeight": "600",
                    },
                )
            )
            children_list.append(health_alert)

            # Optional drilldown table
            per_plant = extra.get("per_plant") if isinstance(extra, dict) else None
            if isinstance(per_plant, pd.DataFrame) and not per_plant.empty:
                drill = per_plant.copy()
                # Prefer showing worst plants first (missing blocks proxy: fewer present blocks)
                if "present_blocks" in drill.columns:
                    drill = drill.sort_values("present_blocks")
                drill_table = dash_table.DataTable(
                    columns=[{"name": c, "id": c} for c in drill.columns],
                    data=drill.to_dict("records"),
                    sort_action="native",
                    page_size=10,
                    style_table={"overflowX": "auto"},
                    style_cell={"padding": "6px", "textAlign": "left"},
                    style_header={"fontWeight": "600", "backgroundColor": "#f8f9fa"},
                )
                children_list.append(
                    html.Details(
                        [
                            html.Summary("Show per-plant health details", style={"cursor": "pointer", "fontWeight": "600"}),
                            html.Div(drill_table, style={"marginTop": "0.75rem"}),
                        ],
                        open=False,
                        style={"marginBottom": "0.75rem"},
                    )
                )
        except Exception as e:
            print(f"DEBUG - Health panel failed: {e}")
        
        # Add table
        children_list.append(
            dbc.Row([dbc.Col(plant_summary_table, md=12)])
        )
        
        # Add download button
        children_list.append(
            html.Div(download_btn, className="text-end mt-2")
        )

        return dcc.Loading(
            type="circle",
            children=html.Div(children_list)
        )
    except Exception as e:
        return dbc.Alert(f"Failed to render results: {e}", color="danger")


@app.callback(
    Output("agg-tab-content", "children"),
    Input("agg-results-store", "data"),
)
def render_agg_tabs(stored):
    """Render results for Aggregation Analysis using the same layout as main Analysis,
    with an additional note indicating the plants included in the aggregation.
    """
    try:
        if not stored or (isinstance(stored, dict) and stored.get("error")):
            err = stored.get("error") if isinstance(stored, dict) else ""
            msg = err or "Click Plot Now to compute."
            return dbc.Alert(msg, color="warning")

        df = pd.DataFrame(json.loads(stored["df"])) if isinstance(stored.get("df"), str) else pd.DataFrame(stored.get("df", []))
        plant_summary_df = pd.DataFrame(stored.get("plant_summary", []))
        used = stored.get("used_settings", {}) if isinstance(stored, dict) else {}
        filters = stored.get("filters", {}) if isinstance(stored, dict) else {}

        if df.empty:
            return dbc.Alert("No data found for the selected filters.", color="warning")

        # Plant summary table (aggregated)
        plant_summary_table = dash_table.DataTable(
            columns=[
                {"name": c, "id": c, "type": "numeric", "format": {"specifier": ".2f"}} 
                if c in ["PPA", "Revenue Loss (%)", "Revenue Loss (p/k)", "Plant Capacity", "DSM Loss"]
                else {"name": c, "id": c}
                for c in plant_summary_df.columns
            ] if not plant_summary_df.empty else [],
            data=plant_summary_df.to_dict("records"),
            sort_action="native",
            style_table={"overflowX": "auto"},
            style_cell={"padding": "6px", "textAlign": "left"},
            style_header={"fontWeight": "600", "backgroundColor": "#f8f9fa"},
        )

        # Optional comparison chart (still valid for multiple presets, but on aggregated profile)
        graph_div = None
        chart_heading = None
        try:
            if not plant_summary_df.empty and ("Plant name" in plant_summary_df.columns) and ("Revenue Loss (%)" in plant_summary_df.columns):
                ps = plant_summary_df.copy()
                ps = ps.sort_values("Revenue Loss (%)")
                print(f"DEBUG - Creating aggregation chart with {len(ps)} rows, columns: {list(ps.columns)}")

                preset_x_map = {}
                all_runs = stored.get("_all_runs", []) if isinstance(stored, dict) else []
                for run in all_runs:
                    preset_name = run.get("name", "")
                    err_mode = str(run.get("err_mode", "default")).lower()
                    x_pct = float(run.get("x_pct", 100))
                    x_threshold = 100.0 if err_mode == "default" else x_pct
                    preset_x_map[preset_name] = {
                        "x_value": x_threshold,
                        "err_mode": err_mode,
                        "x_pct": x_pct
                    }

                def label_for_setting(setting_name: str) -> str:
                    info = preset_x_map.get(setting_name, None)
                    if not info:
                        return ""
                    x_val = int(round(float(info.get("x_value", 0))))
                    return f"X = {x_val} %"

                if "Custom Setting" in ps.columns:
                    ps["X Label"] = ps["Custom Setting"].apply(label_for_setting)
                    fig = px.bar(
                        ps,
                        x="Revenue Loss (%)",
                        y="Plant name",
                        color="Custom Setting",
                        text="X Label",
                        orientation="h",
                        barmode="group",
                    )
                    fig.update_layout(legend_title_text="Setting")
                else:
                    if len(preset_x_map) > 0:
                        single_name = list(preset_x_map.keys())[0]
                        x_val = int(round(float(preset_x_map[single_name].get("x_value", 0))))
                        ps["X Label"] = [f"X = {x_val} %"] * len(ps)
                    else:
                        ps["X Label"] = [""] * len(ps)
                    fig = px.bar(
                        ps,
                        x="Revenue Loss (%)",
                        y="Plant name",
                        text="X Label",
                        orientation="h",
                    )

                for tr in getattr(fig, "data", []):
                    try:
                        tr.textposition = "outside"
                        if hasattr(tr, "marker") and hasattr(tr.marker, "color"):
                            tr.textfont.color = tr.marker.color
                            tr.textfont.size = 12
                    except Exception:
                        pass

                n_rows = len(ps.index)
                fig.update_layout(
                    template="plotly_white",
                    title="Revenue Loss % Comparison (Aggregated Profile)",
                    xaxis_title="Revenue Loss (%)",
                    yaxis_title="Plant Name",
                    margin=dict(l=150, r=60, t=60, b=60),
                    height=max(450, 220 + 40 * n_rows),
                    bargap=0.25,
                )
                chart_height = max(450, 220 + 40 * n_rows)
                graph_div = dcc.Graph(
                    id="agg-revenue-loss-chart",
                    figure=fig,
                    config={"displayModeBar": True},
                    style={"width": "100%", "height": f"{chart_height}px"}
                )
                chart_heading = html.H6(
                    "Revenue Loss % — Aggregated Comparison", 
                    style={"marginTop": "1rem", "marginBottom": "0.5rem", "color": "#333", "fontWeight": "600"}
                )
        except Exception as e:
            import traceback
            print(f"DEBUG - Aggregation chart build failed: {e}")
            traceback.print_exc()

        download_btn = dbc.Button("Download Full Calculation (Excel)", id="agg-btn-download", color="primary", className="mt-2")

        # Aggregation note
        agg_plants = used.get("aggregated_plants") or []
        agg_region = used.get("aggregated_region") or ""
        note_children = []
        if agg_plants:
            plants_text = ", ".join(map(str, agg_plants))
            note_text = f"Aggregated output for {len(agg_plants)} plants: {plants_text}"
            if agg_region:
                note_text = f"{note_text} (Region: {agg_region})"
            note_children.append(
                html.Div(
                    note_text,
                    style={
                        "fontSize": "0.9rem",
                        "color": "#555",
                        "backgroundColor": "#f8f9fa",
                        "borderRadius": "6px",
                        "padding": "8px 12px",
                        "marginBottom": "0.75rem",
                        "border": "1px dashed #ddd",
                    },
                )
            )

        children_list = note_children + [
            html.H5("Plant Summary (Aggregated)", style={"marginTop": "0.5rem", "marginBottom": "1rem", "color": "#333"}),
        ]

        # -------------------------
        # Portfolio KPI cards (selected filters only)
        # -------------------------
        try:
            total_plants_kpi = 0
            flt_plants = filters.get("plants")
            if isinstance(flt_plants, list):
                total_plants_kpi = len([p for p in flt_plants if p and p != "SELECT_ALL"])
            elif flt_plants and flt_plants != "SELECT_ALL":
                total_plants_kpi = 1
            if total_plants_kpi <= 0:
                total_plants_kpi = len(agg_plants) if agg_plants else 0

            regions_kpi = 0
            flt_regions = filters.get("regions")
            if isinstance(flt_regions, list):
                regions_kpi = len([r for r in flt_regions if r and r != "SELECT_ALL"])
            elif flt_regions and flt_regions != "SELECT_ALL":
                regions_kpi = 1

            qca_kpi = int(pd.Series(df.get("qca")).dropna().astype(str).str.strip().replace("", pd.NA).dropna().nunique()) if "qca" in df.columns else 0
            dts = pd.to_datetime(df.get("date"), errors="coerce")
            cov_from = dts.dt.date.min() if hasattr(dts, "dt") else None
            cov_to = dts.dt.date.max() if hasattr(dts, "dt") else None
            coverage = f"{cov_from} → {cov_to}" if cov_from and cov_to else "—"
            penalty_sum = float(pd.to_numeric(df["penalty"], errors="coerce").fillna(0.0).sum()) if "penalty" in df.columns else None

            children_list.append(
                dbc.Row(
                    [
                        dbc.Col(kpi_card("Plants (agg)", f"{total_plants_kpi}"), md=2),
                        dbc.Col(kpi_card("Regions", f"{regions_kpi}"), md=2),
                        dbc.Col(kpi_card("QCAs", f"{qca_kpi}"), md=2),
                        dbc.Col(kpi_card("Coverage", coverage), md=3),
                        dbc.Col(
                            kpi_card("DSM penalty (sum)", f"{penalty_sum:,.2f}" if penalty_sum is not None else "—"),
                            md=3,
                        ),
                    ],
                    className="mb-3 g-3",
                )
            )
        except Exception as e:
            print(f"DEBUG - (agg) KPI build failed: {e}")

        if graph_div is not None:
            if chart_heading:
                children_list.append(chart_heading)
            children_list.append(graph_div)

        # -------------------------
        # Heatmap + Availability Pie (selected filters only)
        # -------------------------
        try:
            sd = filters.get("start_date") or str(pd.to_datetime(df["date"], errors="coerce").min().date())
            ed = filters.get("end_date") or str(pd.to_datetime(df["date"], errors="coerce").max().date())
            sd_d = pd.to_datetime(sd, errors="coerce").date()
            ed_d = pd.to_datetime(ed, errors="coerce").date()
            if ed_d < sd_d:
                ed_d = sd_d
            days = int((ed_d - sd_d).days) + 1

            expected = max(0, days * 96)  # aggregated profile = single time series
            present = int(df[["date", "time_block"]].drop_duplicates().shape[0])
            missing = max(0, expected - present)

            pie_df = pd.DataFrame({"Status": ["Present blocks", "Missing blocks"], "Blocks": [present, missing]})
            pie_fig = px.pie(
                pie_df,
                names="Status",
                values="Blocks",
                title=f"Data Availability ({sd_d} → {ed_d})",
                hole=0.45,
            )
            pie_fig.update_layout(template="plotly_white", height=340, margin=dict(l=10, r=10, t=55, b=10))

            hm = df.copy()
            hm["date"] = pd.to_datetime(hm["date"], errors="coerce").dt.date
            hm["time_block"] = pd.to_numeric(hm.get("time_block"), errors="coerce")
            if "penalty" in hm.columns:
                hm["intensity"] = pd.to_numeric(hm["penalty"], errors="coerce").fillna(0.0).abs()
                metric_name = "ABS(Penalty)"
            else:
                hm["intensity"] = pd.to_numeric(hm.get("error_pct"), errors="coerce").fillna(0.0).abs()
                metric_name = "ABS(Error%)"

            heat_long = (
                hm.dropna(subset=["date", "time_block"])
                .groupby(["date", "time_block"], as_index=False)["intensity"]
                .sum()
            )
            heat_fig = px.density_heatmap(
                heat_long,
                x="date",
                y="time_block",
                z="intensity",
                histfunc="sum",
                color_continuous_scale="YlOrRd",
                title=f"DSM Heatmap (Aggregated) — {metric_name} (sum)",
            )
            heat_fig.update_layout(template="plotly_white", height=460, margin=dict(l=30, r=10, t=55, b=30))
            heat_fig.update_yaxes(title="Time Block", autorange="reversed")
            heat_fig.update_xaxes(title="Date")

            children_list.append(
                html.H6(
                    "DSM Heatmap & Data Availability",
                    style={"marginTop": "1.25rem", "marginBottom": "0.5rem", "color": "#333", "fontWeight": "600"},
                )
            )
            children_list.append(
                dbc.Row(
                    [
                        dbc.Col(dcc.Graph(id="agg-dsm-heatmap", figure=heat_fig, config={"displayModeBar": True}), md=8),
                        dbc.Col(dcc.Graph(id="agg-data-availability-pie", figure=pie_fig, config={"displayModeBar": False}), md=4),
                    ],
                    className="mb-3",
                )
            )
        except Exception as e:
            print(f"DEBUG - (agg) Heatmap/Pie build failed: {e}")

        # -------------------------
        # QCA Analytics (aggregated profile - single time series, may have 0–1 QCA)
        # -------------------------
        try:
            qca_col = "QCA" if "QCA" in df.columns else ("qca" if "qca" in df.columns else None)
            if qca_col and qca_col in df.columns:
                qca_vals = df[qca_col].dropna().astype(str).str.strip()
                qca_vals = qca_vals[qca_vals != ""]
                if not qca_vals.empty:
                    qca_agg = df.loc[qca_vals.index].copy()
                    qca_agg["_qca_label"] = qca_agg[qca_col].astype(str).str.strip()
                    qca_agg["_penalty_num"] = pd.to_numeric(qca_agg.get("penalty"), errors="coerce").fillna(0.0) if "penalty" in qca_agg.columns else 0.0
                    agg_by_qca = qca_agg.groupby("_qca_label", as_index=False).agg(
                        Blocks=("_qca_label", "size"),
                        Penalty_Sum=("_penalty_num", "sum"),
                    )
                    agg_by_qca = agg_by_qca.rename(columns={"_qca_label": "QCA", "Penalty_Sum": "Penalty (₹)"})
                    agg_by_qca = agg_by_qca.sort_values("Penalty (₹)", ascending=False)

                    qca_fig = px.bar(
                        agg_by_qca,
                        x="QCA",
                        y="Penalty (₹)",
                        title="DSM Penalty by QCA (Aggregated)",
                        text="Penalty (₹)",
                        text_auto=",.0f",
                    )
                    qca_fig.update_layout(template="plotly_white", height=320, margin=dict(l=10, r=10, t=50, b=80))
                    qca_fig.update_xaxes(tickangle=-45)

                    qca_table = dash_table.DataTable(
                        columns=[
                            {"name": c, "id": c, "type": "numeric", "format": {"specifier": ",.2f"}}
                            if c == "Penalty (₹)" else {"name": c, "id": c}
                            for c in agg_by_qca.columns
                        ],
                        data=agg_by_qca.to_dict("records"),
                        sort_action="native",
                        style_table={"overflowX": "auto"},
                        style_cell={"padding": "6px", "textAlign": "left"},
                        style_header={"fontWeight": "600", "backgroundColor": "#f8f9fa"},
                    )
                    children_list.append(
                        html.H6(
                            "QCA Analytics",
                            style={"marginTop": "1.25rem", "marginBottom": "0.5rem", "color": "#333", "fontWeight": "600"},
                        )
                    )
                    children_list.append(
                        dbc.Row(
                            [
                                dbc.Col(dcc.Graph(id="agg-qca-penalty-chart", figure=qca_fig, config={"displayModeBar": True}), md=7),
                                dbc.Col(html.Div([html.H6("Summary by QCA", className="mb-2"), qca_table], style={"fontSize": "0.9rem"}), md=5),
                            ],
                            className="mb-3",
                        )
                    )
        except Exception as e:
            print(f"DEBUG - (agg) QCA analytics failed: {e}")

        # -------------------------
        # Data Health (selected filters only)
        # -------------------------
        try:
            hs, extra = summarize_health_from_df(
                df,
                start_date=filters.get("start_date"),
                end_date=filters.get("end_date"),
                plant_col="Plant",
            )
            health_lines = []
            if hs.missing_blocks > 0:
                health_lines.append(
                    f"Missing blocks: {hs.missing_blocks:,} (expected {hs.expected_blocks:,}, present {hs.present_blocks:,})"
                )
            if hs.ppa_missing_rows > 0:
                health_lines.append(f"Missing/zero PPA rows: {hs.ppa_missing_rows:,}")
            if hs.avc_zero_rows > 0:
                health_lines.append(f"AVC = 0 rows: {hs.avc_zero_rows:,}")
            if hs.all_zero_rows > 0:
                health_lines.append(
                    f"All-zero rows (AVC=0 & Schedule=0 & Actual=0): {hs.all_zero_rows:,}"
                )

            if not health_lines:
                health_alert = dbc.Alert(
                    "No data quality issues detected for the selected filters.", color="success"
                )
            else:
                health_alert = dbc.Alert([html.Div(x) for x in health_lines], color="warning")

            children_list.append(
                html.H6(
                    "Data Health (selected filters)",
                    style={
                        "marginTop": "0.5rem",
                        "marginBottom": "0.5rem",
                        "color": "#333",
                        "fontWeight": "600",
                    },
                )
            )
            children_list.append(health_alert)

            per_plant = extra.get("per_plant") if isinstance(extra, dict) else None
            if isinstance(per_plant, pd.DataFrame) and not per_plant.empty:
                drill = per_plant.copy()
                if "present_blocks" in drill.columns:
                    drill = drill.sort_values("present_blocks")
                drill_table = dash_table.DataTable(
                    columns=[{"name": c, "id": c} for c in drill.columns],
                    data=drill.to_dict("records"),
                    sort_action="native",
                    page_size=10,
                    style_table={"overflowX": "auto"},
                    style_cell={"padding": "6px", "textAlign": "left"},
                    style_header={"fontWeight": "600", "backgroundColor": "#f8f9fa"},
                )
                children_list.append(
                    html.Details(
                        [
                            html.Summary("Show per-plant health details", style={"cursor": "pointer", "fontWeight": "600"}),
                            html.Div(drill_table, style={"marginTop": "0.75rem"}),
                        ],
                        open=False,
                        style={"marginBottom": "0.75rem"},
                    )
                )
        except Exception as e:
            print(f"DEBUG - (agg) Health panel failed: {e}")

        children_list.append(
            dbc.Row([dbc.Col(plant_summary_table, md=12)])
        )
        children_list.append(
            html.Div(download_btn, className="text-end mt-2")
        )

        return dcc.Loading(
            type="circle",
            children=html.Div(children_list)
        )
    except Exception as e:
        return dbc.Alert(f"Failed to render aggregated results: {e}", color="danger")

@app.callback(
    Output("custom-upload-preset-select", "options"),
    Input("presets-store", "data"),
)
def load_custom_preset_options(presets):
    presets = presets or []
    return [{"label": p.get("name", "Unnamed"), "value": p.get("name", "Unnamed")} for p in presets]

@app.callback(
    Output("dl-sample-xlsx", "data"),
    Input("btn-download-sample-xlsx", "n_clicks"),
    prevent_initial_call=True
)
def dl_sample_xlsx(n):
    if not n:
        raise PreventUpdate
    import pandas as pd
    from io import BytesIO
    # Build sample dataframe
    def block_to_times(block):
        start = (int(block) - 1) * 15
        h1, m1 = divmod(start, 60)
        end = start + 15
        h2, m2 = divmod(end, 60)
        return f"{h1:02d}:{m1:02d}", f"{h2:02d}:{m2:02d}"
    rows = []
    for blk in [21, 22, 23, 24]:
        ft, tt = block_to_times(blk)
        rows.append({
            "region": "NRPC",
            "plant_name": "Plant-A",
            "date": "2025-01-01",
            "time_block": blk,
            "from_time": ft,
            "to_time": tt,
            "AvC_MW": 50,
            "Scheduled_MW": 45,
            "Actual_MW": 40,
            "PPA": 3.0,
        })
    df = pd.DataFrame(rows)
    out = BytesIO()
    with pd.ExcelWriter(out, engine="xlsxwriter") as xw:
        df.to_excel(xw, sheet_name="Template", index=False)
    return dcc.send_bytes(out.getvalue(), filename="custom_upload_template.xlsx")

@app.callback(
    Output("dl-sample-csv", "data"),
    Input("btn-download-sample-csv", "n_clicks"),
    prevent_initial_call=True
)
def dl_sample_csv(n):
    if not n:
        raise PreventUpdate
    import pandas as pd
    from io import StringIO
    sio = StringIO()
    df = pd.DataFrame([{
        "region": "NRPC",
        "plant_name": "Plant-A",
        "date": "2025-01-01",
        "time_block": 21,
        "from_time": "05:00",
        "to_time": "05:15",
        "AvC_MW": 50,
        "Scheduled_MW": 45,
        "Actual_MW": 40,
        "PPA": 3.0,
    }])
    df.to_csv(sio, index=False)
    return dict(content=sio.getvalue(), filename="custom_upload_template.csv", type="text/csv")

@app.callback(
    Output("custom-upload-store", "data"),
    Output("custom-upload-validate", "children"),
    Input("upload-custom-file", "contents"),
    State("upload-custom-file", "filename"),
    prevent_initial_call=True
)
def handle_upload(contents, filename):
    if not contents:
        raise PreventUpdate
    import base64, io
    header, b64 = contents.split(",", 1)
    data = base64.b64decode(b64)
    try:
        if filename.lower().endswith(".csv"):
            df = pd.read_csv(io.BytesIO(data))
        else:
            df = pd.read_excel(io.BytesIO(data))
    except Exception as e:
        return None, dbc.Alert(f"Unable to read file: {e}", color="danger")

    def canon(s):
        return str(s).strip()
    df.columns = [canon(c) for c in df.columns]

    required = ["region","plant_name","date","time_block","AvC_MW","Scheduled_MW","Actual_MW","PPA"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        return None, dbc.Alert(f"Missing required columns: {', '.join(missing)}", color="warning")

    for opt in ["from_time","to_time"]:
        if opt not in df.columns:
            df[opt] = ""

    for c in ["time_block","AvC_MW","Scheduled_MW","Actual_MW","PPA"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    return df.to_json(orient="records"), dbc.Alert(f"Loaded {len(df)} rows from {filename}", color="success")

@app.callback(
    Output("btn-run-custom", "disabled"),
    Input("custom-upload-store", "data"),
    Input("custom-upload-preset-select", "value"),
)
def toggle_run(data, presets):
    return not (data and (presets or []))

@app.callback(
    Output("custom-results", "children"),
    Output("custom-results-store", "data"),
    Input("btn-run-custom", "n_clicks"),
    State("custom-upload-store", "data"),
    State("custom-upload-preset-select", "value"),
    State("presets-store", "data"),
    prevent_initial_call=True
)
def run_custom(n, data_json, preset_names, presets_store):
    if not n:
        raise PreventUpdate
    if not data_json:
        raise PreventUpdate
    df = pd.DataFrame(json.loads(data_json))
    if df.empty:
        return dbc.Alert("No rows found in upload.", color="warning"), None

    presets_store = presets_store or []
    name_to_settings = {p.get("name"): p.get("settings", {}) for p in presets_store if isinstance(p, dict)}
    selected = list(preset_names or [])
    if not selected:
        return dbc.Alert("Select at least one preset.", color="warning"), None

    combined_ps = []
    first_settings = None
    first_bands = None
    detail_first = None

    for nm in selected:
        st = name_to_settings.get(nm)
        if not st:
            print(f"DEBUG - Preset '{nm}' not found; skipping")
            continue
        err_mode = str(st.get("err_mode", "default")).lower()
        mode_upper = MODE_DEFAULT if err_mode == "default" else MODE_DYNAMIC
        try:
            x_pct = float(st.get("x_pct", 50))
        except Exception:
            x_pct = 50.0
        dyn_x = (x_pct / 100.0) if mode_upper == MODE_DYNAMIC else 0.0
        bands_rows = st.get("bands", DEFAULT_BANDS.copy())
        bands_list, bands_table = parse_bands_from_settings(bands_rows)

        out_rows = []
        for slot in df.to_dict("records"):
            # Ensure numeric coercion
            slot_local = {
                **slot,
                "AvC_MW": float(pd.to_numeric(slot.get("AvC_MW", 0), errors="coerce") or 0),
                "Scheduled_MW": float(pd.to_numeric(slot.get("Scheduled_MW", 0), errors="coerce") or 0),
                "Actual_MW": float(pd.to_numeric(slot.get("Actual_MW", 0), errors="coerce") or 0),
                "PPA": float(pd.to_numeric(slot.get("PPA", 0), errors="coerce") or 0),
            }
            calc = compute_slot_row(slot_local, bands_list, mode_upper, dyn_x)
            out_rows.append({**slot_local, **calc})
        out_df = pd.DataFrame(out_rows)

        plant_summary = (
            out_df.groupby(["region","plant_name"], as_index=False)
            .agg({
                "PPA":"median",
                "AvC_MW":"median",
                "Revenue_Loss":"sum",
                "Revenue_as_per_generation":"sum",
                "Total_DSM":"sum"
            })
            .rename(columns={"plant_name":"Plant name","AvC_MW":"Plant Capacity","Total_DSM":"DSM Loss"})
        )
        # Avoid division by zero
        plant_summary["Revenue Loss (%)"] = plant_summary.apply(
            lambda r: (float(r["Revenue_Loss"]) / float(r["Revenue_as_per_generation"])) * 100.0 if float(r["Revenue_as_per_generation"]) > 0 else 0.0,
            axis=1
        )
        plant_summary["Revenue Loss (p/k)"] = plant_summary["Revenue_Loss"] / 1000.0
        plant_summary["Custom Setting"] = nm
        combined_ps.append(plant_summary)

        if first_settings is None:
            first_settings = {"err_mode": err_mode, "x_pct": x_pct}
            first_bands = bands_rows
            detail_first = out_df

    if not combined_ps:
        return dbc.Alert("No valid preset selected.", color="warning"), None

    combined_df = pd.concat(combined_ps, ignore_index=True) if len(combined_ps) > 1 else combined_ps[0]
    table = dash_table.DataTable(
        columns=[{"name": c, "id": c} for c in combined_df.columns],
        data=combined_df.to_dict("records"),
        sort_action="native",
        style_table={"overflowX":"auto"},
        style_header={"fontWeight":"600"},
    )
    stored = {
        "df": detail_first.to_dict("records") if isinstance(detail_first, pd.DataFrame) else [],
        "final_bands": first_bands or [],
        "err_mode": (first_settings or {}).get("err_mode", "default"),
        "x_pct": (first_settings or {}).get("x_pct", 50.0),
    }
    return table, stored

@app.callback(
    Output("download-custom-output", "data"),
    Input("download-custom-output-btn", "n_clicks"),
    State("custom-results-store", "data"),
    prevent_initial_call=True
)
def download_custom(n, stored):
    if not n:
        raise PreventUpdate
    if not stored:
        raise PreventUpdate
    try:
        df_main = pd.DataFrame(stored.get("df", []))
    except Exception:
        df_main = pd.DataFrame()
    if df_main.empty:
        raise PreventUpdate

    err_mode = str(stored.get("err_mode", "default")).lower()
    mode_upper = MODE_DEFAULT if err_mode == "default" else MODE_DYNAMIC
    try:
        x_pct = float(stored.get("x_pct", 50))
    except Exception:
        x_pct = 50.0
    dyn_x = (x_pct / 100.0) if mode_upper == MODE_DYNAMIC else 0.0

    bands_rows = stored.get("final_bands", [])
    bands_list, bands_table = parse_bands_from_settings(bands_rows)

    base_cols = [
        "region","plant_name","date","time_block","from_time","to_time",
        "AvC_MW","Scheduled_MW","Actual_MW","PPA"
    ]
    missing = [c for c in base_cols if c not in df_main.columns]
    if missing:
        df_exp = df_main.copy()
        if "plant_name" not in df_exp.columns and "Plant" in df_exp.columns:
            df_exp["plant_name"] = df_exp["Plant"]
        if "time_block" not in df_exp.columns and "block" in df_exp.columns:
            df_exp["time_block"] = df_exp["block"]
        missing2 = [c for c in base_cols if c not in df_exp.columns]
        if missing2:
            raise PreventUpdate
        detail_for_excel = df_exp[base_cols].copy()
    else:
        detail_for_excel = df_main[base_cols].copy()

    # Calculate per-slot metrics as in main export
    detail_calculated_rows = []
    for _, row in detail_for_excel.iterrows():
        slot = {
            "region": row["region"],
            "plant_name": row["plant_name"],
            "date": row["date"],
            "time_block": row["time_block"],
            "from_time": row["from_time"],
            "to_time": row["to_time"],
            "AvC_MW": float(row["AvC_MW"]),
            "Scheduled_MW": float(row["Scheduled_MW"]),
            "Actual_MW": float(row["Actual_MW"]),
            "PPA": float(row["PPA"]),
        }
        calc = compute_slot_row(slot, bands_list, mode_upper, dyn_x)
        detail_calculated_rows.append({**slot, **calc})
    detail_calculated_df = pd.DataFrame(detail_calculated_rows)

    # Build workbook (mirrors main export structure)
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils import column_index_from_string
    from openpyxl.workbook.defined_name import DefinedName

    buf = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    # Config
    ws_config = wb.create_sheet("Config")
    ws_config.cell(row=1, column=1, value="Key")
    ws_config.cell(row=1, column=2, value="Value")
    ws_config.cell(row=2, column=1, value="MODE")
    ws_config.cell(row=2, column=2, value=mode_upper)
    ws_config.cell(row=3, column=1, value="DYN_X")
    ws_config.cell(row=3, column=2, value=float(dyn_x))
    try:
        wb.defined_names.append(DefinedName(name="CFG_MODE", attr_text="Config!$B$2"))
        wb.defined_names.append(DefinedName(name="CFG_DYNX", attr_text="Config!$B$3"))
    except Exception:
        pass

    # Bands
    ws_bands = wb.create_sheet("Bands")
    bands_cols = ["direction","lower_pct","upper_pct","rate_type","rate_value","rate_slope","loss_zone"]
    for col_idx, col_name in enumerate(bands_cols, start=1):
        ws_bands.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(bands_table.itertuples(index=False), start=2):
        ws_bands.cell(row=row_idx, column=1, value=str(getattr(row, "direction", "")))
        ws_bands.cell(row=row_idx, column=2, value=float(getattr(row, "lower_pct", 0)))
        ws_bands.cell(row=row_idx, column=3, value=float(getattr(row, "upper_pct", 0)))
        ws_bands.cell(row=row_idx, column=4, value=str(getattr(row, "rate_type", "")))
        ws_bands.cell(row=row_idx, column=5, value=float(getattr(row, "rate_value", 0)))
        ws_bands.cell(row=row_idx, column=6, value=float(getattr(row, "rate_slope", 0)))
        ws_bands.cell(row=row_idx, column=7, value=bool(getattr(row, "loss_zone", False)))
    n_bands = max(len(bands_table), 1)
    bands_end_ref = 200 if n_bands < 199 else (1 + n_bands)
    try:
        wb.defined_names.append(DefinedName(name="Bands_Dir", attr_text=f"Bands!$A$2:$A${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_Lower", attr_text=f"Bands!$B$2:$B${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_Upper", attr_text=f"Bands!$C$2:$C${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_RateType", attr_text=f"Bands!$D$2:$D${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_RateVal", attr_text=f"Bands!$E$2:$E${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_RateSlope", attr_text=f"Bands!$F$2:$F${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_LossZone", attr_text=f"Bands!$G$2:$G${bands_end_ref}"))
    except Exception:
        pass

    # Detail
    ws_detail = wb.create_sheet("Detail")
    detail_headers = {
        'A': 'Region',
        'B': 'Plant Name',
        'C': 'Date',
        'D': 'Block',
        'E': 'From Time',
        'F': 'To Time',
        'G': 'Schedule Power (MW)',
        'H': 'AvC (MW)',
        'I': 'Injected Power (MW)',
        'J': 'PPA or MCP',
        'K': 'Error %',
        'L': 'Absolute error %',
        'M': 'Direction',
        'N': 'Deviation (MW)',
        'U': 'Revenue as per Generation (INR)',
        'V': 'Scheduled Revenue (INR)',
        'AE': '_basis'
    }
    for col_letter, header in detail_headers.items():
        col_idx = column_index_from_string(col_letter)
        ws_detail.cell(row=1, column=col_idx, value=header)
    ws_detail.column_dimensions['AE'].hidden = True

    n_rows = len(detail_for_excel)
    for row_idx, row_data in enumerate(detail_for_excel.itertuples(index=False), start=2):
        calc_row = detail_calculated_df.iloc[row_idx - 2] if row_idx - 2 < len(detail_calculated_df) else {}
        ws_detail.cell(row=row_idx, column=1, value=getattr(row_data, "region", ""))
        ws_detail.cell(row=row_idx, column=2, value=getattr(row_data, "plant_name", ""))
        ws_detail.cell(row=row_idx, column=3, value=getattr(row_data, "date", ""))
        ws_detail.cell(row=row_idx, column=4, value=getattr(row_data, "time_block", ""))
        ws_detail.cell(row=row_idx, column=5, value=getattr(row_data, "from_time", ""))
        ws_detail.cell(row=row_idx, column=6, value=getattr(row_data, "to_time", ""))
        ws_detail.cell(row=row_idx, column=7, value=float(getattr(row_data, "Scheduled_MW", 0)))
        ws_detail.cell(row=row_idx, column=8, value=float(getattr(row_data, "AvC_MW", 0)))
        ws_detail.cell(row=row_idx, column=9, value=float(getattr(row_data, "Actual_MW", 0)))
        ws_detail.cell(row=row_idx, column=10, value=float(getattr(row_data, "PPA", 0)))
        basis = denominator_and_basis(float(getattr(row_data, "AvC_MW", 0)), float(getattr(row_data, "Scheduled_MW", 0)), mode_upper, dyn_x)
        ws_detail.cell(row=row_idx, column=31, value=float(basis))
        ws_detail.cell(row=row_idx, column=11, value=float(calc_row.get("error_pct", 0.0)))
        ws_detail.cell(row=row_idx, column=12, value=float(calc_row.get("abs_err", 0.0)))
        ws_detail.cell(row=row_idx, column=13, value=str(calc_row.get("direction", "")))
        deviation = float(getattr(row_data, "Actual_MW", 0)) - float(getattr(row_data, "Scheduled_MW", 0))
        ws_detail.cell(row=row_idx, column=14, value=float(deviation))
        ws_detail.cell(row=row_idx, column=21, value=float(calc_row.get("Revenue_as_per_generation", 0.0)))
        ws_detail.cell(row=row_idx, column=22, value=float(calc_row.get("Scheduled_Revenue_as_per_generation", 0.0)))

    # Totals
    ws_detail.cell(row=1, column=29, value='Total DSM (INR)')
    ws_detail.cell(row=1, column=30, value='Revenue Loss (INR)')
    for data_row in range(2, len(detail_calculated_df) + 2):
        calc_row = detail_calculated_df.iloc[data_row - 2]
        ws_detail.cell(row=data_row, column=29, value=float(calc_row.get("Total_DSM", 0.0)))
        ws_detail.cell(row=data_row, column=30, value=float(calc_row.get("Revenue_Loss", 0.0)))

    wb.save(buf)
    buf.seek(0)
    return dcc.send_bytes(lambda x: x.write(buf.read()), "custom_analysis.xlsx")

@app.callback(
    Output("download-excel", "data"),
    Input("btn-download", "n_clicks"),
    State("results-store", "data"),
    prevent_initial_call=True
)
def download_full(n, stored):
    """Export Analysis results to Excel. Uses same bands/err_mode as pipeline; STU and RPC treated identically."""
    if not n or n == 0:
        raise PreventUpdate
    if not stored or stored.get("error"):
        raise PreventUpdate

    # Multi-preset export path
    all_runs = stored.get("_all_runs") if isinstance(stored, dict) else None
    if all_runs:
        from io import BytesIO
        output = BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as xw:
            # Combined Plant Summary
            try:
                plant_summary_df = pd.DataFrame(stored.get("plant_summary", []))
                if not plant_summary_df.empty:
                    plant_summary_df.to_excel(xw, sheet_name="Plant_Summary", index=False)
            except Exception as e:
                print(f"DEBUG - Failed writing Plant_Summary: {e}")

            # One Detail/Config/Bands per preset
            for r in all_runs:
                nm = str(r.get("name", "Preset"))
                try:
                    df_detail = pd.DataFrame(json.loads(r.get("df", "[]")))
                except Exception:
                    df_detail = pd.DataFrame(r.get("df", []))
                bands_df = _normalize_bands_df(pd.DataFrame(r.get("final_bands", [])))
                # Detail
                try:
                    df_detail.to_excel(xw, sheet_name=f"Detail_{nm}", index=False)
                except Exception as e:
                    print(f"DEBUG - Writing Detail_{nm} failed: {e}")
                # Config
                try:
                    pd.DataFrame({"Key":["MODE","DYN_X"], "Value":[r.get("err_mode"), r.get("x_pct")]}) \
                        .to_excel(xw, sheet_name=f"Config_{nm}", index=False)
                except Exception as e:
                    print(f"DEBUG - Writing Config_{nm} failed: {e}")
                # Bands
                try:
                    bands_df.to_excel(xw, sheet_name=f"Bands_{nm}", index=False)
                except Exception as e:
                    print(f"DEBUG - Writing Bands_{nm} failed: {e}")
        return dcc.send_bytes(output.getvalue(), filename="DSM_Full_Calculation_MultiPresets.xlsx")

    df_main = pd.DataFrame(json.loads(stored["df"])) if isinstance(stored.get("df"), str) else pd.DataFrame(stored.get("df", []))
    if df_main.empty:
        raise PreventUpdate

    used = stored.get("used_settings", {}) if isinstance(stored, dict) else {}
    err_mode = str(used.get("err_mode", "default")).lower()
    mode_upper = MODE_DEFAULT if err_mode == "default" else MODE_DYNAMIC
    try:
        x_pct = float(used.get("x_pct", 50))
    except Exception:
        x_pct = 50.0
    dyn_x = (x_pct / 100.0) if mode_upper == MODE_DYNAMIC else 0.0

    bands_rows = stored.get("final_bands", []) if isinstance(stored, dict) else []
    bands_list, bands_table = parse_bands_from_settings(bands_rows)

    # Select input columns for formula-driven export
    base_cols = [
        "region","plant_name","date","time_block","from_time","to_time",
        "AvC_MW","Scheduled_MW","Actual_MW","PPA"
    ]
    missing = [c for c in base_cols if c not in df_main.columns]
    if missing:
        # Try to map Plant->plant_name and block->time_block if present
        df_exp = df_main.copy()
        if "plant_name" not in df_exp.columns and "Plant" in df_exp.columns:
            df_exp["plant_name"] = df_exp["Plant"]
        if "time_block" not in df_exp.columns and "block" in df_exp.columns:
            df_exp["time_block"] = df_exp["block"]
        missing2 = [c for c in base_cols if c not in df_exp.columns]
        if missing2:
            raise PreventUpdate
        detail_for_excel = df_exp[base_cols].copy()
    else:
        detail_for_excel = df_main[base_cols].copy()

    # Pre-calculate all slot values and build complete detail with per-band columns
    detail_calculated_rows = []
    for _, row in detail_for_excel.iterrows():
        slot = {
            "region": row["region"],
            "plant_name": row["plant_name"],
            "date": row["date"],
            "time_block": row["time_block"],
            "from_time": row["from_time"],
            "to_time": row["to_time"],
            "AvC_MW": float(row["AvC_MW"]),
            "Scheduled_MW": float(row["Scheduled_MW"]),
            "Actual_MW": float(row["Actual_MW"]),
            "PPA": float(row["PPA"]),
        }
        calc = compute_slot_row(slot, bands_list, mode_upper, dyn_x)
        # Merge input + calculated
        detail_calculated_rows.append({**slot, **calc})
    detail_calculated_df = pd.DataFrame(detail_calculated_rows)

    # Build per-band columns: one column per band showing deviation energy (MW) AND DSM amount (INR)
    per_band_cols = {}
    per_band_dsm_cols = {}
    for band in bands_list:
        dir_label = "UI" if band.direction == "UI" else "OI"
        if band.upper_pct >= 999:
            band_label = f"{dir_label} Energy deviation >{int(band.lower_pct)}%"
            dsm_label = f"{dir_label} DSM due to Deviation >{int(band.lower_pct)}% (INR)"
        else:
            band_label = f"{dir_label} Energy deviation between {int(band.lower_pct)}-{int(band.upper_pct)}%"
            dsm_label = f"{dir_label} DSM between {int(band.lower_pct)}-{int(band.upper_pct)}% (INR)"
        per_band_cols[band_label] = []
        per_band_dsm_cols[dsm_label] = []

    # Calculate per-band deviation and DSM for each row
    for _, row in detail_calculated_df.iterrows():
        abs_err = row.get("abs_err", 0.0)
        direction = row.get("direction", "")
        avc = row.get("AvC_MW", 0.0)
        sch = row.get("Scheduled_MW", 0.0)
        ppa = row.get("PPA", 0.0)
        denom = denominator_and_basis(avc, sch, mode_upper, dyn_x)

        for band_idx, band in enumerate(bands_list):
            energy_col_name = list(per_band_cols.keys())[band_idx]
            dsm_col_name = list(per_band_dsm_cols.keys())[band_idx]
            if band.direction != direction:
                per_band_cols[energy_col_name].append(0.0)
                per_band_dsm_cols[dsm_col_name].append(0.0)
                continue
            sp = slice_pct(abs_err, band.lower_pct, band.upper_pct)
            if sp > 0:
                kwh = kwh_from_slice(sp, denom)
                rate = band_rate(ppa, band.rate_type, band.rate_value, band.rate_slope, abs_err)
                amount = kwh * rate
                per_band_cols[energy_col_name].append(kwh / 1000.0)  # Convert to MW for display
                per_band_dsm_cols[dsm_col_name].append(amount)
            else:
                per_band_cols[energy_col_name].append(0.0)
                per_band_dsm_cols[dsm_col_name].append(0.0)

    # Add per-band columns to detail dataframe
    for col_name, values in per_band_cols.items():
        detail_calculated_df[col_name] = values
    for col_name, values in per_band_dsm_cols.items():
        detail_calculated_df[col_name] = values

    # Reorder columns: base columns, then calculated standard columns, then per-band columns (energy + DSM)
    # Removed generic UI_Energy_deviation_bands and OI_Energy_deviation_bands - replaced with per-band columns
    standard_calc_cols = [
        "error_pct", "direction", "abs_err", "band_level",
        "Revenue_as_per_generation", "Scheduled_Revenue_as_per_generation",
        "UI_DSM", "OI_DSM", "OI_Loss", "Total_DSM", "Revenue_Loss"
    ]
    # Interleave per-band columns: energy then DSM for each band
    # Sort by direction and lower_pct for proper ordering
    def get_band_sort_key(col_name):
        # Extract direction and range for sorting
        if "UI" in col_name:
            dir_val = 0
        elif "OI" in col_name:
            dir_val = 1
        else:
            dir_val = 2
        # Extract lower percentage
        import re
        nums = re.findall(r'\d+', col_name)
        lower_val = int(nums[0]) if nums else 9999
        return (dir_val, lower_val)
    
    per_band_col_names = sorted(list(per_band_cols.keys()), key=get_band_sort_key)
    per_band_dsm_col_names = sorted(list(per_band_dsm_cols.keys()), key=get_band_sort_key)
    # Group by direction and range to interleave properly
    interleaved_per_band = []
    for energy_col in per_band_col_names:
        interleaved_per_band.append(energy_col)
        # Find corresponding DSM column by matching the range
        # Extract the range portion (e.g., "0-10%" or ">20%")
        import re
        energy_range = re.search(r'(\d+-?\d*%|>\d+%)', energy_col)
        if energy_range:
            range_str = energy_range.group(1)
            for dsm_col in per_band_dsm_col_names:
                if range_str in dsm_col and energy_col.split()[0] in dsm_col:  # Same direction and range
                    interleaved_per_band.append(dsm_col)
                    break
    final_cols = base_cols + [c for c in standard_calc_cols if c in detail_calculated_df.columns] + interleaved_per_band
    detail_calculated_df = detail_calculated_df[[c for c in final_cols if c in detail_calculated_df.columns]]

    # Export with Office 365 dynamic array formulas
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    
    buf = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # ===== CONFIG SHEET =====
    ws_config = wb.create_sheet("Config")
    ws_config.cell(row=1, column=1, value="Key")
    ws_config.cell(row=1, column=2, value="Value")
    ws_config.cell(row=2, column=1, value="MODE")
    ws_config.cell(row=2, column=2, value=mode_upper)
    ws_config.cell(row=3, column=1, value="DYN_X")
    ws_config.cell(row=3, column=2, value=float(dyn_x))
    
    # Named cells for Config
    from openpyxl.workbook.defined_name import DefinedName
    try:
        wb.defined_names.append(DefinedName(name="CFG_MODE", attr_text="Config!$B$2"))
        wb.defined_names.append(DefinedName(name="CFG_DYNX", attr_text="Config!$B$3"))
    except Exception:
        pass
    
    # ===== BANDS SHEET =====
    ws_bands = wb.create_sheet("Bands")
    bands_cols = ["direction","lower_pct","upper_pct","rate_type","rate_value","rate_slope","loss_zone"]
    for col_idx, col_name in enumerate(bands_cols, start=1):
        ws_bands.cell(row=1, column=col_idx, value=col_name)
    
    for row_idx, row in enumerate(bands_table.itertuples(index=False), start=2):
        ws_bands.cell(row=row_idx, column=1, value=str(getattr(row, "direction", "")))
        ws_bands.cell(row=row_idx, column=2, value=float(getattr(row, "lower_pct", 0)))
        ws_bands.cell(row=row_idx, column=3, value=float(getattr(row, "upper_pct", 0)))
        ws_bands.cell(row=row_idx, column=4, value=str(getattr(row, "rate_type", "")))
        ws_bands.cell(row=row_idx, column=5, value=float(getattr(row, "rate_value", 0)))
        ws_bands.cell(row=row_idx, column=6, value=float(getattr(row, "rate_slope", 0)))
        ws_bands.cell(row=row_idx, column=7, value=bool(getattr(row, "loss_zone", False)))
    
    n_bands = max(len(bands_table), 1)
    bands_end_row = 1 + n_bands
    # Named ranges with headroom (2:200)
    bands_end_ref = 200 if n_bands < 199 else bands_end_row
    
    try:
        wb.defined_names.append(DefinedName(name="Bands_Dir", attr_text=f"Bands!$A$2:$A${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_Lower", attr_text=f"Bands!$B$2:$B${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_Upper", attr_text=f"Bands!$C$2:$C${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_RateType", attr_text=f"Bands!$D$2:$D${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_RateVal", attr_text=f"Bands!$E$2:$E${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_RateSlope", attr_text=f"Bands!$F$2:$F${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_LossZone", attr_text=f"Bands!$G$2:$G${bands_end_ref}"))
    except Exception:
        pass
    
    # ===== DETAIL SHEET =====
    ws_detail = wb.create_sheet("Detail")
    
    # Column mapping per spec
    # A=Region, B=Plant Name, C=Date, D=Block, E=From Time, F=To Time,
    # G=Schedule Power (MW), H=AvC (MW), I=Injected Power (MW), J=PPA or MCP,
    # K=Error %, L=Absolute error %, M=Direction, N=Deviation (MW),
    # U=Revenue as per Generation, V=Scheduled Revenue, AE=_basis
    detail_headers = {
        'A': 'Region',
        'B': 'Plant Name',
        'C': 'Date',
        'D': 'Block',
        'E': 'From Time',
        'F': 'To Time',
        'G': 'Schedule Power (MW)',
        'H': 'AvC (MW)',
        'I': 'Injected Power (MW)',
        'J': 'PPA or MCP',
        'K': 'Error %',
        'L': 'Absolute error %',
        'M': 'Direction',
        'N': 'Deviation (MW)',
        'U': 'Revenue as per Generation (INR)',
        'V': 'Scheduled Revenue (INR)',
        'AE': '_basis'
    }
    
    # Write headers - use column_index_from_string to handle multi-letter columns
    for col_letter, header in detail_headers.items():
        col_idx = column_index_from_string(col_letter)
        ws_detail.cell(row=1, column=col_idx, value=header)
    
    # Hide _basis column
    ws_detail.column_dimensions['AE'].hidden = True
    
    # Write data rows with calculated values (no formulas)
    n_rows = len(detail_for_excel)
    for row_idx, row_data in enumerate(detail_for_excel.itertuples(index=False), start=2):
        calc_row = detail_calculated_df.iloc[row_idx - 2] if row_idx - 2 < len(detail_calculated_df) else {}
        
        # Input values
        ws_detail.cell(row=row_idx, column=1, value=getattr(row_data, "region", ""))  # A
        ws_detail.cell(row=row_idx, column=2, value=getattr(row_data, "plant_name", ""))  # B
        ws_detail.cell(row=row_idx, column=3, value=getattr(row_data, "date", ""))  # C
        ws_detail.cell(row=row_idx, column=4, value=getattr(row_data, "time_block", ""))  # D
        ws_detail.cell(row=row_idx, column=5, value=getattr(row_data, "from_time", ""))  # E
        ws_detail.cell(row=row_idx, column=6, value=getattr(row_data, "to_time", ""))  # F
        ws_detail.cell(row=row_idx, column=7, value=float(getattr(row_data, "Scheduled_MW", 0)))  # G
        ws_detail.cell(row=row_idx, column=8, value=float(getattr(row_data, "AvC_MW", 0)))  # H
        ws_detail.cell(row=row_idx, column=9, value=float(getattr(row_data, "Actual_MW", 0)))  # I
        ws_detail.cell(row=row_idx, column=10, value=float(getattr(row_data, "PPA", 0)))  # J
        
        # Calculated values (not formulas)
        avc = float(getattr(row_data, "AvC_MW", 0))
        sch = float(getattr(row_data, "Scheduled_MW", 0))
        act = float(getattr(row_data, "Actual_MW", 0))
        ppa = float(getattr(row_data, "PPA", 0))
        
        # _basis (AE) - calculated value
        basis = denominator_and_basis(avc, sch, mode_upper, dyn_x)
        ws_detail.cell(row=row_idx, column=31, value=float(basis))
        
        # Error % (K)
        error_pct = calc_row.get("error_pct", 0.0)
        ws_detail.cell(row=row_idx, column=11, value=float(error_pct))
        
        # Absolute error % (L)
        abs_err = calc_row.get("abs_err", 0.0)
        ws_detail.cell(row=row_idx, column=12, value=float(abs_err))
        
        # Direction (M)
        direction = calc_row.get("direction", "")
        ws_detail.cell(row=row_idx, column=13, value=str(direction))
        
        # Deviation (MW) (N)
        deviation = act - sch
        ws_detail.cell(row=row_idx, column=14, value=float(deviation))
        
        # Revenue as per Generation (U)
        rev_act = calc_row.get("Revenue_as_per_generation", act * 0.25 * 1000 * ppa)
        ws_detail.cell(row=row_idx, column=21, value=float(rev_act))
        
        # Scheduled Revenue (V)
        rev_sch = calc_row.get("Scheduled_Revenue_as_per_generation", sch * 0.25 * 1000 * ppa)
        ws_detail.cell(row=row_idx, column=22, value=float(rev_sch))
    
    # ===== BAND-WISE ENERGY DEVIATION COLUMNS (O-T) =====
    # Use user-defined bands directly with proper cumulative slicing
    ui_bands = sorted([b for b in bands_list if b.direction == "UI"], key=lambda x: x.lower_pct)
    oi_bands = sorted([b for b in bands_list if b.direction == "OI"], key=lambda x: x.lower_pct)
    
    # Combine bands: UI first, then OI, limit to 6 columns (O-T = columns 15-20)
    # Show all UI bands first, then OI bands
    all_bands_for_energy = (ui_bands + oi_bands)[:6]
    
    # Write band-wise energy deviation headers and values (columns O-T)
    for col_idx, band in enumerate(all_bands_for_energy, start=15):  # O=15, P=16, ..., T=20
        dir_label = "UI" if band.direction == "UI" else "OI"
        
        if band.upper_pct >= 999:
            header = f"{dir_label} Energy >{int(band.lower_pct)}% (MW)"
        else:
            header = f"{dir_label} Energy {int(band.lower_pct)}-{int(band.upper_pct)}% (MW)"
        ws_detail.cell(row=1, column=col_idx, value=header)
        
        # Write calculated energy values for each data row with proper cumulative slicing
        # Formula: SIGN(deviation) * MAX(0, MIN(abs_err, upper) - lower) / 100 * basis
        for data_row in range(2, n_rows + 2):
            calc_row = detail_calculated_df.iloc[data_row - 2] if data_row - 2 < len(detail_calculated_df) else {}
            abs_err = calc_row.get("abs_err", 0.0)
            direction = calc_row.get("direction", "")
            avc = detail_for_excel.iloc[data_row - 2]["AvC_MW"]
            sch = detail_for_excel.iloc[data_row - 2]["Scheduled_MW"]
            act = detail_for_excel.iloc[data_row - 2]["Actual_MW"]
            basis = denominator_and_basis(avc, sch, mode_upper, dyn_x)
            deviation = act - sch  # N2 = I2 - G2 (can be negative for UI, positive for OI)
            
            # Calculate energy deviation for this specific band with cumulative slicing
            # Formula: SIGN(deviation) * MAX(0, MIN(abs_err, upper) - lower) / 100 * basis
            # This preserves sign: negative for UI, positive for OI
            energy_mw = 0.0
            if direction == dir_label:
                # Calculate slice percentage
                sp = slice_pct(abs_err, band.lower_pct, band.upper_pct)
                if sp > 0:
                    # SIGN(deviation) preserves the correct sign
                    sign = -1.0 if deviation < 0 else (1.0 if deviation > 0 else 0.0)
                    # Convert slice percentage to MW: SIGN * slice% / 100 * basis
                    energy_mw = sign * sp / 100.0 * basis
            
            ws_detail.cell(row=data_row, column=col_idx, value=float(energy_mw))
    
    # ===== PER-BAND DSM COLUMNS (after revenue columns) =====
    # Use actual user-defined bands for DSM columns too (for consistency)
    # Start after column V (Revenue columns)
    current_col = 23  # Column W
    
    # UI DSM columns - one column per UI band
    ui_dsm_start_col = current_col
    for band in ui_bands:
        if band.upper_pct >= 999:
            header = f"UI >{int(band.lower_pct)}% DSM"
        else:
            header = f"UI {int(band.lower_pct)}-{int(band.upper_pct)}% DSM"
        ws_detail.cell(row=1, column=current_col, value=header)
        
        # Write calculated values for each data row
        for data_row in range(2, n_rows + 2):
            calc_row = detail_calculated_df.iloc[data_row - 2] if data_row - 2 < len(detail_calculated_df) else {}
            abs_err = calc_row.get("abs_err", 0.0)
            direction = calc_row.get("direction", "")
            avc = detail_for_excel.iloc[data_row - 2]["AvC_MW"]
            sch = detail_for_excel.iloc[data_row - 2]["Scheduled_MW"]
            ppa = detail_for_excel.iloc[data_row - 2]["PPA"]
            basis = denominator_and_basis(avc, sch, mode_upper, dyn_x)
            
            # Calculate DSM for this specific band with cumulative slicing
            dsm_value = 0.0
            if direction == "UI":
                sp = slice_pct(abs_err, band.lower_pct, band.upper_pct)
                if sp > 0:
                    kwh = kwh_from_slice(sp, basis)
                    rate = band_rate(ppa, band.rate_type, band.rate_value, band.rate_slope, abs_err)
                    dsm_value = kwh * rate
            
            ws_detail.cell(row=data_row, column=current_col, value=float(dsm_value))
        
        current_col += 1
    ui_dsm_end_col = current_col - 1
    
    # OI DSM columns (LossZone=FALSE) - one column per OI band
    oi_dsm_start_col = current_col
    for band in oi_bands:
        if not band.loss_zone:  # Only create columns for non-loss-zone OI bands
            if band.upper_pct >= 999:
                header = f"OI >{int(band.lower_pct)}% DSM"
            else:
                header = f"OI {int(band.lower_pct)}-{int(band.upper_pct)}% DSM"
            ws_detail.cell(row=1, column=current_col, value=header)
            
            for data_row in range(2, n_rows + 2):
                calc_row = detail_calculated_df.iloc[data_row - 2] if data_row - 2 < len(detail_calculated_df) else {}
                abs_err = calc_row.get("abs_err", 0.0)
                direction = calc_row.get("direction", "")
                avc = detail_for_excel.iloc[data_row - 2]["AvC_MW"]
                sch = detail_for_excel.iloc[data_row - 2]["Scheduled_MW"]
                ppa = detail_for_excel.iloc[data_row - 2]["PPA"]
                basis = denominator_and_basis(avc, sch, mode_upper, dyn_x)
                
                dsm_value = 0.0
                if direction == "OI":
                    sp = slice_pct(abs_err, band.lower_pct, band.upper_pct)
                    if sp > 0:
                        kwh = kwh_from_slice(sp, basis)
                        rate = band_rate(ppa, band.rate_type, band.rate_value, band.rate_slope, abs_err)
                        dsm_value = kwh * rate
                
                ws_detail.cell(row=data_row, column=current_col, value=float(dsm_value))
            
            current_col += 1
    oi_dsm_end_col = current_col - 1
    
    # OI Loss columns (LossZone=TRUE) - one column per loss-zone OI band
    oi_loss_start_col = current_col
    for band in oi_bands:
        if band.loss_zone:  # Only create columns for loss-zone OI bands
            if band.upper_pct >= 999:
                header = f"OI >{int(band.lower_pct)}% Loss"
            else:
                header = f"OI {int(band.lower_pct)}-{int(band.upper_pct)}% Loss"
            ws_detail.cell(row=1, column=current_col, value=header)
            
            for data_row in range(2, n_rows + 2):
                calc_row = detail_calculated_df.iloc[data_row - 2] if data_row - 2 < len(detail_calculated_df) else {}
                abs_err = calc_row.get("abs_err", 0.0)
                direction = calc_row.get("direction", "")
                avc = detail_for_excel.iloc[data_row - 2]["AvC_MW"]
                sch = detail_for_excel.iloc[data_row - 2]["Scheduled_MW"]
                ppa = detail_for_excel.iloc[data_row - 2]["PPA"]
                basis = denominator_and_basis(avc, sch, mode_upper, dyn_x)
                
                loss_value = 0.0
                if direction == "OI":
                    sp = slice_pct(abs_err, band.lower_pct, band.upper_pct)
                    if sp > 0:
                        kwh = kwh_from_slice(sp, basis)
                        rate = band_rate(ppa, band.rate_type, band.rate_value, band.rate_slope, abs_err)
                        loss_value = kwh * rate
                
                ws_detail.cell(row=data_row, column=current_col, value=float(loss_value))
            
            current_col += 1
    oi_loss_end_col = current_col - 1
    
    # ===== TOTAL COLUMNS (AC and AD) =====
    # AC = column 29, AD = column 30
    ws_detail.cell(row=1, column=29, value='Total DSM (INR)')
    ws_detail.cell(row=1, column=30, value='Revenue Loss (INR)')
    
    for data_row in range(2, n_rows + 2):
        calc_row = detail_calculated_df.iloc[data_row - 2] if data_row - 2 < len(detail_calculated_df) else {}
        
        # Total DSM = calculated value
        total_dsm = calc_row.get("Total_DSM", 0.0)
        ws_detail.cell(row=data_row, column=29, value=float(total_dsm))
        
        # Revenue Loss = calculated value
        revenue_loss = calc_row.get("Revenue_Loss", 0.0)
        ws_detail.cell(row=data_row, column=30, value=float(revenue_loss))
    
    # ===== SUMMARY AREA (optional, at bottom of Detail sheet) =====
    summary_start = n_rows + 5
    ws_detail.cell(row=summary_start, column=1, value='Summary')
    
    # Calculate summary values
    if len(detail_calculated_df) > 0:
        # Get summary from plant_summary if available, otherwise calculate
        region_val = detail_calculated_df["region"].mode().iloc[0] if "region" in detail_calculated_df.columns else ""
        plant_val = detail_calculated_df["plant_name"].mode().iloc[0] if "plant_name" in detail_calculated_df.columns else ""
        avc_mode = safe_mode(detail_for_excel["AvC_MW"].tolist())
        ppa_mode = safe_mode(detail_for_excel["PPA"].tolist())
        rev_loss_sum = detail_calculated_df["Revenue_Loss"].sum() if "Revenue_Loss" in detail_calculated_df.columns else 0.0
        rev_act_sum = detail_calculated_df["Revenue_as_per_generation"].sum() if "Revenue_as_per_generation" in detail_calculated_df.columns else 0.0
        rev_loss_pct = (rev_loss_sum / rev_act_sum * 100.0) if rev_act_sum > 0 else 0.0
        dsm_loss_sum = detail_calculated_df["Total_DSM"].sum() if "Total_DSM" in detail_calculated_df.columns else 0.0
        
        ws_detail.cell(row=summary_start + 1, column=1, value='Region')
        ws_detail.cell(row=summary_start + 1, column=2, value=str(region_val))
        ws_detail.cell(row=summary_start + 2, column=1, value='Plant Name')
        ws_detail.cell(row=summary_start + 2, column=2, value=str(plant_val))
        ws_detail.cell(row=summary_start + 3, column=1, value='Plant Capacity (MODE AvC)')
        ws_detail.cell(row=summary_start + 3, column=2, value=round(float(avc_mode), 2))
        ws_detail.cell(row=summary_start + 4, column=1, value='PPA (MODE)')
        ws_detail.cell(row=summary_start + 4, column=2, value=round(float(ppa_mode), 2))
        ws_detail.cell(row=summary_start + 5, column=1, value='Revenue Loss %')
        ws_detail.cell(row=summary_start + 5, column=2, value=round(float(rev_loss_pct), 2))
        ws_detail.cell(row=summary_start + 6, column=1, value='DSM Loss')
        ws_detail.cell(row=summary_start + 6, column=2, value=round(float(dsm_loss_sum), 2))
    
    wb.save(buf)
    buf.seek(0)
    return dcc.send_bytes(lambda x: x.write(buf.read()), "DSM_calculation.xlsx")


@app.callback(
    Output("download-excel", "data", allow_duplicate=True),
    Input("agg-btn-download", "n_clicks"),
    State("agg-results-store", "data"),
    prevent_initial_call=True
)
def download_full_aggregated(n, stored):
    """Download full calculation for Aggregation Analysis, reusing the same
    Excel structure as the main Analysis export."""
    if not n or n == 0:
        raise PreventUpdate
    if not stored or stored.get("error"):
        raise PreventUpdate

    # Multi-preset aggregated export (same structure as main)
    all_runs = stored.get("_all_runs") if isinstance(stored, dict) else None
    if all_runs:
        from io import BytesIO
        output = BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as xw:
            # Combined Plant Summary
            try:
                plant_summary_df = pd.DataFrame(stored.get("plant_summary", []))
                if not plant_summary_df.empty:
                    plant_summary_df.to_excel(xw, sheet_name="Plant_Summary", index=False)
            except Exception as e:
                print(f"DEBUG - (agg) Failed writing Plant_Summary: {e}")

            # One Detail/Config/Bands per preset
            for r in all_runs:
                nm = str(r.get("name", "Preset"))
                try:
                    df_detail = pd.DataFrame(json.loads(r.get("df", "[]")))
                except Exception:
                    df_detail = pd.DataFrame(r.get("df", []))
                bands_df = _normalize_bands_df(pd.DataFrame(r.get("final_bands", [])))
                # Detail
                try:
                    df_detail.to_excel(xw, sheet_name=f"Detail_{nm}", index=False)
                except Exception as e:
                    print(f"DEBUG - (agg) Writing Detail_{nm} failed: {e}")
                # Config
                try:
                    pd.DataFrame({"Key":["MODE","DYN_X"], "Value":[r.get("err_mode"), r.get("x_pct")]}) \
                        .to_excel(xw, sheet_name=f"Config_{nm}", index=False)
                except Exception as e:
                    print(f"DEBUG - (agg) Writing Config_{nm} failed: {e}")
                # Bands
                try:
                    bands_df.to_excel(xw, sheet_name=f"Bands_{nm}", index=False)
                except Exception as e:
                    print(f"DEBUG - (agg) Writing Bands_{nm} failed: {e}")
        return dcc.send_bytes(output.getvalue(), filename="DSM_Full_Calculation_Aggregated_MultiPresets.xlsx")

    # Single-setting path – identical to main download_full, but reading from agg-results-store
    df_main = pd.DataFrame(json.loads(stored["df"])) if isinstance(stored.get("df"), str) else pd.DataFrame(stored.get("df", []))
    if df_main.empty:
        raise PreventUpdate

    used = stored.get("used_settings", {}) if isinstance(stored, dict) else {}
    err_mode = str(used.get("err_mode", "default")).lower()
    mode_upper = MODE_DEFAULT if err_mode == "default" else MODE_DYNAMIC
    try:
        x_pct = float(used.get("x_pct", 50))
    except Exception:
        x_pct = 50.0
    dyn_x = (x_pct / 100.0) if mode_upper == MODE_DYNAMIC else 0.0

    bands_rows = stored.get("final_bands", []) if isinstance(stored, dict) else []
    bands_list, bands_table = parse_bands_from_settings(bands_rows)

    # Select input columns for export
    base_cols = [
        "region","plant_name","date","time_block","from_time","to_time",
        "AvC_MW","Scheduled_MW","Actual_MW","PPA"
    ]
    missing = [c for c in base_cols if c not in df_main.columns]
    if missing:
        df_exp = df_main.copy()
        if "plant_name" not in df_exp.columns and "Plant" in df_exp.columns:
            df_exp["plant_name"] = df_exp["Plant"]
        if "time_block" not in df_exp.columns and "block" in df_exp.columns:
            df_exp["time_block"] = df_exp["block"]
        missing2 = [c for c in base_cols if c not in df_exp.columns]
        if missing2:
            raise PreventUpdate
        detail_for_excel = df_exp[base_cols].copy()
    else:
        detail_for_excel = df_main[base_cols].copy()

    # Reuse same detail_calculated_df logic
    detail_calculated_rows = []
    for _, row in detail_for_excel.iterrows():
        slot = {
            "region": row["region"],
            "plant_name": row["plant_name"],
            "date": row["date"],
            "time_block": row["time_block"],
            "from_time": row["from_time"],
            "to_time": row["to_time"],
            "AvC_MW": float(row["AvC_MW"]),
            "Scheduled_MW": float(row["Scheduled_MW"]),
            "Actual_MW": float(row["Actual_MW"]),
            "PPA": float(row["PPA"]),
        }
        calc = compute_slot_row(slot, bands_list, mode_upper, dyn_x)
        detail_calculated_rows.append({**slot, **calc})
    detail_calculated_df = pd.DataFrame(detail_calculated_rows)

    # The remainder of the export logic (per-band columns, headers, summary, etc.)
    # is identical to the main download_full implementation and is reused here
    # by following the same steps.

    # Build per-band columns
    per_band_cols = {}
    per_band_dsm_cols = {}
    for band in bands_list:
        dir_label = "UI" if band.direction == "UI" else "OI"
        if band.upper_pct >= 999:
            band_label = f"{dir_label} Energy deviation >{int(band.lower_pct)}%"
            dsm_label = f"{dir_label} DSM due to Deviation >{int(band.lower_pct)}% (INR)"
        else:
            band_label = f"{dir_label} Energy deviation between {int(band.lower_pct)}-{int(band.upper_pct)}%"
            dsm_label = f"{dir_label} DSM between {int(band.lower_pct)}-{int(band.upper_pct)}% (INR)"
        per_band_cols[band_label] = []
        per_band_dsm_cols[dsm_label] = []

    for _, row in detail_calculated_df.iterrows():
        abs_err = row.get("abs_err", 0.0)
        direction = row.get("direction", "")
        avc = row.get("AvC_MW", 0.0)
        sch = row.get("Scheduled_MW", 0.0)
        ppa = row.get("PPA", 0.0)
        denom = denominator_and_basis(avc, sch, mode_upper, dyn_x)

        for band_idx, band in enumerate(bands_list):
            energy_col_name = list(per_band_cols.keys())[band_idx]
            dsm_col_name = list(per_band_dsm_cols.keys())[band_idx]
            if band.direction != direction:
                per_band_cols[energy_col_name].append(0.0)
                per_band_dsm_cols[dsm_col_name].append(0.0)
                continue
            sp = slice_pct(abs_err, band.lower_pct, band.upper_pct)
            if sp > 0:
                kwh = kwh_from_slice(sp, denom)
                rate = band_rate(ppa, band.rate_type, band.rate_value, band.rate_slope, abs_err)
                amount = kwh * rate
                per_band_cols[energy_col_name].append(kwh / 1000.0)
                per_band_dsm_cols[dsm_col_name].append(amount)
            else:
                per_band_cols[energy_col_name].append(0.0)
                per_band_dsm_cols[dsm_col_name].append(0.0)

    for col_name, values in per_band_cols.items():
        detail_calculated_df[col_name] = values
    for col_name, values in per_band_dsm_cols.items():
        detail_calculated_df[col_name] = values

    standard_calc_cols = [
        "error_pct", "direction", "abs_err", "band_level",
        "Revenue_as_per_generation", "Scheduled_Revenue_as_per_generation",
        "UI_DSM", "OI_DSM", "OI_Loss", "Total_DSM", "Revenue_Loss"
    ]

    def get_band_sort_key(col_name):
        if "UI" in col_name:
            dir_val = 0
        elif "OI" in col_name:
            dir_val = 1
        else:
            dir_val = 2
        import re
        nums = re.findall(r'\d+', col_name)
        lower_val = int(nums[0]) if nums else 9999
        return (dir_val, lower_val)

    per_band_col_names = sorted(list(per_band_cols.keys()), key=get_band_sort_key)
    per_band_dsm_col_names = sorted(list(per_band_dsm_cols.keys()), key=get_band_sort_key)
    interleaved_per_band = []
    for energy_col in per_band_col_names:
        interleaved_per_band.append(energy_col)
        import re
        energy_range = re.search(r'(\d+-?\d*%|>\d+%)', energy_col)
        if energy_range:
            range_str = energy_range.group(1)
            for dsm_col in per_band_dsm_col_names:
                if range_str in dsm_col and energy_col.split()[0] in dsm_col:
                    interleaved_per_band.append(dsm_col)
                    break
    final_cols = base_cols + [c for c in standard_calc_cols if c in detail_calculated_df.columns] + interleaved_per_band
    detail_calculated_df = detail_calculated_df[[c for c in final_cols if c in detail_calculated_df.columns]]

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter, column_index_from_string

    buf = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    # Reuse same Config/Bands/Detail sheet build as main export
    ws_config = wb.create_sheet("Config")
    ws_config.cell(row=1, column=1, value="Key")
    ws_config.cell(row=1, column=2, value="Value")
    ws_config.cell(row=2, column=1, value="MODE")
    ws_config.cell(row=2, column=2, value=mode_upper)
    ws_config.cell(row=3, column=1, value="DYN_X")
    ws_config.cell(row=3, column=2, value=float(dyn_x))

    from openpyxl.workbook.defined_name import DefinedName
    try:
        wb.defined_names.append(DefinedName(name="CFG_MODE", attr_text="Config!$B$2"))
        wb.defined_names.append(DefinedName(name="CFG_DYNX", attr_text="Config!$B$3"))
    except Exception:
        pass

    ws_bands = wb.create_sheet("Bands")
    bands_cols = ["direction","lower_pct","upper_pct","rate_type","rate_value","rate_slope","loss_zone"]
    for col_idx, col_name in enumerate(bands_cols, start=1):
        ws_bands.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(bands_table.itertuples(index=False), start=2):
        ws_bands.cell(row=row_idx, column=1, value=str(getattr(row, "direction", "")))
        ws_bands.cell(row=row_idx, column=2, value=float(getattr(row, "lower_pct", 0)))
        ws_bands.cell(row=row_idx, column=3, value=float(getattr(row, "upper_pct", 0)))
        ws_bands.cell(row=row_idx, column=4, value=str(getattr(row, "rate_type", "")))
        ws_bands.cell(row=row_idx, column=5, value=float(getattr(row, "rate_value", 0)))
        ws_bands.cell(row=row_idx, column=6, value=float(getattr(row, "rate_slope", 0)))
        ws_bands.cell(row=row_idx, column=7, value=bool(getattr(row, "loss_zone", False)))

    n_bands = max(len(bands_table), 1)
    bands_end_row = 1 + n_bands
    bands_end_ref = 200 if n_bands < 199 else bands_end_row

    try:
        wb.defined_names.append(DefinedName(name="Bands_Dir", attr_text=f"Bands!$A$2:$A${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_Lower", attr_text=f"Bands!$B$2:$B${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_Upper", attr_text=f"Bands!$C$2:$C${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_RateType", attr_text=f"Bands!$D$2:$D${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_RateVal", attr_text=f"Bands!$E$2:$E${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_RateSlope", attr_text=f"Bands!$F$2:$F${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_LossZone", attr_text=f"Bands!$G$2:$G${bands_end_ref}"))
    except Exception:
        pass

    ws_detail = wb.create_sheet("Detail")
    detail_headers = {
        'A': 'Region',
        'B': 'Plant Name',
        'C': 'Date',
        'D': 'Block',
        'E': 'From Time',
        'F': 'To Time',
        'G': 'Schedule Power (MW)',
        'H': 'AvC (MW)',
        'I': 'Injected Power (MW)',
        'J': 'PPA or MCP',
        'K': 'Error %',
        'L': 'Absolute error %',
        'M': 'Direction',
        'N': 'Deviation (MW)',
        'U': 'Revenue as per Generation (INR)',
        'V': 'Scheduled Revenue (INR)',
        'AE': '_basis'
    }

    for col_letter, header in detail_headers.items():
        col_idx = column_index_from_string(col_letter)
        ws_detail.cell(row=1, column=col_idx, value=header)

    ws_detail.column_dimensions['AE'].hidden = True

    n_rows = len(detail_for_excel)
    for row_idx, row_data in enumerate(detail_for_excel.itertuples(index=False), start=2):
        calc_row = detail_calculated_df.iloc[row_idx - 2] if row_idx - 2 < len(detail_calculated_df) else {}

        ws_detail.cell(row=row_idx, column=1, value=getattr(row_data, "region", ""))
        ws_detail.cell(row=row_idx, column=2, value=getattr(row_data, "plant_name", ""))
        ws_detail.cell(row=row_idx, column=3, value=getattr(row_data, "date", ""))
        ws_detail.cell(row=row_idx, column=4, value=getattr(row_data, "time_block", ""))
        ws_detail.cell(row=row_idx, column=5, value=getattr(row_data, "from_time", ""))
        ws_detail.cell(row=row_idx, column=6, value=getattr(row_data, "to_time", ""))
        ws_detail.cell(row=row_idx, column=7, value=float(getattr(row_data, "Scheduled_MW", 0)))
        ws_detail.cell(row=row_idx, column=8, value=float(getattr(row_data, "AvC_MW", 0)))
        ws_detail.cell(row=row_idx, column=9, value=float(getattr(row_data, "Actual_MW", 0)))
        ws_detail.cell(row=row_idx, column=10, value=float(getattr(row_data, "PPA", 0)))

        avc = float(getattr(row_data, "AvC_MW", 0))
        sch = float(getattr(row_data, "Scheduled_MW", 0))
        act = float(getattr(row_data, "Actual_MW", 0))
        ppa = float(getattr(row_data, "PPA", 0))

        basis = denominator_and_basis(avc, sch, mode_upper, dyn_x)
        ws_detail.cell(row=row_idx, column=31, value=float(basis))

        error_pct = calc_row.get("error_pct", 0.0)
        ws_detail.cell(row=row_idx, column=11, value=float(error_pct))

        abs_err = calc_row.get("abs_err", 0.0)
        ws_detail.cell(row=row_idx, column=12, value=float(abs_err))

        direction = calc_row.get("direction", "")
        ws_detail.cell(row=row_idx, column=13, value=str(direction))

        deviation = act - sch
        ws_detail.cell(row=row_idx, column=14, value=float(deviation))

        rev_act = calc_row.get("Revenue_as_per_generation", 0.0)
        ws_detail.cell(row=row_idx, column=21, value=float(rev_act))

        rev_sch = calc_row.get("Scheduled_Revenue_as_per_generation", 0.0)
        ws_detail.cell(row=row_idx, column=22, value=float(rev_sch))

    # Reuse final summary section pattern if desired, or keep as-is
    wb.save(buf)
    buf.seek(0)
    return dcc.send_bytes(lambda x: x.write(buf.read()), "DSM_calculation_aggregated.xlsx")

if __name__ == "__main__":
    host = os.getenv("DSM_HOST", "127.0.0.1")
    port = int(os.getenv("DSM_PORT", "8050"))
    debug = os.getenv("DSM_DEBUG", "true").lower() in ("1", "true", "yes")
    app.run(
        debug=debug,
        host=host,
        port=port,
        use_reloader=False,
        threaded=True,
    )